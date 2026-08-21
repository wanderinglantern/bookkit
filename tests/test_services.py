from __future__ import annotations

import sqlite3
from datetime import date, timedelta

import pytest

from bookkit import seed
from bookkit.money import weighted_cents
from bookkit.repo import (
    contacts,
    events,
    interactions,
    opportunities,
    orgs,
    placements,
    rfi,
    submissions,
    tasks,
)
from bookkit.repo import projects as projects_repo
from bookkit.services import book, capture, hit_rate, pipeline, renewals, sla, staleness, undo
from bookkit.services.export_open_items import compose

_WITHDRAWN_SCOPE_NOTE = (
    "This report lists items owned by you or by us on your account. "
    "Internal administrative items are not included."
)
"""The standing scope line, REMOVED from the workbook 2026-08-21 (Grant).

Kept here as a literal so the tests below can assert it never returns. The
WITHHOLDING it described is unchanged and still asserted — what went is the
sentence explaining the document to a reader who had not asked."""

TODAY = date(2026, 8, 11)


@pytest.fixture
def seeded(conn: sqlite3.Connection) -> sqlite3.Connection:
    counts = seed.seed(conn, today=TODAY)
    assert counts["orgs"] == 35
    return conn


def test_seed_is_realistic(seeded) -> None:
    assert len(orgs.list_orgs(seeded, kind="client")) == 20
    assert len(orgs.list_orgs(seeded, kind="market")) == 15


def test_renewal_buckets(seeded) -> None:
    buckets = renewals.bucketed(seeded, TODAY)
    assert set(buckets) == {"overdue", "0-30", "31-60", "61-90", "91-120"}
    items = renewals.upcoming(seeded, TODAY)
    assert items, "seed should have renewals inside 120 days"
    for item in items:
        if item.bucket == "overdue":  # expired, never renewed — stays visible
            assert item.days_remaining < 0
            continue
        assert 0 <= item.days_remaining <= 120
        lo, hi = (int(x) for x in item.bucket.split("-"))
        assert lo <= item.days_remaining <= hi
    upcoming_only = [i.days_remaining for i in items if i.days_remaining >= 0]
    assert upcoming_only == sorted(upcoming_only)


def test_the_attention_window_is_exactly_120_days(seeded) -> None:
    """`0 <= days_remaining <= 120` in the test above is CLAMPED by the query
    under test, and the seed's largest in-window value is 104 — so cutting
    `upcoming(days=)` from 120 to 35 left the whole suite green while the
    attention window silently shrank (2026-08-18). Pin BOTH edges with
    placements planted on them: day 120 must be in, day 121 must be out.

    Unlinked on purpose — no program_path means `_renewal_on` falls back to
    period_to, so the date under test is the one written here."""
    org = orgs.create(seeded, name="Horizon Edge Co", kind="client", status="active")
    on_the_edge = placements.create(
        seeded, org.id, "Edge Day 120",
        (TODAY - timedelta(days=245)).isoformat(),
        (TODAY + timedelta(days=120)).isoformat(),
        status="bound",
    )
    just_past = placements.create(
        seeded, org.id, "Edge Day 121",
        (TODAY - timedelta(days=244)).isoformat(),
        (TODAY + timedelta(days=121)).isoformat(),
        status="bound",
    )

    found = {i.placement.id: i for i in renewals.upcoming(seeded, TODAY)}
    assert on_the_edge.id in found, "day 120 fell outside the 120-day window"
    assert found[on_the_edge.id].days_remaining == 120
    assert found[on_the_edge.id].bucket == "91-120"
    assert just_past.id not in found, "day 121 is past the window and must not show"


def test_staleness_weighted_by_premium(seeded) -> None:
    """The seed leaves only two stale accounts and ordering them by days_stale
    gives the SAME order as days_stale x premium, so dropping the premium
    factor entirely (`days_stale * 1`) was green (2026-08-18). The fixture
    below plants the inversion the name promises: a big account stale for a
    short while must outrank a tiny one stale for far longer."""
    whale = orgs.create(seeded, name="Whale Holdings", kind="client", status="active")
    minnow = orgs.create(seeded, name="Minnow LLC", kind="client", status="active")
    pauper = orgs.create(seeded, name="Pauper Co", kind="client", status="active")
    plan = [
        # org, days since last contact, current bound premium (cents)
        (whale, 70, 5_000_000_00),
        (minnow, 400, 1_00),
        (pauper, 300, 0),  # no bound placement at all
    ]
    for org, gap, premium in plan:
        interactions.log(
            seeded, org.id, "call", "checked in",
            (TODAY - timedelta(days=gap)).isoformat(),
        )
        if premium:
            placements.create(
                seeded, org.id, f"{org.name} Program",
                (TODAY - timedelta(days=200)).isoformat(),
                (TODAY + timedelta(days=165)).isoformat(),
                status="bound", total_premium=premium,
            )

    stale = staleness.stale_accounts(seeded, TODAY, threshold_days=60)
    assert stale, "seed deliberately leaves accounts stale"
    weights = [s.weight for s in stale]
    assert weights == sorted(weights, reverse=True)
    for account in stale:
        assert account.days_stale > 60
        assert account.org.status == "active"

    by_id = {s.org.id: s for s in stale}
    assert {whale.id, minnow.id, pauper.id} <= set(by_id)
    # the inversion: days_stale alone would put minnow first
    assert by_id[minnow.id].days_stale > by_id[whale.id].days_stale
    order = [s.org.id for s in stale]
    assert order.index(whale.id) < order.index(minnow.id), (
        "premium is not weighting the ranking — a 5m account 70 days cold "
        "ranks below a $1 account"
    )
    assert by_id[whale.id].premium == 5_000_000_00
    # a premium-less account still gets a rank, not a zero
    assert by_id[pauper.id].premium == 0
    assert by_id[pauper.id].weight == by_id[pauper.id].days_stale
    assert order.index(minnow.id) < order.index(pauper.id)


def test_stage_transitions_enforced(seeded) -> None:
    org = orgs.list_orgs(seeded, kind="client")[0]
    opp = opportunities.create(seeded, org.id, "Test opp", target_premium=100_000_00)
    with pytest.raises(pipeline.StageError):
        pipeline.move_stage(seeded, opp.id, "quoted")  # can't skip gates
    pipeline.move_stage(seeded, opp.id, "qualified")
    moved = pipeline.move_stage(seeded, opp.id, "lost", loss_reason="budget")
    assert moved.stage == "lost" and moved.outcome == "lost" and moved.probability_pct == 0
    with pytest.raises(pipeline.StageError):
        pipeline.move_stage(seeded, opp.id, "won")  # closed is closed
    log = pipeline.stage_history(seeded, opp.id)
    assert [(e.old_value, e.new_value) for e in log] == [
        ("qualified", "lost"), ("identified", "qualified"),
    ]


def test_pipeline_metrics(seeded) -> None:
    rows = pipeline.metrics(seeded)
    assert [r.stage for r in rows] == list(pipeline.STAGES)
    by_stage = {r.stage: r for r in rows}
    total = sum(r.count for r in rows)
    assert total == len(opportunities.by_stage(seeded))
    for r in rows:
        # `weighted <= total` CANNOT FAIL — weighted is total * pct // 100 with
        # pct <= 100 — so halving every stage's weighted pipeline was green
        # (2026-08-18). Pin the figure to the opportunities it is drawn from,
        # via the separately unit-tested money helper.
        stage_opps = opportunities.by_stage(seeded, r.stage)
        assert r.weighted_cents == sum(
            weighted_cents(o.target_premium or 0, o.probability_pct)
            for o in stage_opps
        ), f"{r.stage}: not the probability-weighted sum of its own opportunities"
        assert r.total_cents == sum(o.target_premium or 0 for o in stage_opps)
    # anti-vacuity: the seed must actually exercise weighting, or the loop
    # above holds just as well for an empty or unweighted pipeline
    assert sum(r.weighted_cents for r in rows) > 0
    assert any(0 < r.weighted_cents < r.total_cents for r in rows)
    assert by_stage["qualified"].avg_days_in_stage is not None  # seed advanced through it


def test_sla(seeded) -> None:
    overdue = sla.past_sla(seeded, TODAY, sla_days=10)
    assert overdue, "seed leaves submissions out past SLA"
    assert all(o.days_out >= 10 for o in overdue)
    assert all(o.submission.status == "out" for o in overdue)
    days = [o.days_out for o in overdue]
    assert days == sorted(days, reverse=True)


def test_hit_rate(seeded) -> None:
    rows = hit_rate.by_market(seeded)
    assert rows
    for r in rows:
        assert 0 <= r.quote_rate <= 1
        assert r.bound <= r.quoted <= r.sent
    total = hit_rate.overall(seeded)
    assert total.sent == sum(r.sent for r in rows)


def test_book_summary(seeded) -> None:
    summary = book.summary(seeded)
    assert summary.accounts > 0
    assert summary.total_premium > 0
    assert 0 < summary.total_commission < summary.total_premium
    assert summary.by_program


def test_undo_field_change(seeded) -> None:
    """`u` reverts a TUI write. It is batch-granular now, so the write has to
    be made the way the TUI makes it — inside a batch."""
    from bookkit.services import batches as batches_svc

    org = orgs.list_orgs(seeded, kind="client")[0]
    with batches_svc.open_batch(
        seeded, source="tui", tool="edit_account", summary="set status"
    ):
        orgs.update(seeded, org.id, status="dormant")

    result = undo.undo_last(seeded)
    assert result is not None and result.applied
    assert orgs.get(seeded, org.id).status == org.status


def test_undo_soft_delete(seeded) -> None:
    from bookkit.services import batches as batches_svc

    org = orgs.list_orgs(seeded, kind="client")[0]
    task = tasks.create(seeded, "Ephemeral", org_id=org.id)
    with batches_svc.open_batch(
        seeded, source="tui", tool="task_delete", summary="deleted a task"
    ):
        tasks.delete(seeded, task.id)

    result = undo.undo_last(seeded)
    assert result is not None and result.applied
    assert tasks.get(seeded, task.id).status == "open"


def test_undo_ignores_a_write_this_app_did_not_make(seeded) -> None:
    """Unbatched writes are sync projections and imports. `u` used to revert
    placement.synced_at on a freshly opened app; now it finds nothing."""
    org = orgs.list_orgs(seeded, kind="client")[0]
    orgs.update(seeded, org.id, status="dormant")      # no batch = not ours

    assert undo.undo_last(seeded) is None
    assert orgs.get(seeded, org.id).status == "dormant"


def test_capture_suggests_task() -> None:
    got = capture.suggest_task("Spoke to Alice. Follow up Tuesday about loss runs.", TODAY)
    assert got is not None
    assert got.due_on > TODAY and got.due_on.weekday() == 1
    got = capture.suggest_task("call him next week re: quote", TODAY)
    assert got is not None and got.due_on > TODAY
    assert capture.suggest_task("Bound the renewal, all done.", TODAY) is None


def test_event_log_answers_when_did_this_move(seeded) -> None:
    org = orgs.list_orgs(seeded, kind="client")[0]
    placement = placements.for_org(seeded, org.id)[0]
    placements.update(seeded, placement.id, period_to="2027-03-01", note="insured pushed expiry")
    history = events.field_history(seeded, "placement", placement.id, "period_to")
    assert history[0].new_value == "2027-03-01"
    assert history[0].note == "insured pushed expiry"


def test_renewal_date_survives_timezones(seeded, monkeypatch) -> None:
    """DATE columns are text and never converted: reading under a different
    TZ yields the same date (§3.1)."""
    import os
    import time

    org = orgs.list_orgs(seeded, kind="client")[0]
    placement = placements.for_org(seeded, org.id)[0]
    before = placement.period_to
    monkeypatch.setenv("TZ", "Pacific/Auckland")
    time.tzset()
    try:
        assert placements.get(seeded, placement.id).period_to == before
        assert date.fromisoformat(before) == date.fromisoformat(
            placements.get(seeded, placement.id).period_to
        )
    finally:
        os.environ.pop("TZ", None)
        time.tzset()


class TestRenewalsScanAllPrograms:
    def test_overdue_unrenewed_program_stays_on_the_radar(self, conn) -> None:
        org = orgs.create(conn, kind="client", name="Multi Program Co")
        placements.create(conn, org.id, "Property", "2025-11-01", "2026-11-01",
                          status="bound")
        expired = placements.create(conn, org.id, "Casualty", "2025-06-01",
                                    "2026-06-01", status="bound")
        items = renewals.upcoming(conn, TODAY, days=120)
        by_id = {item.placement.id: item for item in items}
        assert expired.id in by_id  # expired 2 months ago, never renewed
        assert by_id[expired.id].bucket == "overdue"
        assert by_id[expired.id].days_remaining < 0

    def test_renewed_and_lapsed_programs_do_not_nag(self, conn) -> None:
        org = orgs.create(conn, kind="client", name="Tidy Co")
        placements.create(conn, org.id, "2025 Casualty Program", "2025-06-01",
                          "2026-06-01", status="bound")
        placements.create(conn, org.id, "2026 Casualty Program", "2026-06-01",
                          "2027-06-01", status="bound")  # renew-at-birth successor
        placements.create(conn, org.id, "Old Marine", "2024-01-01", "2025-01-01",
                          status="lapsed")  # deliberately let go
        items = renewals.upcoming(conn, TODAY, days=120)
        assert all(item.bucket != "overdue" for item in items)

    def test_next_for_org_prefers_overdue_across_programs(self, conn) -> None:
        org = orgs.create(conn, kind="client", name="Two Towers Co")
        placements.create(conn, org.id, "Property", "2025-10-01", "2026-10-01",
                          status="bound")
        placements.create(conn, org.id, "Casualty", "2025-06-01", "2026-06-01",
                          status="bound")
        nxt = renewals.next_for_org(conn, org.id, TODAY)
        assert nxt is not None
        assert nxt.placement.program_name == "Casualty"  # overdue beats upcoming
        assert nxt.days_remaining < 0


def test_renewal_items_carry_line_labels(tmp_path) -> None:
    from bookkit import db as db_mod
    from bookkit import sync as sync_mod

    connection = db_mod.connect(tmp_path / "t.db")
    try:
        seed.seed(connection, today=TODAY, programs_dir=tmp_path / "programs")
        sync_mod.project_all(connection, [tmp_path / "programs"])
        items = renewals.upcoming(connection, TODAY)
        linked = [i for i in items if i.placement.program_path]
        assert linked, "seed has file-linked placements in the window"
        assert all(item.lines for item in linked)  # e.g. "GL, AL, EL"
        unlinked = [i for i in items if not i.placement.program_path]
        assert all(item.lines == "" for item in unlinked)
    finally:
        connection.close()


def test_line_of_cover_drives_renewal_clock(conn, tmp_path) -> None:
    """A line whose layer expires before the program period must surface on
    ITS clock: the IM policy dying in 30 days pulls the placement into the
    renewal radar even though the program end is 180 days out."""
    from datetime import timedelta

    from towerkit.model import Layer, Line, Participant, Period, Program, dump_program
    from towerkit.model import Placement as TkPlacement

    org = orgs.create(conn, name="Line Clock Co", kind="client", status="active")
    start = TODAY - timedelta(days=185)
    end = TODAY + timedelta(days=180)
    p = placements.create(
        conn, org.id, "2026 Package Program",
        start.isoformat(), end.isoformat(), status="bound",
    )
    im_end = TODAY + timedelta(days=30)
    program = Program(
        insured=org.name, program="Package Program", placement=TkPlacement.BOUND,
        period=Period(start=start, end=end),
        lines=[
            Line(id="pr", name="Property", abbr="PR"),
            Line(id="im", name="Inland Marine", abbr="IM"),
        ],
        layers=[
            Layer(
                id="pr1", name="Primary Property", applies_to=["pr"],
                attach=0, limit=1_000_000,
                participants=[Participant(carrier="Zurich", share_bps=10_000)],
            ),
            Layer(
                id="im1", name="Primary IM", applies_to=["im"],
                period=Period(start=start, end=im_end),
                attach=0, limit=1_000_000,
                participants=[Participant(carrier="CNA", share_bps=10_000)],
            ),
        ],
    )
    path = tmp_path / "line-clock.json"
    dump_program(program, path)
    placements.update(conn, p.id, program_path=str(path))

    items = [i for i in renewals.upcoming(conn, TODAY) if i.org.id == org.id]
    assert len(items) == 1, "the IM line's clock must pull the placement in"
    item = items[0]
    assert item.renewal_on == im_end.isoformat()
    assert item.days_remaining == 30
    assert item.bucket == "0-30"
    assert item.line_ends[0] == ("IM", im_end.isoformat())
    assert dict(item.line_ends)["PR"] == end.isoformat()

    nxt = renewals.next_for_org(conn, org.id, TODAY)
    assert nxt is not None and nxt.renewal_on == im_end.isoformat()


def test_flatten_markdown_strips_marks_keeps_bullets():
    from bookkit.services.export_open_items import flatten_markdown

    text = "## Head\n- **bold** item\n* second [link](http://x)\n`code`"
    assert flatten_markdown(text) == "Head\n- bold item\n- second link\ncode"


def test_compose_groups_by_program_project_and_general(conn):
    # build: client with an org-level task, a placement-attached task,
    # an outstanding submission on that placement, and a project need
    client = orgs.create(conn, kind="client", name="Acme", status="active", owner="grant")
    market = orgs.create(conn, kind="market", name="Zurich", status="active")

    org_task = tasks.create(
        conn, "Chase updated loss runs", org_id=client.id,
        description="waiting on brief line from the client",
        detail="**Please** call the broker",
    )

    p = placements.create(
        conn, client.id, "Acme Property 25-26", "2025-10-01", "2026-10-01"
    )
    placement_task = tasks.create(conn, "Confirm bound terms", placement_id=p.id)
    sub = submissions.create(conn, market.id, "2026-07-01", placement_id=p.id)

    project = projects_repo.create_project(conn, client.id, "Warehouse Expansion")
    need = projects_repo.add_need(conn, project.id, "Builder's Risk", "2026-09-01")

    today = date.today()
    sections = compose(conn, client.id, today)
    labels = [s.label for s in sections]
    assert labels[0].startswith("General")
    assert any(lbl.startswith("Acme Property") for lbl in labels)
    assert any(lbl.startswith("Project — ") for lbl in labels)
    task_row = sections[0].rows[0]
    assert task_row.kind == "Task"
    assert task_row.description == "waiting on brief line from the client"
    assert task_row.detail == "Please call the broker"  # markdown flattened
    # per-client rows carry a ref — the exact id an MCP caller reads and
    # later hands to task_complete (task_complete has no ref of its own to
    # offer from the export path otherwise)
    assert task_row.ref == org_task.id

    # a placement-only task (org_id NULL, placement_id set — legal per
    # repo/tasks.py) must still reach the workbook, in the placement's section
    placement_section = next(s for s in sections if s.label.startswith("Acme Property"))
    assert any(r.item == "Confirm bound terms" for r in placement_section.rows)
    placement_task_row = next(
        r for r in placement_section.rows if r.item == "Confirm bound terms")
    assert placement_task_row.ref == placement_task.id
    submission_row = next(r for r in placement_section.rows if r.kind == "Submission")
    assert submission_row.ref == sub.id

    # need status is client-facing prose, not raw vocab ("identified", not
    # the DB's underscored form)
    project_section = next(s for s in sections if s.label.startswith("Project — "))
    assert project_section.rows[0].status == "Identified"
    assert project_section.rows[0].ref == need.id


def test_status_label_prettifies_underscored_vocab():
    from bookkit.services.export_open_items import _status_label

    assert _status_label("identified") == "Identified"
    assert _status_label("not_needed") == "Not needed"


def test_compose_empty_book_returns_no_sections(conn):
    org = orgs.create(conn, kind="client", name="Empty Co", status="active", owner="grant")
    assert compose(conn, org.id, date(2026, 8, 12)) == []


def test_compose_sections_org_tasks_by_category(conn):
    org = orgs.create(conn, name="Cat Co", kind="client")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    tasks.create(conn, "renew AL", org_id=org.id, category="Renewal")
    tasks.create(conn, "send COI", org_id=org.id, category="Certificates")
    tasks.create(conn, "misc", org_id=org.id)
    labels = [s.label for s in compose(conn, org.id, date(2026, 8, 12))]
    assert labels == ["Certificates — Cat Co", "Renewal — Cat Co", "General — Cat Co"]


def test_compose_categories_bucket_case_insensitively(conn):
    # "renewal" and "Renewal" are the same section, case-insensitively
    # (repo/vocab.py's _dedupe rule); first-seen spelling wins the label
    org = orgs.create(conn, name="Case Co", kind="client")
    tasks.create(conn, "renew GL", org_id=org.id, category="renewal")
    tasks.create(conn, "renew AL", org_id=org.id, category="Renewal")
    sections = compose(conn, org.id, date(2026, 8, 12))
    renewal_sections = [s for s in sections if s.label.startswith(("renewal", "Renewal"))]
    assert len(renewal_sections) == 1
    assert renewal_sections[0].label == "renewal — Case Co"  # first-seen spelling
    assert len(renewal_sections[0].rows) == 2


# --- C7: overdue leads ------------------------------------------------------
#
# Alphabetical is the one ordering that serves nobody. Sections sort by their
# most-overdue member and rows sort inside their section; nothing is
# regrouped, so no item can appear twice.


TODAY = date(2026, 8, 18)


def test_the_overdue_section_leads_the_one_that_sorts_first_alphabetically(conn):
    """The reviewer's own copy: their only overdue item, 12 days late, sat
    below a certificate not due for three days, because C came before R."""
    org = orgs.create(conn, name="Order Co", kind="client")
    tasks.create(conn, "issue COI", org_id=org.id, category="Certificates",
                 due_on="2026-08-21")            # due in 3 days
    tasks.create(conn, "renewal submission", org_id=org.id, category="Renewal",
                 due_on="2026-08-06")            # 12 days overdue

    labels = [s.label for s in compose(conn, org.id, TODAY)]
    assert labels == ["Renewal — Order Co", "Certificates — Order Co"]


def test_the_most_overdue_section_leads_the_overdue_ones(conn):
    org = orgs.create(conn, name="Deep Co", kind="client")
    tasks.create(conn, "a", org_id=org.id, category="Audit", due_on="2026-08-15")
    tasks.create(conn, "b", org_id=org.id, category="Binder", due_on="2026-05-01")
    tasks.create(conn, "c", org_id=org.id, category="Claims", due_on="2026-07-01")

    labels = [s.label for s in compose(conn, org.id, TODAY)]
    assert labels == ["Binder — Deep Co", "Claims — Deep Co", "Audit — Deep Co"]


def test_overdue_rows_lead_inside_their_own_section(conn):
    """The same defect at smaller scale: a section that opens on a row due
    next month while holding one two weeks late.

    End-to-end this currently passes with or without the row-level sort,
    because repo/tasks.open_tasks_for_client and repo/projects.needs_for_project
    both already ORDER BY the due date and submissions (undated) are appended
    last. That redundancy is the point of the two direct _overdue_first tests
    below: they are what actually holds the row rule, so a repo ORDER BY
    changing cannot reintroduce the defect unnoticed."""
    org = orgs.create(conn, name="Inside Co", kind="client")
    tasks.create(conn, "later", org_id=org.id, category="Renewal", due_on="2026-09-30")
    tasks.create(conn, "undated", org_id=org.id, category="Renewal")
    tasks.create(conn, "late", org_id=org.id, category="Renewal", due_on="2026-08-04")
    tasks.create(conn, "latest", org_id=org.id, category="Renewal", due_on="2026-07-04")

    section = compose(conn, org.id, TODAY)[0]
    assert [r.item for r in section.rows] == ["latest", "late", "later", "undated"]


def test_a_placement_section_outranks_a_category_section_when_it_is_overdue(conn):
    """Sheet 1 carries TWO kinds of section. Both sort in one ordering, so an
    overdue placement item leads a merely-upcoming category item and vice
    versa — a rule applied to only one kind would leave the other alphabetical."""
    org = orgs.create(conn, name="Both Co", kind="client")
    p = placements.create(conn, org.id, "Both Co Property 25-26",
                          "2025-10-01", "2026-10-01")
    tasks.create(conn, "chase subjectivities", placement_id=p.id, due_on="2026-08-01")
    tasks.create(conn, "issue COI", org_id=org.id, category="Certificates",
                 due_on="2026-08-21")

    labels = [s.label for s in compose(conn, org.id, TODAY)]
    assert labels == ["Both Co Property 25-26", "Certificates — Both Co"]


def test_an_overdue_project_need_leads_although_its_status_is_not_overdue(conn):
    """`_overdue_on` reads the DUE date, not the status word. Only task rows
    ever say "Overdue"; a project need carries its own vocabulary status and
    reaches sheet 1 through a different branch entirely."""
    org = orgs.create(conn, name="Need Co", kind="client")
    project = projects_repo.create_project(conn, org.id, "Warehouse")
    projects_repo.add_need(conn, project.id, "Builder's Risk", "2026-06-01")
    tasks.create(conn, "issue COI", org_id=org.id, category="Certificates",
                 due_on="2026-08-21")

    sections = compose(conn, org.id, TODAY)
    assert [s.label for s in sections] == [
        "Project — Warehouse", "Certificates — Need Co"]
    assert sections[0].rows[0].status == "Identified"  # not the word "Overdue"


def test_nothing_appears_twice_when_overdue_leads(conn):
    """Reordering, never regrouping: every ref still appears exactly once
    across the whole sheet, and the row count is unchanged."""
    org = orgs.create(conn, name="Once Co", kind="client")
    p = placements.create(conn, org.id, "Once Co Package", "2025-10-01", "2026-10-01")
    project = projects_repo.create_project(conn, org.id, "Fitout")
    tasks.create(conn, "late cat", org_id=org.id, category="Renewal",
                 due_on="2026-07-01")
    tasks.create(conn, "soon cat", org_id=org.id, category="Certificates",
                 due_on="2026-09-01")
    tasks.create(conn, "loose", org_id=org.id)
    tasks.create(conn, "late placement", placement_id=p.id, due_on="2026-06-01")
    projects_repo.add_need(conn, project.id, "Builder's Risk", "2026-05-01")

    refs = [r.ref for s in compose(conn, org.id, TODAY) for r in s.rows]
    assert len(refs) == len(set(refs)) == 5


def test_with_nothing_overdue_the_composition_order_survives(conn):
    """The stable-sort tiebreak, pinned: categories alphabetical, then
    General, then placements, then projects — unchanged from before C7."""
    org = orgs.create(conn, name="Calm Co", kind="client")
    p = placements.create(conn, org.id, "Calm Co Package", "2025-10-01", "2026-10-01")
    project = projects_repo.create_project(conn, org.id, "Fitout")
    tasks.create(conn, "renew", org_id=org.id, category="Renewal", due_on="2026-09-05")
    tasks.create(conn, "cert", org_id=org.id, category="Certificates")
    tasks.create(conn, "loose", org_id=org.id)
    tasks.create(conn, "bind", placement_id=p.id, due_on="2026-09-09")
    projects_repo.add_need(conn, project.id, "Builder's Risk", "2026-12-01")

    assert [s.label for s in compose(conn, org.id, TODAY)] == [
        "Certificates — Calm Co", "Renewal — Calm Co", "General — Calm Co",
        "Calm Co Package", "Project — Fitout",
    ]


def _row(item: str, due: str) -> object:
    from bookkit.services.export_open_items import ExportRow

    return ExportRow(item=item, description="", detail="", kind="Task",
                     due=due, status="", ref=item)


def test_overdue_first_sorts_rows_inside_a_section_directly():
    """Directly on the ordering function, with rows handed to it out of
    order — the compose path cannot currently produce that (see above), so
    this is the test that makes the row-level sort load-bearing."""
    from bookkit.services.export_open_items import ExportSection, _overdue_first

    section = ExportSection("S", (
        _row("soon", "2026-09-30"), _row("undated", ""),
        _row("late", "2026-08-04"), _row("latest", "2026-07-04"),
    ))
    out = _overdue_first([section], TODAY)
    assert [r.item for r in out[0].rows] == ["latest", "late", "soon", "undated"]


def test_a_section_is_ranked_by_its_most_overdue_member_not_its_least():
    """"Most overdue" is the EARLIEST past-due date in the section. Ranking on
    the latest one instead buries a section holding something six months late
    behind one holding a single item two days late."""
    from bookkit.services.export_open_items import ExportSection, _overdue_first

    deep = ExportSection("Deep", (
        _row("barely", "2026-08-16"), _row("ancient", "2026-01-01")))
    middling = ExportSection("Middling", (_row("mid", "2026-06-01"),))

    assert [s.label for s in _overdue_first([middling, deep], TODAY)] == [
        "Deep", "Middling"]


def test_a_task_due_today_is_not_overdue(conn):
    """The same boundary C3 draws: today is not late."""
    org = orgs.create(conn, name="Edge Co", kind="client")
    tasks.create(conn, "due today", org_id=org.id, category="Zeta",
                 due_on=TODAY.isoformat())
    tasks.create(conn, "no date", org_id=org.id, category="Alpha")

    assert [s.label for s in compose(conn, org.id, TODAY)] == [
        "Alpha — Edge Co", "Zeta — Edge Co"]


# --- the Internal category: never leaves the building ------------------------
#
# The filter sits on the task list in compose(), before the split into
# category sections and placement sections, so an Internal task cannot ride
# out to the client through the placement half (which ignores category
# entirely). These tests pin both halves, the match rule, and the count the
# exporter reports back.


def test_compose_omits_the_internal_category_section(conn):
    org = orgs.create(conn, name="Internal Co", kind="client")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    tasks.create(conn, "chase our own file note", org_id=org.id, category="Internal")
    sections = compose(conn, org.id, date(2026, 8, 12))
    assert [s.label for s in sections] == ["Renewal — Internal Co"]
    items = [r.item for s in sections for r in s.rows]
    assert "chase our own file note" not in items


def test_internal_match_ignores_case_and_surrounding_space(conn):
    org = orgs.create(conn, name="Space Co", kind="client")
    tasks.create(conn, "a", org_id=org.id, category=" internal ")
    tasks.create(conn, "b", org_id=org.id, category="INTERNAL")
    tasks.create(conn, "c", org_id=org.id, category="Internal")
    tasks.create(conn, "d", org_id=org.id, category="Renewal")
    sections = compose(conn, org.id, date(2026, 8, 12))
    assert [s.label for s in sections] == ["Renewal — Space Co"]


def test_internal_prefix_is_not_internal(conn):
    """D1: exact equality, not a prefix. "Internal Review" is a real
    client-facing broking category; excluding it would drop a task from the
    deliverable with no signal anywhere. THE ROW STILL SHIPS — that half is
    unchanged and is what this test guards.

    D1's other half — that a wrong inclusion is loud because "the client sees
    a section header naming it" — was overturned by C9: the loudness landed on
    the client, not on us, and a heading reading Internal is worse than the
    item beneath it. So the heading is now suppressed and the row is filed
    under General (see the C9 block below); the withholding rule itself is
    untouched."""
    org = orgs.create(conn, name="Prefix Co", kind="client")
    tasks.create(conn, "walk the client through the audit", org_id=org.id,
                 category="Internal Review")
    sections = compose(conn, org.id, date(2026, 8, 12))
    assert [s.label for s in sections] == ["General — Prefix Co"]
    assert sections[0].rows[0].item == "walk the client through the audit"


def test_internal_task_on_a_placement_is_withheld_too(conn):
    """Sheet 1 carries a section per PLACEMENT, built from the same task list
    with category IGNORED. A filter applied inside the category branch leaves
    that half untouched and the Internal task ships anyway."""
    org = orgs.create(conn, name="Placement Co", kind="client")
    p = placements.create(conn, org.id, "Placement Co Property 25-26",
                          "2025-10-01", "2026-10-01")
    tasks.create(conn, "Confirm bound terms", placement_id=p.id)
    tasks.create(conn, "our own reserve note", placement_id=p.id, category="Internal")
    # also reachable through a category section, so a category-branch-only
    # filter looks like it works while the placement half leaks
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")

    sections = compose(conn, org.id, date(2026, 8, 12))
    items = [r.item for s in sections for r in s.rows]
    assert "our own reserve note" not in items, "an Internal placement task shipped"
    assert "our own file note" not in items
    placement_section = next(
        s for s in sections if s.label.startswith("Placement Co Property"))
    assert [r.item for r in placement_section.rows] == ["Confirm bound terms"]


def test_all_internal_account_exports_the_no_open_items_sheet(conn, tmp_path):
    """The pinned empty-section answer: compose() never emits an empty
    section, so an account whose only open item is Internal composes to
    nothing and write() falls back to its placeholder row."""
    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Quiet Co", kind="client")
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    assert compose(conn, org.id, date(2026, 8, 12)) == []
    path = write(conn, org.id, tmp_path / "q.xlsx", date(2026, 8, 12))
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    # The body starts at A2 now that no scope line precedes it.
    assert ws["A2"].value == "No open items as of 2026-08-12"


def test_compose_can_be_asked_for_the_internal_rows(conn):
    """The exclusion is a default, not a law: MCP asks for them explicitly."""
    from bookkit.services.export_open_items import compose as compose_fn

    org = orgs.create(conn, name="Ask Co", kind="client")
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    sections = compose_fn(conn, org.id, date(2026, 8, 12), include_internal=True)
    rows = [r for s in sections for r in s.rows]
    assert [r.item for r in rows] == ["our own file note"]
    assert rows[0].internal is True


def test_export_rows_flag_internal_but_the_workbook_cannot_print_it(conn, tmp_path):
    """ExportRow.internal follows `ref`: carried for MCP, absent from the
    writer's explicit column tuple, so it can never reach the client."""
    from bookkit.services.export_open_items import compose as compose_fn
    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Flag Co", kind="client")
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    rows = {r.item: r.internal
            for s in compose_fn(conn, org.id, date(2026, 8, 12), include_internal=True)
            for r in s.rows}
    assert rows == {"our own file note": True, "renew GL": False}

    path = write(conn, org.id, tmp_path / "f.xlsx", date(2026, 8, 12))
    from openpyxl import load_workbook
    sheet = load_workbook(path).active
    values = [c.value for row in sheet.iter_rows() for c in row]
    assert "our own file note" not in values

    # `True not in values` cannot fail here: write() composes with the default,
    # so every surviving row already has internal=False and no True exists to
    # find. These two can. A bool in ANY cell is the flag having reached the
    # workbook, and the Status column is where a wrong column tuple would put
    # it — so pin what that column actually says.
    assert not any(isinstance(v, bool) for v in values), (
        "a bool reached a workbook cell — ExportRow.internal is in write()'s "
        "column tuple"
    )
    # header, the scope line (column A only), the "Renewal — Flag Co" section
    # label (column A only), the one surviving row. A wrong column tuple puts
    # a False here.
    status_column = [row[5].value for row in sheet.iter_rows()]
    assert status_column == ["Status", None, "Open"], status_column


def test_withheld_internal_lists_what_the_client_did_not_get(conn):
    from bookkit.services.export_open_items import withheld_internal

    org = orgs.create(conn, name="Held Co", kind="client")
    held = tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    tasks.create(conn, "misc", org_id=org.id)
    assert [t.id for t in withheld_internal(conn, org.id)] == [held.id]


def test_near_miss_internal_names_what_the_rule_did_not_catch(conn):
    """Exact equality is the rule, so "Internal Review" ships. The near-miss
    list is the positive signal that says so — an absence teaches nobody who
    has never seen the presence."""
    from bookkit.services.export_open_items import near_miss_internal

    org = orgs.create(conn, name="Near Co", kind="client")
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    review = tasks.create(conn, "audit support", org_id=org.id, category="Internal Review")
    lower = tasks.create(conn, "note", org_id=org.id, category=" internal note ")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    tasks.create(conn, "misc", org_id=org.id)
    assert {t.id for t in near_miss_internal(conn, org.id)} == {review.id, lower.id}


def test_withheld_note_says_only_what_was_withheld(conn):
    from bookkit.services.export_open_items import withheld_note

    org = orgs.create(conn, name="Held Only Co", kind="client")
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    note = withheld_note(conn, org.id)
    assert note == " — 1 internal task withheld"

    tasks.create(conn, "our own reserve note", org_id=org.id, category="internal")
    assert withheld_note(conn, org.id) == " — 2 internal tasks withheld"


def test_withheld_note_names_the_near_miss_that_was_exported(conn):
    """The ROADMAP's own stated worry — "a prefix match and an equality match
    behave very differently the first time someone types 'Internal Review'" —
    answered where both surfaces already read the same sentence."""
    from bookkit.services.export_open_items import withheld_note

    org = orgs.create(conn, name="Miss Co", kind="client")
    tasks.create(conn, "audit support", org_id=org.id, category="Internal Review")
    assert withheld_note(conn, org.id) == (
        ' — 1 task categorised "Internal Review" WAS exported '
        '(only the exact category "Internal" is withheld)'
    )

    # plural, and the same spelling twice names the category once
    tasks.create(conn, "more audit support", org_id=org.id, category="Internal Review")
    tasks.create(conn, "file note", org_id=org.id, category="internal note")
    assert withheld_note(conn, org.id) == (
        ' — 3 tasks categorised "internal note", "Internal Review" WERE exported '
        '(only the exact category "Internal" is withheld)'
    )


def test_withheld_note_says_both_when_both_happened(conn):
    from bookkit.services.export_open_items import withheld_note

    org = orgs.create(conn, name="Both Co", kind="client")
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    tasks.create(conn, "audit support", org_id=org.id, category="Internal Review")
    assert withheld_note(conn, org.id) == (
        ' — 1 internal task withheld; 1 task categorised "Internal Review" '
        'WAS exported (only the exact category "Internal" is withheld)'
    )


def test_withheld_note_is_empty_when_there_is_nothing_to_say(conn):
    """Silence is still correct for the ordinary account: nothing internal,
    nothing that reads internal. It is only silence about a NEAR MISS that
    taught nobody anything."""
    from bookkit.services.export_open_items import withheld_note

    org = orgs.create(conn, name="Quiet Co", kind="client")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    tasks.create(conn, "misc", org_id=org.id)
    assert withheld_note(conn, org.id) == ""


def test_is_internal_category_predicate():
    from bookkit.models import INTERNAL_CATEGORY, is_internal_category

    assert INTERNAL_CATEGORY == "Internal"
    assert is_internal_category("Internal")
    assert is_internal_category(" internal ")
    assert is_internal_category("INTERNAL")
    assert not is_internal_category("Internal Review")
    assert not is_internal_category("Renewal")
    assert not is_internal_category("")
    assert not is_internal_category(None)


# --- C9: no section heading in the client's copy may read "Internal" ---------


def _rendered_sections(ws) -> list[tuple[str, list[str]]]:
    """The sheet read back the way a client reads it: (section label, items).

    A section label is rendered by towerkit's band_row, which writes column A
    and leaves the rest of the row empty; a body row fills every column. So
    the shape of the row IS the distinction, and reading it back this way is
    what makes the assertion about what the CLIENT sees rather than about what
    compose() returned. Row 1 is the column header and is skipped. A body row
    before any label row raises here on purpose — that is the headerless
    section this design rejected, and it must not appear silently."""
    out: list[tuple[str, list[str]]] = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        if all(v is None for v in values[1:]):
            out.append((str(values[0]), []))
        else:
            out[-1][1].append(str(values[0]))
    return out


def _banner_fixture(conn):
    """One account carrying every shape the rule has to separate."""
    org = orgs.create(conn, name="Banner Co", kind="client")
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    tasks.create(conn, "audit support", org_id=org.id, category="Internal Review")
    tasks.create(conn, "reserve note", org_id=org.id, category=" internal note ")
    tasks.create(conn, "walk the site", org_id=org.id, category="Client internal audit")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    tasks.create(conn, "misc", org_id=org.id)
    return org


def test_the_generated_workbook_has_no_section_label_reading_internal(conn, tmp_path):
    """C9, load-bearing: drive a real workbook and read the rendered bands.

    A task categorised exactly "Internal" is withheld already. "Internal
    Review" ships — correctly, by the exact-match rule — and used to ship
    under a banner literally naming it, which reads as a leak of our private
    list whatever the item beneath it says. The rows stay; the heading goes."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    org = _banner_fixture(conn)
    path = write(conn, org.id, tmp_path / "b.xlsx", date(2026, 8, 18))
    sections = _rendered_sections(load_workbook(path).active)

    labels = [label for label, _ in sections]
    assert not any(label.strip().lower().startswith("internal") for label in labels), labels

    # and the rows themselves are still there, and still counted
    items = [item for _, rows in sections for item in rows]
    assert sorted(items) == [
        "audit support", "misc", "renew GL", "reserve note", "walk the site",
    ], items
    assert "our own file note" not in items  # the exact category, still withheld


def test_the_de_labelled_rows_land_in_general(conn, tmp_path):
    """Not a headerless block: towerkit renders sections back to back with no
    separator and restarts its banding per section, so rows under no banner
    read as belonging to whichever section printed above them — an "Internal
    Review" row silently filed under "Compliance". They join General, which is
    what a row whose category cannot be shown honestly is."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    org = _banner_fixture(conn)
    path = write(conn, org.id, tmp_path / "g.xlsx", date(2026, 8, 18))
    sections = dict(_rendered_sections(load_workbook(path).active))

    assert sorted(sections["General — Banner Co"]) == [
        "audit support", "misc", "reserve note",
    ]
    assert sections["Renewal — Banner Co"] == ["renew GL"]


def test_the_suppression_is_a_prefix_not_a_contains(conn):
    """"Client internal audit" is a real client-facing broking task and its
    heading says nothing about our private list. Equality is already what
    withholds, so this rule has to be wider than equality and narrower than
    containment — exactly the prefix."""
    org = _banner_fixture(conn)
    labels = [s.label for s in compose(conn, org.id, date(2026, 8, 18))]
    assert "Client internal audit — Banner Co" in labels
    assert "Internal Review — Banner Co" not in labels
    assert "internal note — Banner Co" not in labels


def test_the_exact_internal_category_is_still_withheld_not_relabelled(conn):
    """The two rules must not merge. Suppressing the heading moves rows;
    withholding removes them. A task categorised exactly "Internal" must not
    reappear in General wearing no label."""
    org = orgs.create(conn, name="Still Co", kind="client")
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    items = [r.item for s in compose(conn, org.id, date(2026, 8, 18)) for r in s.rows]
    assert items == ["renew GL"]


def test_general_appears_for_de_labelled_rows_alone(conn):
    """compose() emits General only `if general_rows`. When the near-miss rows
    are the ONLY rows on the account, that guard has to be satisfied by them —
    otherwise suppressing the heading deletes the section and the rows with
    it."""
    org = orgs.create(conn, name="Only Co", kind="client")
    tasks.create(conn, "audit support", org_id=org.id, category="Internal Review")
    sections = compose(conn, org.id, date(2026, 8, 18))
    assert [(s.label, [r.item for r in s.rows]) for s in sections] == [
        ("General — Only Co", ["audit support"])
    ]


def test_our_own_copy_keeps_the_internal_headings(conn):
    """include_internal is the client-copy switch, and C9 rides it: MCP's
    open_items is Grant's book, not the deliverable, and a category is context
    there rather than a leak."""
    org = _banner_fixture(conn)
    labels = [
        s.label
        for s in compose(conn, org.id, date(2026, 8, 18), include_internal=True)
    ]
    assert "Internal — Banner Co" in labels
    assert "Internal Review — Banner Co" in labels


def test_suppressing_the_heading_does_not_swallow_the_operators_line(conn, tmp_path):
    """Two readers, two lines. withheld_note still names the near miss to us
    at write time — the client's copy losing the banner is exactly why that
    line matters more, not less — and it still never enters the workbook."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import withheld_note, write

    org = _banner_fixture(conn)
    note = withheld_note(conn, org.id)
    assert note == (
        ' — 1 internal task withheld; 3 tasks categorised "Client internal audit", '
        '"internal note", "Internal Review" WERE exported '
        '(only the exact category "Internal" is withheld)'
    )

    path = write(conn, org.id, tmp_path / "n.xlsx", date(2026, 8, 18))
    values = [str(c.value) for row in load_workbook(path).active.iter_rows() for c in row]
    assert not any("only the exact category" in v for v in values)

# --- C8: the withholding rule, stated once, in fixed wording ----------------


def test_sheet_one_opens_straight_on_the_body(conn, tmp_path):
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Scope Co", kind="client")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    path = write(conn, org.id, tmp_path / "s.xlsx", date(2026, 8, 18))
    ws = load_workbook(path).active
    # Row 1 is the header; the FIRST section label follows it directly. The
    # scope line that used to sit between them was removed 2026-08-21 (Grant).
    assert ws["A2"].value == "Renewal — Scope Co"
    assert ws["A3"].value != _WITHDRAWN_SCOPE_NOTE


def test_the_scope_line_is_the_same_words_whether_or_not_anything_was_withheld(
    conn, tmp_path
):
    """The whole point of C8. A sentence that changed shape when something was
    held back would BE the count it was chosen instead of — it converts a
    non-event into a standing question on every export."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import withheld_note, write

    clean = orgs.create(conn, name="Clean Co", kind="client")
    tasks.create(conn, "renew GL", org_id=clean.id, category="Renewal")

    held = orgs.create(conn, name="Held Co", kind="client")
    tasks.create(conn, "renew GL", org_id=held.id, category="Renewal")
    tasks.create(conn, "our own file note", org_id=held.id, category="Internal")
    tasks.create(conn, "our own reserve note", org_id=held.id, category="internal")

    assert withheld_note(conn, held.id)  # something WAS withheld on this one
    assert withheld_note(conn, clean.id) == ""

    a = load_workbook(write(conn, clean.id, tmp_path / "a.xlsx", date(2026, 8, 18)))
    b = load_workbook(write(conn, held.id, tmp_path / "b.xlsx", date(2026, 8, 18)))
    # A2 is each account's own first section label now, so the two workbooks
    # legitimately differ there. The invariant that MATTERED is below and is
    # unchanged: the account that had something held back must not announce it,
    # by the scope line or by any other route.
    assert not any(
        _WITHDRAWN_SCOPE_NOTE == str(c.value)
        for wb in (a, b) for row in wb.active.iter_rows() for c in row
    )

    # and no count of what was withheld reaches the file by any other route
    held_values = [str(c.value) for row in b.active.iter_rows() for c in row]
    assert not any("withheld" in v for v in held_values)
    assert not any("our own file note" in v for v in held_values)


def test_the_scope_line_appears_nowhere_in_the_workbook(conn, tmp_path):
    """Once per export, not once per sheet. Sheet 1 is the only sheet always
    present, so a line there is a line on every export. The sentence covers
    the whole workbook (it says "this report", and every sheet it covers obeys
    the rule — see test_no_sheet_contradicts_the_scope_note), which is why it
    is stated once rather than repeated per sheet where two copies could
    drift."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    client = orgs.create(conn, kind="client", name="Four Co", status="active")
    tasks.create(conn, "renew GL", org_id=client.id, category="Renewal")
    placements.create(conn, client.id, "Four Program", "2026-01-01", "2027-01-01",
                      status="bound", total_premium=100_000_00)
    project = projects_repo.create_project(conn, client.id, "Fitout")
    projects_repo.add_need(conn, project.id, "Builder's Risk", "2026-12-01")
    req = rfi.create_request(conn, client.id, "onboarding docs", "2026-09-05")
    rfi.add_item(conn, req.id, "audited financials")

    wb = load_workbook(write(conn, client.id, tmp_path / "f.xlsx", date(2026, 8, 18)))
    assert len(wb.sheetnames) == 4, wb.sheetnames
    hits = [
        (name, c.coordinate)
        for name in wb.sheetnames
        for row in wb[name].iter_rows()
        for c in row
        if c.value == _WITHDRAWN_SCOPE_NOTE
    ]
    assert hits == [], hits


def test_no_sheet_contradicts_the_scope_note(conn, tmp_path):
    """THE reviewer's finding, at the level a client reads it. Sheet 1 says
    "Internal administrative items are not included" and says it about "this
    report"; sheet 2 shipped an item categorised Internal under a heading
    naming it. A scope note that is false about the document containing it is
    worse than either half alone — a reader who checks is told the wrong thing
    by the document itself. Fixed by extending the rule to sheet 2, not by
    narrowing the sentence to sheet 1."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Contradiction Co", kind="client")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    tasks.create(conn, "our own file note", org_id=org.id, category="Internal")
    req = rfi.create_request(conn, org.id, "onboarding docs", "2026-08-05")
    rfi.add_item(conn, req.id, "audited financials", category="Financials")
    rfi.add_item(conn, req.id, "pull prior loss runs from our file",
                 category="Internal")

    wb = load_workbook(write(conn, org.id, tmp_path / "c.xlsx", date(2026, 8, 18)))
    values = [
        str(c.value)
        for name in wb.sheetnames
        for row in wb[name].iter_rows()
        for c in row
        if c.value is not None
    ]

    assert _WITHDRAWN_SCOPE_NOTE not in values  # the claim is no longer MADE
    assert "audited financials" in values  # the sheet is really there
    # …and the rule it used to state is still OBEYED, which is the half that
    # mattered: no sheet ships anything internal, stated or not.
    assert not any("nternal" in v for v in values)
    assert "pull prior loss runs from our file" not in values
    assert "our own file note" not in values


def test_the_operator_line_names_what_sheet_2_withheld(conn, tmp_path):
    """The other side of extending the rule: a withheld row that nothing
    announces is the silent deletion the whole design warns against. The
    client's file still says nothing — this line is ours."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import withheld_note, write

    org = orgs.create(conn, name="Announce Co", kind="client")
    tasks.create(conn, "renew GL", org_id=org.id, category="Renewal")
    req = rfi.create_request(conn, org.id, "onboarding docs", "2026-08-05")
    rfi.add_item(conn, req.id, "audited financials", category="Financials")
    rfi.add_item(conn, req.id, "reserve note", category="Internal")

    note = withheld_note(conn, org.id)
    assert "1 internal request item withheld" in note

    wb = load_workbook(write(conn, org.id, tmp_path / "a.xlsx", date(2026, 8, 18)))
    values = [
        str(c.value) for name in wb.sheetnames
        for row in wb[name].iter_rows() for c in row if c.value is not None
    ]
    assert not any("withheld" in v for v in values)


def test_the_scope_line_does_not_replace_the_operators_near_miss_line(
    conn, tmp_path
):
    """Two different lines for two different readers. withheld_note names a
    NEAR MISS to us at write time; it is not a count of withheld items and it
    must never enter the client's file."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import withheld_note, write

    org = orgs.create(conn, name="Near Miss Co", kind="client")
    tasks.create(conn, "audit support", org_id=org.id, category="Internal Review")
    note = withheld_note(conn, org.id)
    assert "WAS exported" in note  # still ours, unchanged by C8

    wb = load_workbook(write(conn, org.id, tmp_path / "n.xlsx", date(2026, 8, 18)))
    values = [str(c.value) for row in wb.active.iter_rows() for c in row]
    assert _WITHDRAWN_SCOPE_NOTE not in values
    assert not any("only the exact category" in v for v in values)
    assert _WITHDRAWN_SCOPE_NOTE not in note


def test_write_open_items_deterministic_and_styled(conn, tmp_path):
    from bookkit.services.export_open_items import write

    # reuse the fixture-building from test_compose_groups...
    client = orgs.create(conn, kind="client", name="Acme", status="active", owner="grant")
    market = orgs.create(conn, kind="market", name="Zurich", status="active")

    tasks.create(
        conn, "Chase updated loss runs", org_id=client.id,
        description="waiting on brief line from the client",
    )

    p = placements.create(
        conn, client.id, "Acme Property 25-26", "2025-10-01", "2026-10-01"
    )
    tasks.create(conn, "Confirm bound terms", placement_id=p.id)
    submissions.create(conn, market.id, "2026-07-01", placement_id=p.id)

    project = projects_repo.create_project(conn, client.id, "Warehouse Expansion")
    projects_repo.add_need(conn, project.id, "Builder's Risk", "2026-09-01")

    today = date.today()
    a = write(conn, client.id, tmp_path / "a.xlsx", today)
    b = write(conn, client.id, tmp_path / "b.xlsx", today)
    assert a.read_bytes() == b.read_bytes()
    from openpyxl import load_workbook  # test-only import; src never imports it
    ws = load_workbook(a).active
    assert [c.value for c in ws[1]] == [
        "Item", "Description", "Detail", "Type", "Due / Needed by", "Status",
        "Owner"]
    # the placement-only task (org_id NULL) must actually be in the workbook,
    # not silently dropped by an org_id-only fetch
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "Confirm bound terms" in values


def test_write_empty_book_says_so(conn, tmp_path):
    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Empty Co", kind="client")
    path = write(conn, org.id, tmp_path / "e.xlsx", date(2026, 8, 12))
    from openpyxl import load_workbook
    assert load_workbook(path).active["A2"].value == "No open items as of 2026-08-12"


def test_write_three_tab_order_and_headers(conn, tmp_path):
    from bookkit.services.export_open_items import write

    client = orgs.create(conn, kind="client", name="Acme", status="active", owner="grant")
    tasks.create(conn, "Chase updated loss runs", org_id=client.id)
    placements.create(
        conn, client.id, "Acme Property 25-26", "2025-10-01", "2026-10-01",
        status="bound", total_premium=250_000_00,
    )
    live = projects_repo.create_project(conn, client.id, "Warehouse Expansion")
    projects_repo.add_need(conn, live.id, "Builder's Risk", "2026-09-01")

    path = write(conn, client.id, tmp_path / "w.xlsx", date(2026, 8, 13))
    from openpyxl import load_workbook  # test-only import; src never imports it
    wb = load_workbook(path)
    # spec-fixed sheet order; sheet 1 title unchanged from the single-sheet era
    assert wb.sheetnames == ["Open Items — Acme", "Projects", "Schedule of Insurance"]
    assert [c.value for c in wb["Projects"][1]] == [
        "Line", "Notes", "Needed by", "Status", "Limit"]
    assert [c.value for c in wb["Schedule of Insurance"][1]] == [
        "Insured", "Line of Coverage", "Status", "Carrier", "Policy Number",
        "Effective Date", "Expiration Date", "Limits",
        "Deductible / SIR / Retention", "Premium"]
    # show_premiums=True end to end: the book-data premium in whole dollars
    soi_values = [c.value for row in wb["Schedule of Insurance"].iter_rows() for c in row]
    assert 250_000 in soi_values


def test_write_omits_sheets_without_data(conn, tmp_path):
    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Solo Co", kind="client")
    tasks.create(conn, "one open task", org_id=org.id)
    path = write(conn, org.id, tmp_path / "s.xlsx", date(2026, 8, 13))
    from openpyxl import load_workbook
    assert load_workbook(path).sheetnames == ["Open Items — Solo Co"]


def test_write_three_tab_deterministic(conn, tmp_path):
    from bookkit.services.export_open_items import write

    client = orgs.create(conn, kind="client", name="Det Co", status="active")
    placements.create(conn, client.id, "Det Program", "2025-01-01", "2026-01-01")
    live = projects_repo.create_project(conn, client.id, "Det Project")
    projects_repo.add_need(conn, live.id, "GL", "2026-09-01")
    a = write(conn, client.id, tmp_path / "a.xlsx", date(2026, 8, 13))
    b = write(conn, client.id, tmp_path / "b.xlsx", date(2026, 8, 13))
    assert a.read_bytes() == b.read_bytes()  # three sheets, one finalize, same bytes


def test_onboarding_completeness_derives_from_data(conn):
    from bookkit.services import onboarding

    org = orgs.create(conn, name="Newco", kind="client")
    states = {s.step.key: s.state for s in onboarding.completeness(conn, org.id)}
    assert states == {
        "org": onboarding.PARTIAL,       # name/kind exist by construction
        "contacts": onboarding.UNTOUCHED,
        "program": onboarding.UNTOUCHED,
        "projects": onboarding.UNTOUCHED,
        "followups": onboarding.UNTOUCHED,
    }
    assert onboarding.first_incomplete(conn, org.id) == "org"
    assert not onboarding.is_complete(conn, org.id)

    orgs.update(conn, org.id, owner="grant", industry="construction")
    contacts.create(conn, org.id, first_name="Ann", last_name="Lee",
                    email="ann@newco.com")
    placements.create(conn, org_id=org.id, program_name="Newco Package 26-27",
                      period_from="2026-09-01", period_to="2027-09-01")
    states = {s.step.key: s.state for s in onboarding.completeness(conn, org.id)}
    assert states["org"] == states["contacts"] == states["program"] == onboarding.COMPLETE
    assert onboarding.is_complete(conn, org.id)  # optional steps don't gate
    assert onboarding.first_incomplete(conn, org.id) == "projects"


def test_onboarding_contact_without_reach_is_partial(conn):
    from bookkit.services import onboarding

    org = orgs.create(conn, name="Newco2", kind="client")
    contacts.create(conn, org.id, first_name="Bo", last_name="Nil")  # no email/phone
    status = {s.step.key: s for s in onboarding.completeness(conn, org.id)}
    assert status["contacts"].state == onboarding.PARTIAL
    assert "email or phone" in status["contacts"].summary


def test_onboarding_followups_sees_placement_attached_task(conn):
    from bookkit.services import onboarding

    org = orgs.create(conn, name="Newco3", kind="client")
    p = placements.create(conn, org_id=org.id, program_name="Newco3 Package 26-27",
                          period_from="2026-09-01", period_to="2027-09-01")
    # placement-attached only: org_id NULL, placement_id set — legal per
    # tasks_repo.open_tasks_for_client's docstring. open_tasks(org_id=...)
    # alone drops this row; the followups step must not.
    tasks.create(conn, "Confirm bound terms", placement_id=p.id)

    status = {s.step.key: s for s in onboarding.completeness(conn, org.id)}
    assert status["followups"].state == onboarding.COMPLETE
    assert "1 open task" in status["followups"].summary


def test_incomplete_clients_lists_missing_labels(conn):
    from bookkit.services import onboarding

    org = orgs.create(conn, name="Fresh LLC", kind="client")  # status defaults to prospect
    got = onboarding.incomplete_clients(conn, date(2026, 8, 12))
    assert [o.id for o, _ in got] == [org.id]
    _, missing = got[0]
    assert "contacts" in missing and "program" in missing


def test_compose_projects_full_report_live_projects_only(conn):
    from bookkit.services.export_open_items import compose_projects

    org = orgs.create(conn, name="Proj Co", kind="client", status="active", owner="grant")
    live = projects_repo.create_project(
        conn, org.id, "Warehouse Expansion", status="active",
        start_on="2026-06-01", end_on="2027-06-01",
    )
    projects_repo.add_need(
        conn, live.id, "Builder's Risk", "2026-09-01",
        limit_cents=25_000_000_00, notes="GC requires evidence",
    )
    projects_repo.add_need(conn, live.id, "GL", "2026-09-15", status="placed")
    done = projects_repo.create_project(conn, org.id, "Old HQ Fit-out", status="completed")
    projects_repo.add_need(conn, done.id, "Property", "2025-01-01")
    projects_repo.create_project(conn, org.id, "Shelved", status="cancelled")

    sections = compose_projects(conn, org.id)
    assert len(sections) == 1  # completed and cancelled projects are not live
    section = sections[0]
    assert section.label == "Warehouse Expansion — Active (2026-06-01 → 2027-06-01)"
    # every need regardless of status; line, notes, needed-by, prettified
    # status, formatted limit — and NO days-open (five columns, no date math)
    assert section.rows == (
        ("Builder's Risk", "GC requires evidence", "2026-09-01", "Identified", "$25,000,000"),
        ("GL", "", "2026-09-15", "Placed", ""),
    )


def test_compose_projects_needless_live_project_still_sections(conn):
    from bookkit.services.export_open_items import compose_projects

    org = orgs.create(conn, name="Plan Co", kind="client")
    projects_repo.create_project(conn, org.id, "Planning Stage")  # status "planned"
    sections = compose_projects(conn, org.id)
    assert sections[0].label == "Planning Stage — Planned"
    assert sections[0].rows == ()


def test_compose_projects_empty_when_no_live_projects(conn):
    from bookkit.services.export_open_items import compose_projects

    org = orgs.create(conn, name="No Proj Co", kind="client")
    assert compose_projects(conn, org.id) == []
    projects_repo.create_project(conn, org.id, "Done", status="completed")
    assert compose_projects(conn, org.id) == []


def test_compose_soi_linked_placement_uses_towerkit_soi(conn, tmp_path):
    from towerkit.model import Layer, Line, Participant, Period, Program, dump_program
    from towerkit.model import Placement as TkPlacement

    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Linked Co", kind="client", status="active")
    p = placements.create(
        conn, org.id, "2026 Package", "2025-10-01", "2026-10-01", status="bound"
    )
    program = Program(
        insured="Linked Co", program="Package Program", placement=TkPlacement.BOUND,
        period=Period(start=date(2025, 10, 1), end=date(2026, 10, 1)),
        lines=[Line(id="gl", name="General Liability", abbr="GL")],
        layers=[
            Layer(
                id="gl1", name="Primary GL", applies_to=["gl"],
                attach=0, limit=1_000_000, premium=52_000,
                participants=[Participant(carrier="Zurich", share_bps=10_000)],
            )
        ],
    )
    path = tmp_path / "package.json"
    dump_program(program, path)
    placements.update(conn, p.id, program_path=str(path))

    sections = compose_soi(conn, org.id, date(2026, 1, 1))
    assert len(sections) == 1
    section = sections[0]
    # build_soi's unlabeled section takes the program name as its label — the
    # FILE's, not the book's. The two agree on any projected placement
    # (sync.project writes program_name=program.program); where they disagree
    # the book's is the stale one, and heading a casualty tower "2025 Property
    # Program" is the milder half of the mislink defect.
    assert section.label == "Package Program"
    row = section.rows[0]
    assert row.insured == "Linked Co"
    assert row.coverage == "General Liability"
    assert row.carrier == "Zurich"
    assert row.effective == date(2025, 10, 1)
    assert row.premium == 52_000  # whole dollars, straight from the file


def test_compose_soi_unlinked_placement_gets_book_data_section(conn):
    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Paper Co", kind="client")
    placements.create(
        conn, org.id, "Legacy Property", "2025-01-01", "2026-01-01",
        status="bound", total_premium=12_345_00,
    )
    sections = compose_soi(conn, org.id, date(2025, 6, 1))
    assert len(sections) == 1
    assert sections[0].label == "Legacy Property"  # the status is a column now
    row = sections[0].rows[0]
    assert row.insured == "Paper Co"
    assert row.coverage == "Legacy Property"
    assert row.carrier == "See policy documents"
    assert row.policy_number == ""
    assert row.effective == date(2025, 1, 1)
    assert row.expiration == date(2026, 1, 1)
    assert row.limits == "" and row.retention == ""
    assert row.premium == 12_345  # cents → whole dollars via the money boundary


def test_compose_soi_unreadable_file_falls_back_to_book_data(conn, tmp_path):
    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Moved Co", kind="client")
    placements.create(
        conn, org.id, "Moved Program", "2025-01-01", "2026-01-01",
        program_path=str(tmp_path / "gone.json"),  # linked, but the file moved
    )
    sections = compose_soi(conn, org.id, date(2025, 6, 1))
    assert len(sections) == 1  # the policy list is never silently partial
    # ...and it says so. The placement is prospective, so there are no policy
    # documents to send the reader to (see _UNPLACED_CARRIER), and the heading
    # states that a file was expected here and not used.
    assert sections[0].rows[0].carrier == "To be placed"
    assert "policy detail unavailable" in (sections[0].label or "")


def test_compose_soi_empty_when_no_placements(conn):
    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Bare Co", kind="client")
    assert compose_soi(conn, org.id, date(2026, 8, 18)) == []


# --- C3: expired policy years are not a current Schedule of Insurance --------


def test_compose_soi_excludes_an_expired_policy_year(conn):
    """A prior year renders IDENTICALLY to current cover — same columns, same
    styling, and a Status column still reading `Bound`, which is true in the
    past tense. Exclusion is decided on period_to, the same date the
    Expiration column prints."""
    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Two Years Co", kind="client")
    placements.create(conn, org.id, "2024 Casualty Program",
                      "2024-09-07", "2025-09-07", status="bound")
    placements.create(conn, org.id, "2025 Casualty Program",
                      "2025-09-07", "2026-09-07", status="bound")

    labels = [s.label for s in compose_soi(conn, org.id, date(2026, 8, 18))]
    assert labels == ["2025 Casualty Program"]


def test_compose_soi_keeps_cover_expiring_today(conn):
    """Cover runs to the END of its last day: the comparison is strictly `<`.
    A `<=` here drops a client's live policy from their schedule on its final
    day, which is the day they are most likely to be reading it."""
    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Last Day Co", kind="client")
    placements.create(conn, org.id, "Expiring Today", "2025-08-18", "2026-08-18",
                      status="bound")

    today = date(2026, 8, 18)
    assert [s.label for s in compose_soi(conn, org.id, today)] == ["Expiring Today"]
    # and gone the next morning
    assert compose_soi(conn, org.id, date(2026, 8, 19)) == []


def test_compose_soi_keeps_a_program_whose_earliest_line_has_already_lapsed(
    conn, tmp_path
):
    """THE renewal_on TRAP. renewal_on is the EARLIEST line end, because
    attention must open when the first layer runs out. Exclusion asks the
    opposite question. This program's IM layer ended in March; its GL runs to
    October. Deciding exclusion on the earliest line end would take a live
    General Liability policy off the client's Schedule of Insurance."""
    from towerkit.model import Layer, Line, Participant, Period, Program, dump_program
    from towerkit.model import Placement as TkPlacement

    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Staggered Co", kind="client", status="active")
    p = placements.create(
        conn, org.id, "2026 Package", "2025-10-01", "2026-10-01", status="bound"
    )
    program = Program(
        insured="Staggered Co", program="Package", placement=TkPlacement.BOUND,
        period=Period(start=date(2025, 10, 1), end=date(2026, 10, 1)),
        lines=[
            Line(id="gl", name="General Liability", abbr="GL"),
            Line(id="im", name="Inland Marine", abbr="IM"),
        ],
        layers=[
            Layer(
                id="gl1", name="Primary GL", applies_to=["gl"],
                attach=0, limit=1_000_000, premium=52_000,
                participants=[Participant(carrier="Zurich", share_bps=10_000)],
            ),
            Layer(
                id="im1", name="IM", applies_to=["im"],
                period=Period(start=date(2025, 10, 1), end=date(2026, 3, 1)),
                attach=0, limit=250_000, premium=4_000,
                participants=[Participant(carrier="Chubb", share_bps=10_000)],
            ),
        ],
    )
    path = tmp_path / "package.json"
    dump_program(program, path)
    placements.update(conn, p.id, program_path=str(path))

    from bookkit import sync
    # the trap is real: renewal_on for this placement IS the lapsed IM date
    assert sync.line_ends(str(path))[0][1] == date(2026, 3, 1)

    sections = compose_soi(conn, org.id, date(2026, 6, 1))
    coverages = [r.coverage for s in sections for r in s.rows]
    assert "General Liability" in coverages, (
        "live GL cover fell off the schedule — exclusion used the earliest "
        "line end instead of the program period"
    )


def test_all_expired_account_gets_no_soi_sheet_rather_than_an_empty_one(
    conn, tmp_path
):
    """The sheet-inclusion rule under C3: every placement historic ⇒ the tab
    is omitted, exactly as Information Requests and Projects are when empty.
    What must NEVER happen is a Schedule of Insurance sheet carrying column
    headers and no policies."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Lapsed Co", kind="client")
    placements.create(conn, org.id, "2024 Program", "2024-01-01", "2025-01-01",
                      status="bound")
    tasks.create(conn, "renew the whole account", org_id=org.id, category="Renewal")

    path = write(conn, org.id, tmp_path / "l.xlsx", date(2026, 8, 18))
    wb = load_workbook(path)
    assert "Schedule of Insurance" not in wb.sheetnames

    # and the same account, exported while that year was still running, does
    # get the sheet — so the absence above is C3 and not a broken writer
    path2 = write(conn, org.id, tmp_path / "l2.xlsx", date(2024, 6, 1))
    wb2 = load_workbook(path2)
    assert "Schedule of Insurance" in wb2.sheetnames
    assert wb2["Schedule of Insurance"]["A2"].value == "2024 Program"


def test_premium_dollars_delegates_then_floors_for_display():
    from bookkit.services.export_open_items import _premium_dollars

    assert _premium_dollars(None) is None
    assert _premium_dollars(500_000_00) == 500_000
    # sub-dollar cents: money.cents_to_dollars refuses; the SOI's whole-dollar
    # display column floors instead (format_cents_compact's documented rule)
    assert _premium_dollars(500_000_50) == 500_000


# --- C2: the Status column, and the label suffix it replaces ----------------


def test_soi_status_mapping_covers_every_placement_status():
    """A dict lookup on a status the map has never heard of must FAIL, not
    hand a client's schedule a blank Status cell whose premium then quietly
    lands under `Unbound cover`. The module asserts this at import; this test
    is the same guard where a reader will find it."""
    from bookkit.models import PlacementStatus
    from bookkit.services.export_open_items import _SOI_STATUS

    assert set(_SOI_STATUS) == set(PlacementStatus)


def test_book_data_section_states_status_in_the_column_not_the_label(conn):
    """THE finding. A placement WITHOUT a program file used to carry `(Bound)`
    in its section label while a linked one carried nothing — the more we knew
    about a programme, the less the client could tell whether it was real. The
    status is a row field now, so the suffix is gone from the label and the
    premium of genuinely bound unlinked cover lands under `Bound cover`."""
    from towerkit.soi import NOT_STATED, SoiStatus, premium_subtotal

    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Paper Co", kind="client")
    placements.create(
        conn, org.id, "Legacy Property", "2025-01-01", "2026-01-01",
        status="bound", total_premium=12_345_00,
    )
    (section,) = compose_soi(conn, org.id, date(2025, 6, 1))
    assert section.label == "Legacy Property"  # no "(Bound)" suffix
    row = section.rows[0]
    assert row.status == SoiStatus.BOUND
    assert row.is_bound is True
    # and the money lands on the right subtotal line, which is the whole point
    assert premium_subtotal(section, bound=True) == 12_345
    assert premium_subtotal(section, bound=False) == NOT_STATED


@pytest.mark.parametrize(
    "book_status,expected",
    [
        ("prospective", "To be placed"),
        ("submitted", "Submitted"),
        ("quoted", "Quoted"),
        ("bound", "Bound"),
        ("lapsed", "Expired"),
    ],
)
def test_book_data_status_maps_every_placement_status(conn, book_status, expected):
    """Each of bookkit's five, in the client-facing words towerkit prints.
    `lapsed` on cover whose period has not run out is a mid-term cancellation:
    the year is not over, so the row is still on the schedule — saying
    `Expired` is exactly how the reader learns the cover is gone."""
    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name=f"Co {book_status}", kind="client")
    placements.create(
        conn, org.id, "A Program", "2025-01-01", "2026-01-01",
        status=book_status, total_premium=10_000_00,
    )
    (section,) = compose_soi(conn, org.id, date(2025, 6, 1))
    assert section.rows[0].status == expected
    # only Bound is cover in force — everything else is unbound premium
    assert section.rows[0].is_bound is (expected == "Bound")


def _staggered_program(insured: str, *, placed: bool):
    """A two-layer program: GL fully placed with Zurich, Property with NO
    participants — towerkit reads the second as `To be placed` on its own."""
    from towerkit.model import Layer, Line, Participant, Period, Program
    from towerkit.model import Placement as TkPlacement

    return Program(
        insured=insured, program="Package",
        placement=TkPlacement.BOUND if placed else TkPlacement.PROPOSED,
        period=Period(start=date(2025, 10, 1), end=date(2026, 10, 1)),
        lines=[
            Line(id="gl", name="General Liability", abbr="GL"),
            Line(id="prop", name="Property", abbr="PROP"),
        ],
        layers=[
            Layer(
                id="gl1", name="Primary GL", applies_to=["gl"],
                attach=0, limit=1_000_000, premium=52_000,
                participants=[Participant(carrier="Zurich", share_bps=10_000)],
            ),
            Layer(
                id="prop1", name="Primary Property", applies_to=["prop"],
                attach=0, limit=5_000_000, premium=31_000,
                participants=[],  # nobody on the risk yet
            ),
        ],
    )


def test_compose_soi_linked_bound_keeps_towerkits_per_layer_status(conn, tmp_path):
    """A BOUND placement whose file says one layer is still unplaced. The
    file's per-layer status is BETTER than the placement's — it can say `To be
    placed` for one layer of an otherwise bound programme — so a bound
    placement must not flatten it. Overriding every row with `Bound` here
    would print bound cover over a layer nobody is on, and sweep its premium
    into the `Bound cover` subtotal."""
    from towerkit.model import dump_program
    from towerkit.soi import SoiStatus

    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Staggered Co", kind="client", status="active")
    p = placements.create(
        conn, org.id, "2026 Package", "2025-10-01", "2026-10-01", status="bound"
    )
    path = tmp_path / "package.json"
    dump_program(_staggered_program("Staggered Co", placed=True), path)
    placements.update(conn, p.id, program_path=str(path))

    sections = compose_soi(conn, org.id, date(2026, 1, 1))
    by_coverage = {r.coverage: r for s in sections for r in s.rows}
    assert by_coverage["General Liability"].status == SoiStatus.BOUND
    assert by_coverage["Property"].status == SoiStatus.TO_BE_PLACED, (
        "a bound placement flattened the file's per-layer status — an unplaced "
        "layer is being printed as bound cover"
    )
    assert by_coverage["Property"].is_bound is False
    # both layers share one section (neither line carries a group), so the two
    # subtotal lines are where the split shows: GL's premium is cover in force,
    # Property's is not
    (section,) = sections
    assert section.bound_premium_total == 52_000
    assert section.unbound_premium_total == 31_000


def test_compose_soi_linked_unbound_placement_overrides_every_row(conn, tmp_path):
    """The other direction: the BOOK knows something the file cannot. The file
    is a design the broker has fully populated, so towerkit reads its layers as
    bound — but the book says the placement is only QUOTED, and nothing on a
    quoted placement is cover in force. Every row is overridden, and no
    premium reaches the `Bound cover` subtotal."""
    from towerkit.model import Layer, Line, Participant, Period, Program, dump_program
    from towerkit.model import Placement as TkPlacement
    from towerkit.soi import NOT_STATED, SoiStatus, premium_subtotal

    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Quoted Co", kind="client", status="active")
    p = placements.create(
        conn, org.id, "2026 Casualty", "2025-10-01", "2026-10-01", status="quoted"
    )
    program = Program(
        insured="Quoted Co", program="Casualty", placement=TkPlacement.BOUND,
        period=Period(start=date(2025, 10, 1), end=date(2026, 10, 1)),
        lines=[Line(id="gl", name="General Liability", abbr="GL")],
        layers=[
            Layer(
                id="gl1", name="Primary GL", applies_to=["gl"],
                attach=0, limit=1_000_000, premium=52_000,
                participants=[Participant(carrier="Zurich", share_bps=10_000)],
            )
        ],
    )
    path = tmp_path / "casualty.json"
    dump_program(program, path)
    placements.update(conn, p.id, program_path=str(path))

    (section,) = compose_soi(conn, org.id, date(2026, 1, 1))
    assert section.rows[0].status == SoiStatus.QUOTED
    assert section.rows[0].is_bound is False
    assert premium_subtotal(section, bound=True) == NOT_STATED
    assert premium_subtotal(section, bound=False) == 52_000


def test_soi_sheet_prints_the_status_column_and_both_subtotals(conn, tmp_path):
    """End to end, in the file a client opens: the Status column carries the
    word, and cover in force is in a different BLOCK from cover that is not —
    each with its own subtotal, never one mingled list with two lines under
    it (Grant, 2026-08-21)."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Sheet Co", kind="client", status="active")
    placements.create(
        conn, org.id, "Bound Property", "2025-01-01", "2026-12-01",
        status="bound", total_premium=100_000_00,
    )
    placements.create(
        conn, org.id, "Quoted Casualty", "2025-01-01", "2026-12-01",
        status="quoted", total_premium=40_000_00,
    )
    tasks.create(conn, "something open", org_id=org.id)

    path = write(conn, org.id, tmp_path / "sheet.xlsx", date(2026, 6, 1))
    sheet = load_workbook(path)["Schedule of Insurance"]
    assert [c.value for c in sheet[1]][:3] == ["Insured", "Line of Coverage", "Status"]
    cells = [[c.value for c in row] for row in sheet.iter_rows()]
    flat = [c for row in cells for c in row]
    # section labels carry no status suffix any more
    assert "Bound Property" in flat and "Quoted Casualty" in flat
    assert not any(
        isinstance(c, str) and c.endswith("(Bound)") for c in flat
    ), "the label suffix survived the move to a Status column"
    assert "Bound" in flat and "Quoted" in flat
    # ONE BLOCK PER SECTION PER BOUND-NESS since 2026-08-21 (Grant): the
    # primary schedule is bound cover, and anything not bound is lifted into
    # its own block below with its own single subtotal. So a wholly-bound
    # placement grows no unbound line at all, and a wholly-unbound one appears
    # ONLY under a "— not bound" heading. Walk the sheet and file every
    # subtotal under whichever block heading was above it.
    subtotals: dict[str, object] = {}
    current = ""
    for row in cells:
        head = row[0]
        if not isinstance(head, str):
            continue
        if head.startswith(("Bound Property", "Quoted Casualty")):
            current = head
        elif "premium subtotal" in head:
            subtotals[f"{current} / {head}"] = row[-1]
    assert subtotals == {
        "Bound Property / Bound cover — premium subtotal": 100_000,
        "Quoted Casualty — not bound / Unbound cover — premium subtotal": 40_000,
    }, subtotals
    # And the claim that matters to a reader: nothing anywhere on the sheet
    # adds the two together.
    assert 140_000 not in flat


# --- client safety on the Schedule of Insurance ----------------------------


def _one_layer_program(insured: str, program_name: str, **kw):
    """A single-layer bound casualty program, so a test can vary the ONE thing
    it is about."""
    from towerkit.model import Layer, Line, Participant, Period, Program
    from towerkit.model import Placement as TkPlacement

    return Program(
        insured=insured, program=program_name, placement=TkPlacement.BOUND,
        period=Period(start=date(2025, 10, 1), end=date(2026, 10, 1)),
        lines=[Line(id="gl", name="General Liability", abbr="GL")],
        layers=[Layer(
            id="gl1", name="Primary GL", applies_to=["gl"], attach=0,
            limit=1_000_000, premium=52_000,
            participants=[Participant(carrier="Zurich", share_bps=10_000)],
        )],
        **kw,
    )


def test_a_program_file_confirmed_to_another_account_exports_no_rows(conn, tmp_path):
    """THE ONE THAT SHIPS A COMPETITOR'S TOWER. A placement on this client
    points at a file whose program_link confirms it to a DIFFERENT account:
    every row of that file — insured, carriers, shares, limits, premiums —
    used to land in this client's workbook under this client's file name.

    The file is refused; the EXPORT is not. The placement still prints from
    book data, under a heading that says the detail is unavailable, and the
    operator is told whose file it actually is."""
    from towerkit.model import dump_program

    from bookkit.repo import links
    from bookkit.services.export_open_items import compose_soi, soi_problems

    theirs = orgs.create(conn, name="Atomic Industries, Inc.", kind="client")
    ours = orgs.create(conn, name="Probe Holdings Ltd", kind="client")
    path = tmp_path / "atomic-casualty.json"
    dump_program(_one_layer_program("Atomic Industries, Inc.", "Casualty Program"), path)
    links.confirm(conn, str(path), theirs.id, "Atomic Industries, Inc.")

    p = placements.create(conn, ours.id, "2026 Casualty Program",
                          "2025-10-01", "2026-10-01", status="bound",
                          total_premium=1_000_000_00)
    placements.update(conn, p.id, program_path=str(path))

    (section,) = compose_soi(conn, ours.id, date(2026, 8, 18))
    insureds = {row.insured for row in section.rows}
    assert insureds == {"Probe Holdings Ltd"}, "another client's insured shipped"
    assert not any(row.carrier == "Zurich" for row in section.rows), (
        "another client's carriers shipped"
    )
    assert "policy detail unavailable" in (section.label or "")
    # the client's copy never names the other account; the operator's line does
    assert "Atomic" not in (section.label or "")
    assert "Atomic Industries, Inc." in soi_problems(conn, ours.id, date(2026, 8, 18))[0]


def test_an_unconfirmed_file_naming_a_different_insured_exports_no_rows(conn, tmp_path):
    """No program_link row at all — a program_path set outside sync.project.
    There is no org id to compare, so the insured string is compared, and any
    disagreement loses. Nothing fuzzy: an identity we cannot vouch for costs
    this client a summary row; being wrong ships someone else's tower."""
    from towerkit.model import dump_program

    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Probe Holdings Ltd", kind="client")
    path = tmp_path / "someone-else.json"
    dump_program(_one_layer_program("Atomic Industries, Inc.", "Casualty Program"), path)
    p = placements.create(conn, org.id, "2026 Casualty Program",
                          "2025-10-01", "2026-10-01", status="bound")
    placements.update(conn, p.id, program_path=str(path))

    (section,) = compose_soi(conn, org.id, date(2026, 8, 18))
    assert {r.insured for r in section.rows} == {"Probe Holdings Ltd"}
    assert "policy detail unavailable" in (section.label or "")


def test_a_matching_insured_with_no_link_row_is_accepted(conn, tmp_path):
    """The other side of the same guard: case and stray whitespace are not a
    different company, so a hand-linked file that names this account is USED.
    Without this the refusal would blank out every honestly linked schedule,
    which is the cost that makes an all-or-nothing refusal wrong."""
    from towerkit.model import dump_program

    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Probe Holdings Ltd", kind="client")
    path = tmp_path / "probe.json"
    dump_program(_one_layer_program("probe   holdings ltd", "Casualty Program"), path)
    p = placements.create(conn, org.id, "2026 Casualty Program",
                          "2025-10-01", "2026-10-01", status="bound")
    placements.update(conn, p.id, program_path=str(path))

    (section,) = compose_soi(conn, org.id, date(2026, 8, 18))
    assert [r.carrier for r in section.rows] == ["Zurich"]


def test_a_stale_program_name_loses_to_the_file_and_is_reported(conn, tmp_path):
    """The milder half, which the seeded book reproduces on its own: the
    placement is called "2025 Property Program" and its linked file is a
    casualty tower. A stale LABEL is not grounds to discard a correct TOWER,
    so the file wins the heading and the operator is told to re-project."""
    from towerkit.model import dump_program

    from bookkit.services.export_open_items import compose_soi, soi_problems

    org = orgs.create(conn, name="Atomic Industries, Inc.", kind="client")
    path = tmp_path / "atomic-casualty.json"
    dump_program(_one_layer_program("Atomic Industries, Inc.", "Casualty Program"), path)
    p = placements.create(conn, org.id, "2025 Property Program",
                          "2025-10-01", "2026-10-01", status="bound")
    placements.update(conn, p.id, program_path=str(path))

    (section,) = compose_soi(conn, org.id, date(2026, 8, 18))
    assert section.label == "Casualty Program"
    assert "2025 Property Program" not in (section.label or "")
    assert [r.carrier for r in section.rows] == ["Zurich"]  # the tower survives
    (problem,) = soi_problems(conn, org.id, date(2026, 8, 18))
    assert "Casualty Program" in problem and "re-project" in problem


def test_a_run_off_layer_says_expired_and_leaves_the_bound_subtotal(conn, tmp_path):
    """CRITICAL 2. An excess layer whose own period ended 49 days ago, inside
    a programme whose other layers run to November. It printed `Bound`, with
    its premium in the `Bound cover` subtotal: $10,000,000 xs $2,000,000 of
    cover the client does not hold, and money they are not paying.

    The placement-level EXCLUSION rule is untouched (the programme is still on
    the schedule, which is the point of _expired's docstring); what changes is
    what each ROW says."""
    from towerkit.model import Layer, Line, Participant, Period, Program, dump_program
    from towerkit.model import Placement as TkPlacement
    from towerkit.soi import SoiStatus, premium_subtotal

    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Vertex Manufacturing Co", kind="client")
    program = Program(
        insured="Vertex Manufacturing Co", program="Casualty Program",
        placement=TkPlacement.BOUND,
        period=Period(start=date(2025, 11, 1), end=date(2026, 11, 1)),
        lines=[Line(id="gl", name="General Liability", abbr="GL")],
        layers=[
            Layer(id="gl1", name="Primary GL", applies_to=["gl"], attach=0,
                  limit=2_000_000, premium=1_250_000,
                  participants=[Participant(carrier="Zurich", share_bps=10_000)]),
            Layer(id="gl2", name="Excess GL", applies_to=["gl"],
                  attach=2_000_000, limit=10_000_000, premium=500_000,
                  period=Period(start=date(2025, 11, 1), end=date(2026, 6, 30)),
                  participants=[Participant(carrier="Swiss Re", share_bps=6_000),
                                Participant(carrier="Chubb", share_bps=4_000)]),
        ],
    )
    path = tmp_path / "vertex.json"
    dump_program(program, path)
    p = placements.create(conn, org.id, "Casualty Program",
                          "2025-11-01", "2026-11-01", status="bound")
    placements.update(conn, p.id, program_path=str(path))

    today = date(2026, 8, 18)
    (section,) = compose_soi(conn, org.id, today)
    by_coverage = {r.coverage: r for r in section.rows}
    ran_off = by_coverage["General Liability — Excess GL"]
    assert ran_off.expiration == date(2026, 6, 30)
    assert ran_off.status == SoiStatus.EXPIRED
    assert ran_off.is_bound is False
    # the live layer is untouched, and the programme is still on the schedule
    assert by_coverage["General Liability — Primary GL"].status == SoiStatus.BOUND
    # and the money moved with the word
    assert premium_subtotal(section, bound=True) == 1_250_000
    assert premium_subtotal(section, bound=False) == 500_000


def test_a_layer_expiring_today_is_still_bound(conn, tmp_path):
    """Cover runs to the end of its last day — the same strictly-`<` rule
    _expired keeps one level up. A `<=` here would call a client's live layer
    expired on the day they are most likely to be reading the schedule."""
    from towerkit.model import Layer, Line, Participant, Period, Program, dump_program
    from towerkit.model import Placement as TkPlacement
    from towerkit.soi import SoiStatus

    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Last Layer Co", kind="client")
    program = Program(
        insured="Last Layer Co", program="Casualty Program",
        placement=TkPlacement.BOUND,
        period=Period(start=date(2025, 11, 1), end=date(2026, 11, 1)),
        lines=[Line(id="gl", name="General Liability", abbr="GL")],
        layers=[Layer(
            id="gl1", name="Primary GL", applies_to=["gl"], attach=0,
            limit=2_000_000, premium=1_000,
            period=Period(start=date(2025, 11, 1), end=date(2026, 8, 18)),
            participants=[Participant(carrier="Zurich", share_bps=10_000)],
        )],
    )
    path = tmp_path / "lastday.json"
    dump_program(program, path)
    p = placements.create(conn, org.id, "Casualty Program",
                          "2025-11-01", "2026-11-01", status="bound")
    placements.update(conn, p.id, program_path=str(path))

    (section,) = compose_soi(conn, org.id, date(2026, 8, 18))
    assert section.rows[0].status == SoiStatus.BOUND
    (section,) = compose_soi(conn, org.id, date(2026, 8, 19))
    assert section.rows[0].status == SoiStatus.EXPIRED


def test_an_unreadable_tower_prints_what_the_book_holds_and_says_so(conn, tmp_path):
    """CRITICAL 3. A six-layer D&O tower and a moved file were
    indistinguishable in the output: one row, no limit, no policy number, no
    retention, no carrier, and nothing anywhere saying data had been lost.

    The book still held the total limits; the sheet now prints them, and the
    heading states that the detail is missing."""
    from bookkit.services.export_open_items import compose_soi, soi_problems

    org = orgs.create(conn, name="Zenith Foods Holdings, LLC", kind="client")
    p = placements.create(
        conn, org.id, "2026 Executive Risk Program", "2026-01-01", "2027-01-01",
        status="bound", total_limit=10_000_000_00, total_premium=485_000_00,
    )
    placements.update(conn, p.id, program_path=str(tmp_path / "gone.json"))

    today = date(2026, 8, 18)
    (section,) = compose_soi(conn, org.id, today)
    assert "policy detail unavailable" in (section.label or "")
    assert section.rows[0].limits == "Total limits $10,000,000"
    (problem,) = soi_problems(conn, org.id, today)
    assert "could not read" in problem and "gone.json" in problem
    # the client's copy never carries the path
    assert "gone.json" not in (section.label or "")


def test_a_corrupt_program_file_is_reported_not_swallowed(conn, tmp_path):
    """A file that parses to nothing takes the same visible path as one that
    moved — a bare `except Exception` made both silent, and made a bug in
    this module silent too."""
    from bookkit.services.export_open_items import compose_soi, soi_problems

    org = orgs.create(conn, name="Corrupt Co", kind="client")
    path = tmp_path / "half-written.json"
    path.write_text('{"insured": "Corrupt Co", "program": ', encoding="utf-8")
    p = placements.create(conn, org.id, "2026 Property Program",
                          "2026-02-01", "2027-02-01", status="bound",
                          total_premium=310_000_00)
    placements.update(conn, p.id, program_path=str(path))

    (section,) = compose_soi(conn, org.id, date(2026, 8, 18))
    assert "policy detail unavailable" in (section.label or "")
    (problem,) = soi_problems(conn, org.id, date(2026, 8, 18))
    assert "JSONDecodeError" in problem


def test_an_unreadable_file_does_not_stop_the_rest_of_the_export(conn, tmp_path):
    """The refusal is scoped to the placement, never to the deliverable. A
    file we cannot vouch for must not cost the client their other programmes,
    their open items or their information requests."""
    from openpyxl import load_workbook
    from towerkit.model import dump_program

    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Mixed Co", kind="client", status="active")
    good = tmp_path / "good.json"
    dump_program(_one_layer_program("Mixed Co", "Casualty Program"), good)
    p_good = placements.create(conn, org.id, "Casualty Program",
                               "2025-10-01", "2026-10-01", status="bound")
    placements.update(conn, p_good.id, program_path=str(good))
    p_bad = placements.create(conn, org.id, "Property Program",
                              "2025-11-01", "2026-11-01", status="bound",
                              total_premium=90_000_00)
    placements.update(conn, p_bad.id, program_path=str(tmp_path / "gone.json"))
    tasks.create(conn, "still an open item", org_id=org.id)

    path = write(conn, org.id, tmp_path / "mixed.xlsx", date(2026, 8, 18))
    wb = load_workbook(path)
    assert "Schedule of Insurance" in wb.sheetnames
    flat = [c.value for row in wb["Schedule of Insurance"].iter_rows() for c in row]
    assert "Zurich" in flat, "the readable programme was lost with the bad one"
    assert any(isinstance(c, str) and "policy detail unavailable" in c for c in flat)
    assert any(
        c.value == "still an open item"
        for row in wb[wb.sheetnames[0]].iter_rows() for c in row
    )


def test_unplaced_cover_never_sends_the_reader_to_policy_documents(conn):
    """A row reading `To be placed`, effective next January, told the reader
    to `See policy documents` — documents that do not exist. They conclude
    they misfiled them rather than that there is no policy."""
    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Nothing Yet Co", kind="client")
    placements.create(conn, org.id, "2026 Property Program",
                      "2026-01-01", "2027-01-01", status="prospective")
    placements.create(conn, org.id, "2026 Casualty Program",
                      "2026-01-01", "2027-01-01", status="bound")

    carriers = {
        s.rows[0].status: s.rows[0].carrier
        for s in compose_soi(conn, org.id, date(2026, 8, 18))
    }
    assert carriers["To be placed"] == "To be placed"
    assert carriers["Bound"] == "See policy documents"


def test_cover_that_has_not_incepted_is_not_on_a_schedule_of_what_is_held(conn):
    """`placements.for_org` orders by period_to DESC, so a 2027 renewal was
    the FIRST row a client saw on a schedule of what they hold today. A
    Schedule of Insurance states cover IN FORCE; the window is closed at both
    ends, by the same rule and for the same reason as the expired one."""
    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Northwind Freight LLC", kind="client")
    placements.create(conn, org.id, "2027 Property Program",
                      "2027-01-01", "2028-01-01", status="prospective")
    placements.create(conn, org.id, "2026 Property Program",
                      "2026-01-01", "2027-01-01", status="bound")

    labels = [s.label for s in compose_soi(conn, org.id, date(2026, 8, 18))]
    assert labels == ["2026 Property Program"]


def test_cover_incepting_today_is_in_force(conn):
    """The mirror of the expiring-today rule: cover begins at the start of its
    first day, so the comparison is strictly `>`. Together the two rules leave
    no day of a policy year outside the window."""
    from bookkit.services.export_open_items import compose_soi

    org = orgs.create(conn, name="Day One Co", kind="client")
    placements.create(conn, org.id, "Starts Today", "2026-08-18", "2027-08-18",
                      status="bound")

    assert [s.label for s in compose_soi(conn, org.id, date(2026, 8, 18))] == [
        "Starts Today"
    ]
    assert compose_soi(conn, org.id, date(2026, 8, 17)) == []


def test_an_account_whose_only_cover_is_future_gets_no_soi_sheet(conn, tmp_path):
    """The sheet-inclusion rule, on the new end of the window: omitted, not
    blank — the same rule the all-expired account already follows."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Not Yet Co", kind="client", status="active")
    placements.create(conn, org.id, "2027 Program", "2027-01-01", "2028-01-01",
                      status="bound")
    tasks.create(conn, "bind it", org_id=org.id)

    wb = load_workbook(write(conn, org.id, tmp_path / "ny.xlsx", date(2026, 8, 18)))
    assert "Schedule of Insurance" not in wb.sheetnames


def test_soi_problems_is_silent_when_every_link_is_sound(conn, tmp_path):
    """The operator line only ever says something when something is wrong —
    the same discipline withheld_note keeps."""
    from towerkit.model import dump_program

    from bookkit.services.export_open_items import soi_note, soi_problems

    org = orgs.create(conn, name="Clean Co", kind="client")
    path = tmp_path / "clean.json"
    dump_program(_one_layer_program("Clean Co", "Casualty Program"), path)
    p = placements.create(conn, org.id, "Casualty Program",
                          "2025-10-01", "2026-10-01", status="bound")
    placements.update(conn, p.id, program_path=str(path))
    placements.create(conn, org.id, "Paper Only", "2026-01-01", "2027-01-01",
                      status="bound")

    assert soi_problems(conn, org.id, date(2026, 8, 18)) == []
    assert soi_note(conn, org.id, date(2026, 8, 18)) == ""


# --- sheet 2 (assembler): Information Requests -----------------------------


def test_write_four_sheet_order_when_rfi_outstanding(conn, tmp_path):
    from bookkit.services.export_open_items import write

    client = orgs.create(conn, kind="client", name="Acme", status="active", owner="grant")
    tasks.create(conn, "Chase updated loss runs", org_id=client.id)
    placements.create(
        conn, client.id, "Acme Property 25-26", "2025-10-01", "2026-10-01",
        status="bound", total_premium=250_000_00,
    )
    live = projects_repo.create_project(conn, client.id, "Warehouse Expansion")
    projects_repo.add_need(conn, live.id, "Builder's Risk", "2026-09-01")

    market = orgs.create(conn, name="Sompo", kind="market")
    req = rfi.create_request(
        conn, client.id, "property questions", "2026-08-05", due_on="2026-08-19",
        market_org_id=market.id,
    )
    rfi.add_item(conn, req.id, "how many locations?", detail="Please **list** them all.")

    path = write(conn, client.id, tmp_path / "w.xlsx", date(2026, 8, 13))
    from openpyxl import load_workbook  # test-only import; src never imports it
    wb = load_workbook(path)
    assert wb.sheetnames == [
        "Open Items — Acme", "Information Requests", "Projects",
        "Schedule of Insurance",
    ]
    ws = wb["Information Requests"]
    # No `Type` column since 2026-08-21: it restated the ask beside it.
    assert [c.value for c in ws[1]] == ["Item", "Detail", "Needed by"]
    # No banner row either — the sheet is named Information Requests and the
    # tab says so. Row 2 is the first request's own section label.
    assert ws["A2"].value != "Items we need from you, and what you have already sent"
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "how many locations?" in values
    assert "Please list them all." in values
    assert "Sompo — property questions · due 19 Aug" in values


def test_the_response_column_appears_only_once_something_is_answered(conn, tmp_path):
    """A permanently blank Response band down a client deliverable reads as a
    form we forgot to fill in. It prints when this account has an answer to
    show, and not before."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    client = orgs.create(conn, kind="client", name="Acme", status="active", owner="grant")
    req = rfi.create_request(conn, client.id, "property questions", "2026-08-05")
    item = rfi.add_item(conn, req.id, "how many locations?")

    unanswered = load_workbook(write(conn, client.id, tmp_path / "before.xlsx", date(2026, 8, 13)))
    assert [c.value for c in unanswered["Information Requests"][1]] == [
        "Item", "Detail", "Needed by",
    ]

    rfi.update_item(conn, item.id, response="Fourteen, list attached.")

    answered = load_workbook(write(conn, client.id, tmp_path / "after.xlsx", date(2026, 8, 13)))
    ws = answered["Information Requests"]
    assert [c.value for c in ws[1]] == ["Item", "Detail", "Needed by", "Response"]
    assert "Fourteen, list attached." in [c.value for row in ws.iter_rows() for c in row]


def test_write_omits_information_requests_sheet_when_nothing_outstanding(conn, tmp_path):
    """An RFI request that exists but has no outstanding items (received or
    waived) must not add a 4th tab — omitted, not rendered blank, matching
    the Projects sheet rule."""
    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Solo Co", kind="client")
    tasks.create(conn, "one open task", org_id=org.id)
    req = rfi.create_request(conn, org.id, "closed ask", "2026-08-05")
    item = rfi.add_item(conn, req.id, "already answered")
    rfi.update_item(conn, item.id, status="received", received_on="2026-08-12")

    path = write(conn, org.id, tmp_path / "s.xlsx", date(2026, 8, 13))
    from openpyxl import load_workbook
    assert load_workbook(path).sheetnames == ["Open Items — Solo Co"]


def test_write_four_sheet_deterministic(conn, tmp_path):
    from bookkit.services.export_open_items import write

    client = orgs.create(conn, kind="client", name="Det Co", status="active")
    placements.create(conn, client.id, "Det Program", "2025-01-01", "2026-01-01")
    live = projects_repo.create_project(conn, client.id, "Det Project")
    projects_repo.add_need(conn, live.id, "GL", "2026-09-01")
    req = rfi.create_request(conn, client.id, "onboarding docs", "2026-08-05")
    rfi.add_item(conn, req.id, "audited financials", category="Financials")
    rfi.add_item(conn, req.id, "anything else")

    a = write(conn, client.id, tmp_path / "a.xlsx", date(2026, 8, 13))
    b = write(conn, client.id, tmp_path / "b.xlsx", date(2026, 8, 13))
    assert a.read_bytes() == b.read_bytes()  # four sheets, one finalize, same bytes


# --- sheet 1 / sheet 2 row-height estimate (Grant scope-add, mid-flight) ---


def test_open_items_column_widths_widened(conn, tmp_path):
    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Widths Co", kind="client")
    tasks.create(conn, "one open task", org_id=org.id)
    path = write(conn, org.id, tmp_path / "w.xlsx", date(2026, 8, 13))
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    assert ws.column_dimensions["B"].width == 50.0  # Description
    assert ws.column_dimensions["C"].width == 75.0  # Detail


def test_open_items_long_unbroken_description_gets_three_line_height(conn, tmp_path):
    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Long Co", kind="client")
    tasks.create(
        conn, "chase loss runs", org_id=org.id,
        description="x" * 101,  # ceil(101 / 50) == 3 lines at the new width
    )
    path = write(conn, org.id, tmp_path / "w.xlsx", date(2026, 8, 13))
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    # row 1 header, row 2 the "General — Long Co" section label, row 3 the
    # task. The scope line that used to occupy row 2 is gone.
    assert ws.row_dimensions[3].height >= 54.0


def test_open_items_short_row_keeps_two_line_floor(conn, tmp_path):
    from bookkit.services.export_open_items import write

    org = orgs.create(conn, name="Short Co", kind="client")
    tasks.create(conn, "short one", org_id=org.id, description="brief")
    path = write(conn, org.id, tmp_path / "w.xlsx", date(2026, 8, 13))
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    # row 1 header, row 2 the "General — Short Co" section label, row 3 the
    # task. The scope line that used to occupy row 2 is gone.
    assert ws.row_dimensions[3].height == 36.0


def test_information_requests_detail_column_matches_open_items_width(conn, tmp_path):
    from bookkit.services.export_open_items import write

    client = orgs.create(conn, kind="client", name="Family Co", status="active")
    req = rfi.create_request(conn, client.id, "onboarding docs", "2026-08-05")
    rfi.add_item(conn, req.id, "audited financials", detail="y" * 151)  # ceil(151/75)==3
    path = write(conn, client.id, tmp_path / "w.xlsx", date(2026, 8, 13))
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["Information Requests"]
    assert ws.column_dimensions["B"].width == 75.0  # Detail
    # row 1 header, row 2 the request's own section label, row 3 the item row.
    # The "Items we need from you" banner that used to occupy row 2 is gone.
    assert ws.row_dimensions[3].height >= 54.0


def test_a_nested_open_batch_joins_and_leaves_no_orphan_row(seeded) -> None:
    """ONE WRITER ACTION IS ONE UNDO UNIT, and one row in the changes list.

    db.transaction has always JOINED rather than nested and has always ignored
    an inner `batch=`, so events written by a nested service already landed on
    the outer batch. What open_batch still did was insert a row for the inner
    action anyway — a row with zero events, describing an action, offering a
    Revert, and reverting nothing.

    Found via the program cascade calling services/rfi.remove_request inside
    its own batch (2026-08-21). Asserted here rather than there because it is a
    property of the seam, not of that caller.
    """
    from bookkit.repo import batches as batches_repo
    from bookkit.repo import orgs as orgs_repo
    from bookkit.services import batches as batches_svc

    org = orgs_repo.list_orgs(seeded, kind="client")[0]
    before = len(batches_repo.recent(seeded, "0000", limit=50))

    with batches_svc.open_batch(
        seeded, source="web", tool="outer", summary="outer action",
        org_id=org.id,
    ) as outer:
        with batches_svc.open_batch(
            seeded, source="web", tool="inner", summary="inner action",
            org_id=org.id,
        ) as inner:
            orgs_repo.update(seeded, org.id, notes="touched")
            assert inner.id == outer.id, "the inner call opened its own batch"

    rows = batches_repo.recent(seeded, "0000", limit=50)
    assert len(rows) == before + 1, "a nested open_batch left an extra row"
    assert rows[0].tool == "outer", "the inner action named the undo unit"
    assert batches_repo.events_for(seeded, outer.id), "the events went nowhere"
