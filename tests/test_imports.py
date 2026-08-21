"""Import pipeline pure core: fieldspec, readers, tablemap, staging, matcher,
book mapper, committer, and the bookctl template/import verbs."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest

from bookkit.imports.fieldspec import BOOK_FIELDS, FieldSpec, write_template
from bookkit.imports.readers import read_table
from bookkit.imports.tablemap import apply_mapping, map_headers


def test_book_fields_have_unique_keys_and_aliases() -> None:
    keys = [spec.key for spec in BOOK_FIELDS]
    assert len(keys) == len(set(keys))
    all_hints = [hint for spec in BOOK_FIELDS for hint in (spec.key, *spec.aliases)]
    assert len(all_hints) == len(set(all_hints))
    assert [spec.key for spec in BOOK_FIELDS if spec.required] == ["account"]


def _fill_template(path: Path, **cells: object) -> Path:
    """Fill one data row into a written template's data sheet, the way a user
    would. Templates ship EMPTY — see the example-row tests below."""
    from openpyxl import load_workbook

    from bookkit.imports.fieldspec import DATA_SHEET

    wb = load_workbook(path)
    ws = wb[DATA_SHEET]
    headers = [cell.value for cell in ws[1]]
    ws.append([cells.get(str(header)) for header in headers])
    wb.save(path)
    return path


def test_template_round_trips_with_zero_unmapped(tmp_path: Path) -> None:
    out = _fill_template(
        write_template(BOOK_FIELDS, tmp_path / "book.xlsx"), account="Zephyr Logistics"
    )
    table = read_table(out)
    mapping = map_headers(table.headers, BOOK_FIELDS)
    assert mapping.unmapped == []
    assert mapping.fuzzy == ()  # a template must never need the fuzzy matcher
    rows = apply_mapping(table, mapping)
    assert rows[0]["account"] == "Zephyr Logistics"


def test_template_ships_no_data_rows_and_stages_nothing(tmp_path: Path, conn) -> None:
    """FINDING 1. The worked example used to be appended to the data sheet as
    a real row: fill in rows underneath it, re-import, and you created an
    "Atomic Industries" account carrying a BOUND $250,000 placement that came
    off no document and that nobody typed."""
    out = write_template(BOOK_FIELDS, tmp_path / "book.xlsx")
    table = read_table(out)
    assert table.rows == []
    staged = stage_book(conn, table, map_headers(table.headers, BOOK_FIELDS))
    assert staged.records == []
    assert "Atomic Industries" not in staged.report()


def test_template_commits_nothing_from_a_blank_template(db_path: Path) -> None:
    """The bad outcome end to end: a template committed as-is must not create
    the example account, its bound placement, or even a backup."""
    connection = db.connect(db_path)
    try:
        out = write_template(BOOK_FIELDS, db_path.parent / "book.xlsx")
        table = read_table(out)
        staged = stage_book(connection, table, map_headers(table.headers, BOOK_FIELDS))
        with pytest.raises(ValueError, match="nothing to commit"):
            commit_book(connection, staged, db_path)
        assert orgs_repo.list_orgs(connection, kind="client") == []
        # refused BEFORE the snapshot: an import that changes nothing must not
        # leave a backup behind either
        assert not (db_path.parent / "backups").exists()
    finally:
        connection.close()


def test_template_example_is_still_visible_on_its_own_sheet(tmp_path: Path) -> None:
    """Moving the example must not delete it — showing the expected shape is
    the whole point of a template."""
    from openpyxl import load_workbook

    from bookkit.imports.fieldspec import DATA_SHEET, EXAMPLE_SHEET

    out = write_template(BOOK_FIELDS, tmp_path / "book.xlsx")
    wb = load_workbook(out)
    assert wb.sheetnames == [DATA_SHEET, EXAMPLE_SHEET]
    example = wb[EXAMPLE_SHEET]
    assert "EXAMPLE ONLY" in str(example.cell(row=1, column=1).value)
    values = [cell.value for cell in example[3]]
    assert "Atomic Industries" in values and "250,000" in values


def test_example_sheet_stays_out_of_the_pipeline_when_tabs_are_reordered(
    tmp_path: Path,
) -> None:
    """Tab order decides `worksheets[0]`, and tabs get dragged — so the reader
    asks for the data sheet BY NAME."""
    from openpyxl import load_workbook

    from bookkit.imports.fieldspec import EXAMPLE_SHEET

    out = _fill_template(
        write_template(BOOK_FIELDS, tmp_path / "book.xlsx"), account="Zephyr Logistics"
    )
    wb = load_workbook(out)
    wb.move_sheet(EXAMPLE_SHEET, offset=-1)  # example dragged to the front
    wb.save(out)
    rows = read_table(out).rows
    assert [row.get("account") for row in rows] == ["Zephyr Logistics"]


def test_read_table_csv_and_xlsx_agree(tmp_path: Path) -> None:
    csv_path = tmp_path / "t.csv"
    with csv_path.open("w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Account", "Premium"])
        writer.writerow(["Atomic", "250,000"])
        writer.writerow(["", ""])  # blank row skipped
    table = read_table(csv_path)
    assert table.headers == ["Account", "Premium"]
    assert table.rows == [{"Account": "Atomic", "Premium": "250,000"}]
    assert table.source == "t.csv"
    assert len(table.sha256) == 64

    from openpyxl import Workbook

    wb = Workbook()
    wb.active.append(["Account", "Premium"])
    wb.active.append(["Atomic", "250,000"])
    xlsx_path = tmp_path / "t.xlsx"
    wb.save(xlsx_path)
    assert read_table(xlsx_path).rows == table.rows


def test_read_table_rejects_unknown_suffix(tmp_path: Path) -> None:
    path = tmp_path / "t.pdf"
    path.write_text("nope")
    with pytest.raises(ValueError, match="pdf"):
        read_table(path)


def test_map_headers_exact_alias_fuzzy_and_garbage() -> None:
    specs = (
        FieldSpec("account", True, "A", ("insured", "client")),
        FieldSpec("inception", False, "2026-10-01", ("eff date", "effective date")),
        FieldSpec("premium", False, "1", ()),
    )
    mapping = map_headers(
        ["Insured", "Efective Date", "Annual Premium Zebra Wombat", "Premium"], specs
    )
    assert mapping.assigned["Insured"] == "account"
    assert mapping.assigned["Efective Date"] == "inception"  # fuzzy over typo
    assert mapping.assigned["Premium"] == "premium"
    assert "Annual Premium Zebra Wombat" in mapping.unmapped


def test_map_headers_first_claim_wins() -> None:
    specs = (FieldSpec("account", True, "A", ("insured",)),)
    mapping = map_headers(["Account", "Insured"], specs)
    assert mapping.assigned == {"Account": "account"}
    assert mapping.unmapped == ["Insured"]


# --- staging --------------------------------------------------------------------

from bookkit.imports.staging import (  # noqa: E402
    Issue,
    Severity,
    StagedImport,
    StagedRecord,
)


def _staged(records: list[StagedRecord], unmapped: list[str] | None = None) -> StagedImport:
    return StagedImport("book.xlsx", "ab" * 32, records, unmapped or [])


def test_staged_import_gates_on_errors() -> None:
    clean = StagedRecord("account", "Atomic", {"name": "Atomic"}, source_row=1)
    broken = StagedRecord(
        "placement", "Atomic/Property", {}, source_row=2,
        issues=[Issue(Severity.ERROR, "premium", "cannot parse 'banana'")],
    )
    assert _staged([clean]).ok
    staged = _staged([clean, broken])
    assert not staged.ok
    assert [issue.field for issue in staged.errors] == ["premium"]


def test_staged_report_shows_counts_issues_and_unmapped() -> None:
    records = [
        StagedRecord("account", "Atomic", {}, source_row=1),
        StagedRecord("account", "Borealis", {}, source_row=2, action="update", target_id="x"),
        StagedRecord(
            "contact", "Atomic/Rosa", {}, source_row=1,
            issues=[Issue(Severity.WARNING, "phone", "no digits")],
        ),
    ]
    report = _staged(records, unmapped=["Wibble Col"]).report()
    assert "account: 1 create, 1 update" in report
    assert "contact: 1 create" in report
    assert "⚠ phone: no digits" in report
    assert "Wibble Col" in report
    assert "book.xlsx" in report


# --- committer --------------------------------------------------------------------

from bookkit import db  # noqa: E402
from bookkit.imports.commit import commit_book  # noqa: E402


def _staged_book(conn, rows):
    table = RawTable("book.csv", "cd" * 32, _BOOK_HEADERS, rows)
    return stage_book(conn, table, map_headers(_BOOK_HEADERS, BOOK_FIELDS))


_GOOD_ROW = {
    "account": "Atomic Industries", "status": "active",
    "contact_name": "Rosa Silva", "contact_email": "rosa@atomic.example.com",
    "program": "Property", "inception": "2026-10-01", "expiry": "2027-10-01",
    "premium": "250,000", "commission": "15%",
}


def test_commit_book_end_to_end_with_backup_and_provenance(db_path: Path) -> None:
    connection = db.connect(db_path)
    try:
        staged = _staged_book(connection, [dict(_GOOD_ROW)])
        result = commit_book(connection, staged, db_path)
        assert result.created == {"account": 1, "contact": 1, "placement": 1}
        assert result.backup.exists() and result.backup.parent.name == "backups"

        org = orgs_repo.find_by_name(connection, "Atomic Industries")
        assert org is not None and org.status == "active"
        [contact] = contacts_repo.for_org(connection, org.id)
        assert contact.email == "rosa@atomic.example.com"
        [placement] = placements_repo.for_org(connection, org.id)
        assert placement.total_premium == 25_000_000
        note = connection.execute(
            "SELECT note FROM event_log WHERE note LIKE '%book.csv%'"
        ).fetchone()
        assert note is not None and "cd" * 32 in note["note"]
    finally:
        connection.close()


def test_commit_book_reimport_updates_instead_of_duplicating(db_path: Path) -> None:
    connection = db.connect(db_path)
    try:
        commit_book(connection, _staged_book(connection, [dict(_GOOD_ROW)]), db_path)
        staged = _staged_book(connection, [{**_GOOD_ROW, "premium": "260,000"}])
        result = commit_book(connection, staged, db_path)
        assert result.created == {}
        assert result.updated["placement"] == 1
        assert len(orgs_repo.list_orgs(connection, kind="client")) == 1
        [placement] = placements_repo.for_org(
            connection, orgs_repo.find_by_name(connection, "Atomic Industries").id
        )
        assert placement.total_premium == 26_000_000
    finally:
        connection.close()


def test_commit_book_rolls_back_everything_on_midway_failure(
    db_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from bookkit.imports import commit as commit_mod

    connection = db.connect(db_path)
    try:
        staged = _staged_book(connection, [dict(_GOOD_ROW)])

        def boom(*args: object, **kwargs: object) -> None:
            raise RuntimeError("disk full")

        monkeypatch.setattr(commit_mod.placements, "create", boom)
        with pytest.raises(RuntimeError):
            commit_book(connection, staged, db_path)
        # org and contact were created BEFORE the failure — a real transaction
        # must erase them too
        assert orgs_repo.list_orgs(connection, kind="client") == []
    finally:
        connection.close()


def test_stage_book_fuzzy_match_keeps_canonical_name(conn) -> None:
    orgs_repo.create(conn, kind="client", name="Atomic Industries")
    table, mapping = _book_table([{"account": "Atomic Industries Inc"}])
    staged = stage_book(conn, table, mapping)
    record = staged.records[0]
    assert record.action == "update"
    assert "name" not in record.fields  # never rename the curated account


def test_commit_book_name_only_contacts_do_not_duplicate(db_path: Path) -> None:
    connection = db.connect(db_path)
    try:
        rows = [{"account": "Atomic", "contact_name": "Ken Ito"}]
        commit_book(connection, _staged_book(connection, list(rows)), db_path)
        staged = _staged_book(connection, list(rows))
        contact = next(r for r in staged.records if r.kind == "contact")
        assert contact.action == "update"
        commit_book(connection, staged, db_path)
        org = orgs_repo.find_by_name(connection, "Atomic")
        assert len(contacts_repo.for_org(connection, org.id)) == 1
    finally:
        connection.close()


def test_commit_book_refuses_errors_and_writes_nothing(db_path: Path) -> None:
    connection = db.connect(db_path)
    try:
        staged = _staged_book(connection, [{**_GOOD_ROW, "premium": "banana"}])
        with pytest.raises(ValueError, match="error"):
            commit_book(connection, staged, db_path)
        assert orgs_repo.list_orgs(connection, kind="client") == []
    finally:
        connection.close()


# --- contact paste -------------------------------------------------------------------

from bookkit.imports.commit import commit_contact_paste  # noqa: E402
from bookkit.imports.mappers.contact_paste import stage_contact_paste  # noqa: E402

_SIGNATURE = """Rosa Silva
Director of Risk Management
Atomic Industries, Inc.
rosa.silva@atomic.example.com | (312) 555-0142
https://www.linkedin.com/in/rosasilva
"""


def test_stage_contact_paste_parses_signature(conn) -> None:
    org = orgs_repo.create(conn, kind="client", name="Atomic Industries")
    staged = stage_contact_paste(conn, _SIGNATURE, org.id, org.name)
    contact = next(r for r in staged.records if r.kind == "contact")
    assert contact.fields["first_name"] == "Rosa"
    assert contact.fields["last_name"] == "Silva"
    assert contact.fields["title"] == "Director of Risk Management"
    assert contact.fields["email"] == "rosa.silva@atomic.example.com"
    assert contact.fields["phone"] == "(312) 555-0142"
    assert str(contact.fields["linkedin"]).endswith("/in/rosasilva")
    interaction = next(r for r in staged.records if r.kind == "interaction")
    assert interaction.fields["body"] == _SIGNATURE
    assert staged.ok


def test_stage_contact_paste_matches_existing_by_email(conn) -> None:
    org = orgs_repo.create(conn, kind="client", name="Atomic")
    existing = contacts_repo.create(
        conn, org.id, first_name="Rosa", last_name="Silva",
        email="rosa.silva@atomic.example.com",
    )
    staged = stage_contact_paste(conn, _SIGNATURE, org.id, org.name)
    contact = next(r for r in staged.records if r.kind == "contact")
    assert contact.action == "update" and contact.target_id == existing.id


def test_commit_contact_paste_creates_contact_and_interaction(db_path: Path) -> None:
    connection = db.connect(db_path)
    try:
        org = orgs_repo.create(connection, kind="client", name="Atomic")
        staged = stage_contact_paste(connection, _SIGNATURE, org.id, org.name)
        commit_contact_paste(connection, staged, org.id, db_path)
        [contact] = contacts_repo.for_org(connection, org.id)
        assert contact.name == "Rosa Silva"
        from bookkit.repo import interactions as interactions_repo

        [interaction] = interactions_repo.for_org(connection, org.id)
        assert interaction.type == "note"
    finally:
        connection.close()


def test_stage_contact_paste_ignores_dates_as_phones(conn) -> None:
    org = orgs_repo.create(conn, kind="client", name="Atomic Industries")
    text = "Meeting 2026-08-11 10:30\n\n" + _SIGNATURE
    staged = stage_contact_paste(conn, text, org.id, org.name)
    contact = next(r for r in staged.records if r.kind == "contact")
    assert contact.fields["phone"] == "(312) 555-0142"  # not a formatted date


def test_stage_contact_paste_garbage_still_stages_note(conn) -> None:
    org = orgs_repo.create(conn, kind="client", name="Atomic")
    staged = stage_contact_paste(conn, "12 monkeys\n9931", org.id, org.name)
    kinds = [r.kind for r in staged.records]
    assert "interaction" in kinds
    assert staged.ok  # warnings only — a note is always worth keeping


# --- program from schedule ------------------------------------------------------------

from bookkit.imports.commit import commit_program  # noqa: E402
from bookkit.imports.mappers.program_paste import stage_program  # noqa: E402
from bookkit.repo import aliases as aliases_repo  # noqa: E402

_TOWER = "Primary 10M — Chubb 100% — 250,000\n15M xs 10M — AXA XL 60%, Sompo 40% — 180k\n"


def test_stage_program_flags_known_and_unknown_carriers(conn) -> None:
    market = orgs_repo.create(conn, kind="market", name="Chubb")
    aliases_repo.set_alias(conn, "Chubb", market.id)
    staged, draft = stage_program(
        conn, _TOWER, "Atomic Industries", "Property", "2026-10-01", "2027-10-01"
    )
    assert staged.ok
    carriers = {r.key: r for r in staged.records if r.kind == "carrier"}
    assert carriers["Chubb"].action == "update"  # resolves to an existing market
    assert any(i.field == "carrier" for i in carriers["AXA XL"].issues)  # warning only
    assert draft.period is not None and draft.period.start.isoformat() == "2026-10-01"


def test_commit_program_writes_file_links_and_projects(db_path: Path, tmp_path: Path) -> None:
    connection = db.connect(db_path)
    try:
        org = orgs_repo.create(connection, kind="client", name="Atomic Industries")
        placement = placements_repo.create(
            connection, org.id, "Property", "2026-10-01", "2027-10-01"
        )
        staged, draft = stage_program(
            connection, _TOWER, org.name, "Property", "2026-10-01", "2027-10-01"
        )
        dest = tmp_path / "programs" / "atomic-property-2026.json"
        path, diags = commit_program(
            connection, staged, draft, placement.id, dest, db_path
        )
        assert path == dest and dest.exists()
        assert diags.ok
        refreshed = placements_repo.get(connection, placement.id)
        assert refreshed.program_path == str(dest)
        from bookkit import sync

        layers = sync.layer_details(connection, placement.id)
        assert {layer["name"] for layer in layers} == {"Primary", "$15M xs $10M"}
    finally:
        connection.close()


def test_commit_program_failure_cleans_up_file_and_rolls_back(
    db_path: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from bookkit import sync as sync_mod

    connection = db.connect(db_path)
    try:
        org = orgs_repo.create(connection, kind="client", name="Atomic Industries")
        placement = placements_repo.create(
            connection, org.id, "Property", "2026-10-01", "2027-10-01"
        )
        staged, draft = stage_program(
            connection, _TOWER, org.name, "Property", "2026-10-01", "2027-10-01"
        )

        def boom(*args: object, **kwargs: object) -> None:
            raise RuntimeError("projection exploded")

        monkeypatch.setattr(sync_mod, "project", boom)
        dest = tmp_path / "programs" / "atomic.json"
        with pytest.raises(RuntimeError):
            commit_program(connection, staged, draft, placement.id, dest, db_path)
        assert not dest.exists()  # no orphan blocking the retry
        assert placements_repo.get(connection, placement.id).program_path is None
    finally:
        connection.close()


def test_commit_program_refuses_already_linked(db_path: Path, tmp_path: Path) -> None:
    connection = db.connect(db_path)
    try:
        org = orgs_repo.create(connection, kind="client", name="Atomic")
        placement = placements_repo.create(
            connection, org.id, "Property", "2026-10-01", "2027-10-01",
            program_path="/somewhere/else.json",
        )
        staged, draft = stage_program(
            connection, _TOWER, org.name, "Property", "2026-10-01", "2027-10-01"
        )
        path, diags = commit_program(
            connection, staged, draft, placement.id, tmp_path / "x.json", db_path
        )
        assert path is None and not diags.ok
    finally:
        connection.close()


# --- renewal updates -------------------------------------------------------------------

from bookkit.imports.commit import commit_renewal  # noqa: E402
from bookkit.imports.mappers.renewal_paste import stage_renewal  # noqa: E402


def _linked_placement(connection, tmp_path: Path):
    org = orgs_repo.create(connection, kind="client", name="Atomic Industries")
    placement = placements_repo.create(
        connection, org.id, "Property", "2026-10-01", "2027-10-01"
    )
    staged, draft = stage_program(
        connection, _TOWER, org.name, "Property", "2026-10-01", "2027-10-01"
    )
    dest = tmp_path / "programs" / "atomic-property-2026.json"
    commit_program(connection, staged, draft, placement.id, dest, tmp_path / "unused.db")
    return placements_repo.get(connection, placement.id)


def test_stage_renewal_diffs_and_matches_by_attachment(
    db_path: Path, tmp_path: Path
) -> None:
    connection = db.connect(db_path)
    try:
        placement = _linked_placement(connection, tmp_path)
        staged = stage_renewal(
            connection, placement.id,
            "Primary 10M — Chubb 100% — 275,000\n20M xs 10M — Sompo — 200k\n",
        )
        assert staged.ok
        primary = next(r for r in staged.records if r.key.endswith("Primary"))
        assert primary.action == "update"
        assert primary.fields["premium_cents"] == 27_500_000
        assert "250,000" in str(primary.fields["diff"]) or "$250,000" in str(
            primary.fields["diff"]
        )
        # pasted '20M xs 10M' matches the existing '$15M xs $10M' by ATTACHMENT
        # even though the changed limit changes its auto-generated name
        excess = next(r for r in staged.records if "15M" in r.key)
        assert excess.action == "update"
        assert excess.fields["limit_cents"] == 2_000_000_000
        assert excess.fields["premium_cents"] == 20_000_000
        assert not [r for r in staged.records if r.action == "skip"]
    finally:
        connection.close()


def test_stage_renewal_matches_custom_layer_names_by_attachment(
    db_path: Path, tmp_path: Path
) -> None:
    from towerkit.model import dump_program, load_program

    connection = db.connect(db_path)
    try:
        placement = _linked_placement(connection, tmp_path)
        path = Path(placement.program_path)
        program = load_program(path)  # rename like a real towerkit-authored file
        program.layers[0].name = "Primary GL"
        program.layers[1].name = "1st Excess"
        dump_program(program, path)
        staged = stage_renewal(connection, placement.id, "Primary 10M — Chubb — 300k\n")
        assert staged.ok
        updated = next(r for r in staged.records if r.action == "update")
        assert updated.fields["layer_name"] == "Primary GL"
    finally:
        connection.close()


def test_stage_renewal_nothing_matched_is_error(db_path: Path, tmp_path: Path) -> None:
    connection = db.connect(db_path)
    try:
        placement = _linked_placement(connection, tmp_path)
        staged = stage_renewal(connection, placement.id, "5M xs 50M — Chubb\n")
        assert not staged.ok  # committing would silently drop the whole paste
    finally:
        connection.close()


def test_commit_renewal_surfaces_a_gap_warning(db_path: Path, tmp_path: Path) -> None:
    """Shrinking Primary to 8M opens a gap under the 15M xs 10M layer.
    line-gap is a WARNING, not a refusal (2026-08-21), so the delta COMMITS —
    and the gap must still be visible, now in diags.warnings rather than
    diags.errors (it used to refuse the whole delta; that assertion inverts,
    same as the sync-level and web-level remove tests)."""
    connection = db.connect(db_path)
    try:
        placement = _linked_placement(connection, tmp_path)
        staged = stage_renewal(connection, placement.id, "Primary 8M — Chubb — 200k\n")
        new_id, diags = commit_renewal(connection, staged, placement.id, db_path)
        assert new_id is not None
        assert diags.ok, [d.message for d in diags.errors]
        assert any(d.code == "line-gap" for d in diags.warnings)
    finally:
        connection.close()


def test_commit_renewal_creates_next_period_with_new_premium(
    db_path: Path, tmp_path: Path
) -> None:
    connection = db.connect(db_path)
    try:
        placement = _linked_placement(connection, tmp_path)
        staged = stage_renewal(
            connection, placement.id, "Primary 10M — Chubb 100% — 275,000\n"
        )
        new_id, diags = commit_renewal(connection, staged, placement.id, db_path)
        assert new_id is not None and diags.ok
        renewed = placements_repo.get(connection, new_id)
        assert renewed.period_from == "2027-10-01"
        from bookkit import sync

        primary = next(
            layer for layer in sync.layer_details(connection, new_id)
            if layer["name"] == "Primary"
        )
        assert primary["premium_cents"] == 27_500_000
    finally:
        connection.close()


def test_stage_renewal_unlinked_placement_is_error(conn) -> None:
    org = orgs_repo.create(conn, kind="client", name="Atomic")
    placement = placements_repo.create(conn, org.id, "Property", "2026-10-01", "2027-10-01")
    staged = stage_renewal(conn, placement.id, "Primary 10M — Chubb\n")
    assert not staged.ok


# --- cli ---------------------------------------------------------------------------

from bookkit.cli import main as cli_main  # noqa: E402


def test_cli_template_and_dry_run_import(tmp_path: Path, capsys) -> None:
    dbfile = tmp_path / "book.db"
    template = tmp_path / "book.xlsx"
    assert cli_main(["--db", str(dbfile), "template", "book", str(template)]) == 0
    _fill_template(template, account="Zephyr Logistics")
    code = cli_main(["--db", str(dbfile), "import", "book", str(template), "--dry-run"])
    out = capsys.readouterr().out
    assert code == 0
    assert "account: 1 create" in out and "OK to commit" in out


def test_cli_dry_run_of_an_untouched_template_is_not_success(
    tmp_path: Path, capsys
) -> None:
    """FINDING 8. Zero records is not green, on every surface that reports."""
    dbfile = tmp_path / "book.db"
    template = tmp_path / "book.xlsx"
    cli_main(["--db", str(dbfile), "template", "book", str(template)])
    code = cli_main(["--db", str(dbfile), "import", "book", str(template), "--dry-run"])
    out = capsys.readouterr().out
    assert code == 1
    assert "NOTHING TO COMMIT" in out and "OK to commit" not in out


def test_cli_import_without_dry_run_points_at_tui(tmp_path: Path, capsys) -> None:
    dbfile = tmp_path / "book.db"
    template = tmp_path / "book.xlsx"
    cli_main(["--db", str(dbfile), "template", "book", str(template)])
    _fill_template(template, account="Zephyr Logistics")
    code = cli_main(["--db", str(dbfile), "import", "book", str(template)])
    assert code == 2
    assert "TUI" in capsys.readouterr().out


def test_cli_import_bad_file_exits_one(tmp_path: Path, capsys) -> None:
    dbfile = tmp_path / "book.db"
    bad = tmp_path / "bad.csv"
    bad.write_text("account,premium\nAtomic,banana\n")
    code = cli_main(["--db", str(dbfile), "import", "book", str(bad), "--dry-run"])
    assert code == 1
    assert "banana" in capsys.readouterr().out


# --- TUI pilots -----------------------------------------------------------------------


async def test_import_screen_book_commit(tmp_path: Path) -> None:
    from bookkit.tui.app import BookkitApp
    from bookkit.tui.screens.import_screen import ImportScreen

    dbfile = tmp_path / "tui.db"
    db.connect(dbfile).close()
    template = tmp_path / "book.xlsx"
    from bookkit.imports.fieldspec import write_template

    _fill_template(
        write_template(BOOK_FIELDS, template),
        account="Atomic Industries", program="Property",
        inception="2026-10-01", expiry="2027-10-01", premium="250,000",
        commission="15%",
    )
    app = BookkitApp(dbfile)
    async with app.run_test(size=(120, 40)) as pilot:
        await pilot.press("t")  # navigator is home; the book importer is on Today
        await pilot.pause()
        await pilot.press("i")
        assert isinstance(app.screen, ImportScreen)
        path_input = app.screen.query_one("#import-path")
        path_input.value = str(template)
        await pilot.press("enter")
        await pilot.pause()
        assert app.screen._staged is not None and app.screen._staged.ok
        await pilot.press("ctrl+s")
        await pilot.pause()
        assert orgs_repo.find_by_name(app.conn, "Atomic Industries") is not None


async def test_account_import_chooser_contact_paste(tmp_path: Path) -> None:
    from bookkit.tui.app import BookkitApp
    from bookkit.tui.widgets.paste_import import ImportChooser, PasteImportModal

    dbfile = tmp_path / "tui.db"
    connection = db.connect(dbfile)
    org = orgs_repo.create(connection, kind="client", name="Atomic Industries")
    connection.close()
    app = BookkitApp(dbfile)
    async with app.run_test(size=(120, 40)) as pilot:
        app.open_account(org.id)
        await pilot.pause()
        # I, not i: i is "edit in cell" and InlineTable claims it at widget
        # level, so a screen-level i changed meaning under the cursor
        await pilot.press("I")
        await pilot.pause()
        assert isinstance(app.screen, ImportChooser)
        await pilot.press("enter")  # first option: contact paste
        await pilot.pause()
        assert isinstance(app.screen, PasteImportModal)
        app.screen.query_one("#paste-text").text = _SIGNATURE
        await pilot.press("ctrl+r")
        await pilot.pause()
        assert app.screen._staged is not None and app.screen._staged.ok
        # edit AFTER previewing: commit must re-stage, never use the stale parse
        app.screen.query_one("#paste-text").text = _SIGNATURE.replace(
            "rosa.silva@", "r.silva@"
        )
        await pilot.press("ctrl+s")
        await pilot.pause()
        assert contacts_repo.for_org(app.conn, org.id) == []  # not committed yet
        await pilot.press("ctrl+s")
        await pilot.pause()
        [contact] = contacts_repo.for_org(app.conn, org.id)
        assert contact.name == "Rosa Silva"
        assert contact.email == "r.silva@atomic.example.com"  # the EDITED text won


# --- audit E: the findings, one test each ------------------------------------------


def _percent_book(path: Path, commission: object, number_format: str | None) -> Path:
    """A one-row book with a commission cell written the way Excel writes it."""
    from openpyxl import Workbook

    from bookkit.imports.fieldspec import DATA_SHEET

    wb = Workbook()
    ws = wb.active
    ws.title = DATA_SHEET
    ws.append(
        ["account", "program", "inception", "expiry", "premium", "commission"]
    )
    ws.append(
        ["Atomic Industries", "Property", "2026-10-01", "2027-10-01", "250,000",
         commission]
    )
    if number_format is not None:
        ws.cell(row=2, column=6).number_format = number_format
    wb.save(path)
    return path


def test_percent_formatted_commission_is_refused_not_reinterpreted(
    conn, tmp_path: Path
) -> None:
    """FINDING 2. A percent-formatted 15% cell reaches the reader as the float
    0.15, and the share parser reads a bare number as a percent — so it landed
    as 15 bps (0.15%), a silent 100x understatement. It must refuse, naming
    both readings, and it must NOT auto-correct: 'fix' it upward and a real
    0.5% fee becomes 50%."""
    path = _percent_book(tmp_path / "book.xlsx", 0.15, "0%")
    table = read_table(path)
    assert table.rows[0]["commission"] == 0.15  # what the file really hands over
    staged = stage_book(conn, table, map_headers(table.headers, BOOK_FIELDS))
    placement = next(r for r in staged.records if r.kind == "placement")
    assert "commission_bps" not in placement.fields  # neither 15 nor 1500
    assert not staged.ok
    [issue] = [i for i in placement.issues if i.field == "commission"]
    assert "0.15%" in issue.message and "15%" in issue.message
    assert "row 1" in issue.message


def test_percent_formatted_commission_never_reaches_the_database(
    db_path: Path,
) -> None:
    """The bad outcome, end to end: 0.15 must not commit as 15 bps."""
    connection = db.connect(db_path)
    try:
        path = _percent_book(db_path.parent / "book.xlsx", 0.15, "0%")
        table = read_table(path)
        staged = stage_book(connection, table, map_headers(table.headers, BOOK_FIELDS))
        with pytest.raises(ValueError, match="commit refused"):
            commit_book(connection, staged, db_path)
        assert orgs_repo.list_orgs(connection, kind="client") == []
    finally:
        connection.close()


def test_unambiguous_commission_cells_still_parse(conn, tmp_path: Path) -> None:
    """The refusal is scoped to the collision band: '15%', '15' and '15.0' are
    all 1500 bps and must keep working."""
    for cell in ("15%", "15", 15, 15.0, "0.15%"):
        path = _percent_book(tmp_path / f"book-{cell}.xlsx", cell, None)
        table = read_table(path)
        staged = stage_book(conn, table, map_headers(table.headers, BOOK_FIELDS))
        placement = next(r for r in staged.records if r.kind == "placement")
        expected = 15 if cell == "0.15%" else 1500
        assert placement.fields["commission_bps"] == expected, cell
        assert staged.ok, cell


def test_reimport_that_corrects_a_renewal_date_actually_moves_it(
    db_path: Path,
) -> None:
    """FINDING 3. Matching is period-OVERLAP based, so a corrected expiry
    still lands on the existing placement — and the committer dropped
    period_from/period_to from the update with no Issue anywhere, so
    "updated" meant nothing changed."""
    connection = db.connect(db_path)
    try:
        commit_book(connection, _staged_book(connection, [dict(_GOOD_ROW)]), db_path)
        staged = _staged_book(
            connection, [{**_GOOD_ROW, "expiry": "2027-12-31"}]
        )
        placement_record = next(r for r in staged.records if r.kind == "placement")
        assert placement_record.action == "update"
        assert any(
            "2027-12-31" in issue.message for issue in placement_record.issues
        ), "the preview must say the period moves"
        commit_book(connection, staged, db_path)
        org = orgs_repo.find_by_name(connection, "Atomic Industries")
        [placement] = placements_repo.for_org(connection, org.id)
        assert placement.period_to == "2027-12-31"
    finally:
        connection.close()


def test_reimport_keeps_the_period_a_program_file_owns_and_says_so(
    db_path: Path, tmp_path: Path
) -> None:
    """The other half of the decision: towerkit files are the authority for a
    linked placement's period, so the import keeps its hands off — but says
    it kept them off, which is the part that was missing."""
    connection = db.connect(db_path)
    try:
        placement = _linked_placement(connection, tmp_path)
        staged = _staged_book(connection, [{
            "account": "Atomic Industries", "program": "Property",
            "inception": "2026-10-01", "expiry": "2027-12-31",
            "premium": "300,000",
        }])
        record = next(r for r in staged.records if r.kind == "placement")
        assert record.action == "update"
        assert "period_to" not in record.fields
        assert any("NOT changed" in issue.message for issue in record.issues)
        commit_book(connection, staged, db_path)
        after = placements_repo.get(connection, placement.id)
        assert after.period_to == "2027-10-01"  # the file still owns it
        assert after.total_premium == 30_000_000  # everything else did update
    finally:
        connection.close()


def test_preview_renders_every_header_mapping_including_the_fuzzy_ones(
    conn,
) -> None:
    """FINDING 4. Every fuzzy match at threshold 85 was invisible: nobody
    could see that 'Expiration Date' had become `expiry`."""
    headers = ["Insured", "Efective Date", "Wibble"]
    table = RawTable("book.csv", "ab" * 32, headers, [{"Insured": "Atomic"}])
    mapping = map_headers(headers, BOOK_FIELDS)
    assert mapping.fuzzy == ("Efective Date",)  # matched by fuzz over a typo
    report = stage_book(conn, table, mapping).report()
    assert "'Insured' → account" in report
    assert "'Efective Date' → inception" in report
    assert "fuzzy match" in report
    assert "Wibble" in report  # unmapped is still reported too
    # and the exact match is not slandered as a guess
    insured_line = next(
        line for line in report.splitlines() if "'Insured'" in line
    )
    assert "fuzzy" not in insured_line


def test_verdict_names_the_first_offending_field_and_row(conn) -> None:
    """FINDING 6. '3 error(s)' with no field, row or fix is half a message."""
    table, mapping = _book_table([
        {"account": "Atomic Industries"},
        {"account": "Borealis Foods", "program": "Property",
         "inception": "2026-10-01", "expiry": "2027-10-01", "commission": "0.15"},
    ])
    staged = stage_book(conn, table, mapping)
    verdict = staged.verdict()
    assert "ERRORS — cannot commit" in verdict
    assert "row 2" in verdict and "commission" in verdict
    assert staged.first_error_text() is not None


def test_empty_but_readable_sheet_is_not_green_and_takes_no_backup(
    db_path: Path,
) -> None:
    """FINDING 8. records == [] gave ok is True gave a green 'OK to commit ·
    0 record(s)': success styling over a neutral empty state. The user
    committed, a backup was taken, and nothing happened."""
    connection = db.connect(db_path)
    try:
        empty = db_path.parent / "empty.csv"
        empty.write_text("account,premium\n")
        table = read_table(empty)
        staged = stage_book(connection, table, map_headers(table.headers, BOOK_FIELDS))
        assert staged.ok  # no errors — that was never the question
        assert staged.empty and not staged.committable
        assert "NOTHING TO COMMIT" in staged.verdict()
        assert "OK to commit" not in staged.verdict()
        with pytest.raises(ValueError, match="nothing to commit"):
            commit_book(connection, staged, db_path)
        assert not (db_path.parent / "backups").exists()
    finally:
        connection.close()


def test_renewal_paste_names_the_premium_it_would_carry_forward(
    db_path: Path, tmp_path: Path
) -> None:
    """FINDING 9a. A layer whose premium the parser missed produced NO staged
    record at all, so it renewed carrying the EXPIRING premium with nothing
    on screen."""
    connection = db.connect(db_path)
    try:
        placement = _linked_placement(connection, tmp_path)
        # limit and carrier restated, premium absent
        staged = stage_renewal(connection, placement.id, "Primary 10M — Chubb 100%\n")
        primary = next(r for r in staged.records if r.key.endswith("Primary"))
        messages = " ".join(issue.message for issue in primary.issues)
        assert "premium" in messages
        assert "$250,000" in messages  # the figure it will carry forward
        assert staged.ok  # visible, not blocking: the layer may truly be flat
    finally:
        connection.close()


def test_program_paste_says_the_period_came_off_the_placement(conn) -> None:
    """FINDING 9b. A period is a date off the schedule; borrowing the
    placement's silently is the prefill nobody checks."""
    staged, _draft = stage_program(
        conn, _TOWER, "Atomic Industries", "Property", "2026-10-01", "2027-10-01"
    )
    program = next(r for r in staged.records if r.kind == "program")
    [issue] = [i for i in program.issues if i.field == "period"]
    assert "2026-10-01" in issue.message and "2027-10-01" in issue.message
    assert staged.ok


def test_contact_paste_says_the_note_is_dated_today(conn) -> None:
    """FINDING 9c. `occurred_on` was stamped today on a three-week-old pasted
    thread, with no field to change it and nothing on screen saying so."""
    from datetime import date

    org = orgs_repo.create(conn, kind="client", name="Atomic Industries")
    staged = stage_contact_paste(conn, _SIGNATURE, org.id, org.name)
    note = next(r for r in staged.records if r.kind == "interaction")
    [issue] = [i for i in note.issues if i.field == "occurred_on"]
    assert date.today().isoformat() in issue.message

    dated = "Sent: 2026-07-30\n\n" + _SIGNATURE
    staged = stage_contact_paste(conn, dated, org.id, org.name)
    note = next(r for r in staged.records if r.kind == "interaction")
    [issue] = [i for i in note.issues if i.field == "occurred_on"]
    assert "2026-07-30" in issue.message  # named, never silently adopted
    assert note.fields["occurred_on"] == date.today().isoformat()


async def test_import_screen_preview_shows_parsed_values(tmp_path: Path) -> None:
    """FINDING 5. The file preview called report() without verbose, so it
    showed counts by kind and action and NOT ONE parsed field value — you
    could commit hundreds of accounts having seen none of them. The paste
    flows always showed values; the file flow now does too."""
    from bookkit.tui.app import BookkitApp

    dbfile = tmp_path / "tui.db"
    db.connect(dbfile).close()
    template = _fill_template(
        write_template(BOOK_FIELDS, tmp_path / "book.xlsx"),
        account="Zephyr Logistics", program="Property",
        inception="2026-10-01", expiry="2027-10-01", premium="250,000",
    )
    app = BookkitApp(dbfile)
    async with app.run_test(size=(120, 40)) as pilot:
        await pilot.press("t")
        await pilot.pause()
        await pilot.press("i")
        app.screen.query_one("#import-path").value = str(template)
        await pilot.press("enter")
        await pilot.pause()
        body = str(app.screen.query_one("#import-preview-body").render())
        assert "Zephyr Logistics" in body
        assert "total_premium: 25000000" in body
        assert "period_to: 2027-10-01" in body
        assert "'account' → account" in body  # the mapping, finding 4
        verdict = str(app.screen.query_one("#import-verdict").render())
        assert "OK to commit" in verdict


async def test_import_screen_stale_green_verdict_never_outlives_its_staging(
    tmp_path: Path,
) -> None:
    """FINDING 7. The verdict is a persistent Static, not a toast. The sha256
    guard set `_staged = None` and left the previous run's green "OK to
    commit" standing, so the visible go/no-go contradicted the real one."""
    from bookkit.tui.app import BookkitApp

    dbfile = tmp_path / "tui.db"
    db.connect(dbfile).close()
    template = _fill_template(
        write_template(BOOK_FIELDS, tmp_path / "book.xlsx"), account="Zephyr Logistics"
    )
    app = BookkitApp(dbfile)
    async with app.run_test(size=(120, 40)) as pilot:
        await pilot.press("t")
        await pilot.pause()
        await pilot.press("i")
        app.screen.query_one("#import-path").value = str(template)
        await pilot.press("enter")
        await pilot.pause()
        assert "OK to commit" in str(app.screen.query_one("#import-verdict").render())

        _fill_template(template, account="Borealis Foods")  # edited under us
        await pilot.press("ctrl+s")
        await pilot.pause()
        verdict = str(app.screen.query_one("#import-verdict").render())
        assert app.screen._staged is None
        assert "OK to commit" not in verdict
        assert "re-preview" in verdict

        # and a re-stage that cannot even read the file clears it too
        app.screen.query_one("#import-path").value = str(tmp_path / "nope.xlsx")
        await pilot.press("enter")
        await pilot.pause()
        verdict = str(app.screen.query_one("#import-verdict").render())
        assert "OK to commit" not in verdict and "cannot read" in verdict


async def test_paste_modal_failed_restage_drops_the_previous_parse(
    tmp_path: Path,
) -> None:
    """FINDING 7, the paste half — and worse than a stale line: staging into
    `self._staged` inside the try meant a failed re-stage left the OLD parse
    in place under a green verdict, and ctrl+s would have committed it."""
    from bookkit.imports.staging import StagedImport, StagedRecord
    from bookkit.tui.app import BookkitApp
    from bookkit.tui.widgets.paste_import import PasteImportModal

    dbfile = tmp_path / "tui.db"
    db.connect(dbfile).close()
    seen: list[str] = []

    def stage(text: str) -> StagedImport:
        seen.append(text)
        if len(seen) == 1:
            return StagedImport(
                "paste", "", [StagedRecord("contact", "Rosa", {}, source_row=1)], []
            )
        raise ValueError("unparseable paste")

    app = BookkitApp(dbfile)
    async with app.run_test(size=(120, 40)) as pilot:
        app.push_screen(PasteImportModal("paste test", stage, lambda staged: "done"))
        await pilot.pause()
        screen = app.screen
        screen.query_one("#paste-text").text = "first"
        await pilot.press("ctrl+r")
        await pilot.pause()
        assert "OK to commit" in str(screen.query_one("#paste-verdict").render())

        screen.query_one("#paste-text").text = "second"
        await pilot.press("ctrl+r")
        await pilot.pause()
        verdict = str(screen.query_one("#paste-verdict").render())
        assert screen._staged is None  # the old parse is gone, not committable
        assert "OK to commit" not in verdict
        assert "unparseable paste" in verdict


def test_imports_package_contains_no_sql() -> None:
    import re

    root = Path(__file__).parent.parent / "src" / "bookkit" / "imports"
    pattern = re.compile(r"\b(SELECT|INSERT INTO|UPDATE \w+ SET|DELETE FROM)\b")
    offenders = [
        path.name
        for path in root.rglob("*.py")
        if pattern.search(path.read_text(encoding="utf-8"))
    ]
    assert offenders == [], f"raw SQL in imports/: {offenders}"


# --- matcher --------------------------------------------------------------------

from bookkit.imports.matcher import (  # noqa: E402
    match_contact,
    match_org,
    match_placement,
)
from bookkit.repo import contacts as contacts_repo  # noqa: E402
from bookkit.repo import orgs as orgs_repo  # noqa: E402
from bookkit.repo import placements as placements_repo  # noqa: E402


def test_match_org_exact_fuzzy_conflict_none(conn) -> None:
    a = orgs_repo.create(conn, kind="client", name="Atomic Industries, Inc.")
    orgs_repo.create(conn, kind="client", name="Atomic Industrial Holdings")
    orgs_repo.create(conn, kind="market", name="Atomic Industries, Inc.")  # ignored: market
    exact, candidates = match_org(conn, "Atomic Industries, Inc.")
    assert exact == a.id and candidates == []
    none_id, candidates = match_org(conn, "Atomic Ind")  # too close to two clients
    assert none_id is None
    no_id, no_candidates = match_org(conn, "Zephyr Logistics")
    assert no_id is None and no_candidates == []


def test_match_contact_by_email(conn) -> None:
    org = orgs_repo.create(conn, kind="client", name="Atomic")
    rosa = contacts_repo.create(
        conn, org.id, first_name="Rosa", last_name="Silva", email="rosa@atomic.example.com"
    )
    assert match_contact(conn, org.id, "ROSA@atomic.example.com") == rosa.id
    assert match_contact(conn, org.id, "nobody@atomic.example.com") is None


# --- book mapper ------------------------------------------------------------------

from bookkit.imports.mappers.book import stage_book  # noqa: E402
from bookkit.imports.readers import RawTable  # noqa: E402
from bookkit.imports.tablemap import Mapping  # noqa: E402

_BOOK_HEADERS = [
    "account", "status", "contact_name", "contact_email",
    "program", "inception", "expiry", "premium", "commission",
]


def _book_table(rows: list[dict[str, object]]) -> tuple[RawTable, Mapping]:
    table = RawTable("book.csv", "ab" * 32, _BOOK_HEADERS, rows)
    return table, map_headers(_BOOK_HEADERS, BOOK_FIELDS)


def test_stage_book_stages_account_contact_placement(conn) -> None:
    table, mapping = _book_table([
        {
            "account": "Atomic Industries", "status": "active",
            "contact_name": "Rosa Silva", "contact_email": "Rosa@Atomic.example.com",
            "program": "Property", "inception": "10/1/2026", "expiry": "10/1/2027",
            "premium": "250,000", "commission": "15%",
        },
        {"account": "Atomic Industries", "contact_name": "Ken Ito",
         "contact_email": "ken@atomic.example.com"},
        {"account": "Borealis Foods", "status": "wibble"},
        {"account": "", "program": "Casualty"},
    ])
    staged = stage_book(conn, table, mapping)
    kinds = [(r.kind, r.action) for r in staged.records]
    assert kinds.count(("account", "create")) == 2
    assert kinds.count(("contact", "create")) == 2
    assert kinds.count(("placement", "create")) == 1

    placement = next(r for r in staged.records if r.kind == "placement")
    assert placement.fields["period_from"] == "2026-10-01"
    assert placement.fields["total_premium"] == 25_000_000  # cents
    assert placement.fields["commission_bps"] == 1500
    assert placement.fields["org_key"] == "Atomic Industries"

    contact = next(r for r in staged.records if "Rosa" in r.key)
    assert contact.fields["email"] == "Rosa@atomic.example.com"  # local part keeps case
    assert contact.fields["first_name"] == "Rosa"

    bad_status = next(r for r in staged.records if r.key == "Borealis Foods")
    assert any(i.field == "status" for i in bad_status.issues)
    assert any(i.field == "account" for r in staged.records for i in r.issues)
    assert not staged.ok


def test_stage_book_matches_existing_as_update(conn) -> None:
    org = orgs_repo.create(conn, kind="client", name="Atomic Industries")
    contacts_repo.create(
        conn, org.id, first_name="Rosa", last_name="Silva",
        email="rosa@atomic.example.com",
    )
    placements_repo.create(conn, org.id, "Property", "2026-10-01", "2027-10-01")
    table, mapping = _book_table([
        {
            "account": "Atomic Industries", "contact_name": "Rosa Silva",
            "contact_email": "rosa@atomic.example.com", "program": "Property",
            "inception": "2026-10-01", "expiry": "2027-10-01",
        },
    ])
    staged = stage_book(conn, table, mapping)
    assert staged.ok
    assert all(r.action == "update" and r.target_id for r in staged.records)


def test_stage_book_expiry_before_inception_is_error(conn) -> None:
    table, mapping = _book_table([
        {"account": "Atomic", "program": "Property",
         "inception": "2027-10-01", "expiry": "2026-10-01"},
    ])
    staged = stage_book(conn, table, mapping)
    assert any(i.field == "expiry" for r in staged.records for i in r.issues)
    assert not staged.ok


def test_match_placement_is_period_aware(conn) -> None:
    org = orgs_repo.create(conn, kind="client", name="Atomic")
    p26 = placements_repo.create(conn, org.id, "Property", "2026-10-01", "2027-10-01")
    p27 = placements_repo.create(conn, org.id, "Property", "2027-10-01", "2028-10-01")
    placements_repo.create(conn, org.id, "Casualty", "2026-10-01", "2027-10-01")
    assert match_placement(conn, org.id, "property", "2026-10-01", "2027-10-01") == p26.id
    assert match_placement(conn, org.id, "Property", "2027-12-01", "2028-12-01") == p27.id
    assert match_placement(conn, org.id, "Property", "2030-01-01", "2031-01-01") is None


# --- team colleague paste --------------------------------------------------------

from bookkit.imports.commit import commit_team_paste  # noqa: E402
from bookkit.imports.mappers.team_paste import stage_team_paste  # noqa: E402


def test_stage_team_paste_parses_and_matches_by_email(conn) -> None:
    from bookkit.repo import team as team_repo

    staged = stage_team_paste(conn, _SIGNATURE)
    [record] = staged.records
    assert record.fields["name"] == "Rosa Silva"
    assert record.fields["title"] == "Director of Risk Management"
    assert record.fields["email"] == "rosa.silva@atomic.example.com"
    assert staged.ok

    existing = team_repo.create_member(
        conn, "Rosa Silva", email="rosa.silva@atomic.example.com"
    )
    staged = stage_team_paste(conn, _SIGNATURE)
    assert staged.records[0].action == "update"
    assert staged.records[0].target_id == existing.id


def test_commit_team_paste_creates_member(db_path: Path) -> None:
    from bookkit.repo import team as team_repo

    connection = db.connect(db_path)
    try:
        staged = stage_team_paste(connection, _SIGNATURE)
        commit_team_paste(connection, staged, db_path)
        [member] = team_repo.list_members(connection)
        assert member.name == "Rosa Silva"
        assert member.phone == "(312) 555-0142"
    finally:
        connection.close()


def test_stage_team_paste_email_only_falls_back_with_warning(conn) -> None:
    staged = stage_team_paste(conn, "just-an-email@example.com")
    assert staged.ok  # email-local fallback, flagged for review
    assert staged.records[0].fields["name"] == "just-an-email"
    assert any(i.field == "contact_name" for i in staged.records[0].issues)


def test_stage_team_paste_nothing_usable_is_error(conn) -> None:
    staged = stage_team_paste(conn, "(312) 555-0142")  # phone but nobody
    assert not staged.ok


def test_verbose_report_shows_parsed_fields(conn) -> None:
    org = orgs_repo.create(conn, kind="client", name="Atomic Industries")
    staged = stage_contact_paste(conn, _SIGNATURE, org.id, org.name)
    report = staged.report(verbose=True)
    assert "rosa.silva@atomic.example.com" in report
    assert "(312) 555-0142" in report
    assert "Director of Risk Management" in report
