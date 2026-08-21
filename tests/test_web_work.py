"""The Work tab — open tasks, information requests, and (per request) their
items.

The assertion is deliberately NOT 'the field changed'. A plain outcome check
passes even when a route writes outside a batch (see test_web_writes.py's
docstring for the 33-call-site precedent this guards against). What is
asserted is that the batch exists, that it carries source='web', that it
carries real events, and that reverting it puts the record back.

seed.py seeds ~25 tasks at random but never seeds any RFI request — the
app_and_org fixture below makes both deterministic rather than relying on
chance or skipping (a skipped test protects nothing)."""

from __future__ import annotations

import re

from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from bookkit import db
from bookkit.web.app import create_app


@pytest.fixture
def app_and_org(snapshot_db: Path):
    app = create_app(snapshot_db)
    from bookkit.repo import orgs
    from bookkit.repo import rfi as rfi_repo
    from bookkit.repo import tasks as tasks_repo

    conn = app.state.conn
    org = orgs.list_orgs(conn, kind="client")[0]

    if not tasks_repo.open_tasks_for_client(conn, org.id):
        tasks_repo.create(conn, "Chase loss runs", org_id=org.id, due_on="2026-08-20")

    request = rfi_repo.create_request(conn, org.id, "Loss run refresh", "2026-08-10")
    rfi_repo.add_item(conn, request.id, "loss runs 2021-2025", category="Financials")

    with TestClient(app, base_url="http://127.0.0.1") as client:
        yield client, org, request


def _latest_batch(conn):
    from bookkit.repo import batches as batches_repo

    found = batches_repo.recent(conn, since="", limit=1)
    return found[0] if found else None


# --- the tab shell -----------------------------------------------------------


def test_work_tab_lists_tasks_and_requests(app_and_org):
    client, org, request = app_and_org
    response = client.get(f"/accounts/{org.ref}/work")
    assert response.status_code == 200
    assert request.title in response.text


def test_requests_say_who_was_asked_and_what_they_are_about(app_and_org):
    """A request with no asker and no scope is an ask you cannot chase."""
    client, org, request = app_and_org
    conn = client.app.state.conn
    from bookkit.services import rfi as rfi_svc

    response = client.get(f"/accounts/{org.ref}/work")
    assert rfi_svc.scope_label(conn, request) in response.text
    assert rfi_svc.asker_name(conn, request) in response.text


# --- tasks: inline cell editing -----------------------------------------------


def _category_cell(html: str) -> str:
    """The category cell's own element, start tag to its close — the badge
    must live INSIDE it (a <td> nested in a <td> silently drifts every later
    column out from under its header; see macros/cell.html)."""
    start = html.index('data-field="category"')
    start = html.rindex("<td", 0, start)
    return html[start:html.index("</td>", start)]


def test_internal_task_row_says_not_exported(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    tasks_repo.create(conn, "our own file note", org_id=org.id, category="Internal")
    tasks_repo.create(conn, "renew GL", org_id=org.id, category="Renewal")
    body = client.get(f"/accounts/{org.ref}/work").text
    assert body.count("tag-internal") == 1
    assert body.count("not exported") == 1
    cell = _category_cell(body[body.index("our own file note"):])
    assert "not exported" in cell
    assert "<td" not in cell[3:], "the badge opened a second <td>"


def test_saving_a_category_to_internal_returns_the_badge(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.create(conn, "our own file note", org_id=org.id, category="Renewal")
    response = client.post(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/category", data={"category": "Internal"}
    )
    assert response.status_code == 200
    assert "tag-internal" in response.text and "not exported" in response.text
    assert "<td" not in _category_cell(response.text)[3:]


def test_a_non_internal_category_carries_no_badge(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.create(conn, "audit support", org_id=org.id, category="Renewal")
    response = client.post(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/category",
        data={"category": "Internal Review"},
    )
    assert response.status_code == 200
    assert "not exported" not in response.text


def test_the_task_category_cell_editor_offers_the_vocabulary(app_and_org):
    """The inline cell is the PRIMARY edit path here — click the cell, no
    modal — so the completion that makes "Internal" discoverable has to be in
    it. The whole-record modal already had it; this is where the risk lives."""
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.create(conn, "audit support", org_id=org.id, category="Renewal")
    response = client.get(f"/accounts/{org.ref}/tasks/{task.id}/cell/category/edit")
    assert response.status_code == 200
    assert 'list="cl-category"' in response.text
    assert '<datalist id="cl-category">' in response.text
    # the one category nobody has typed yet, offered anyway (repo.vocab)
    assert '<option value="Internal">' in response.text
    assert '<option value="Renewal">' in response.text


def test_a_non_vocabulary_cell_editor_carries_no_datalist(app_and_org):
    """Only the category field has a vocabulary — a stray datalist on `title`
    would mean the enrichment is being applied by position, not by key."""
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]
    response = client.get(f"/accounts/{org.ref}/tasks/{task.id}/cell/title/edit")
    assert response.status_code == 200
    assert "datalist" not in response.text


def test_a_non_editable_key_is_404_before_the_vocabulary_query(app_and_org):
    """The editable-set guard still runs first on the editor route — the
    vocabulary lookup must not become a way to reach a field the cell
    contract does not expose."""
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]
    assert client.get(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/priority/edit"
    ).status_code == 404


def test_editing_a_task_cell_writes_one_web_batch(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import batches as batches_repo
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]
    response = client.post(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/title",
        data={"title": "Chase the loss runs — updated"},
    )
    assert response.status_code == 200
    assert tasks_repo.get(conn, task.id).title == "Chase the loss runs — updated"

    batch = _latest_batch(conn)
    assert batch is not None, "the edit wrote outside any batch — `R` cannot reach it"
    assert batch.source == "web"
    assert batches_repo.events_for(conn, batch.id), "the batch carries no events to revert"


def test_a_task_cell_edit_reverts(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo
    from bookkit.services import batches as batches_svc

    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]
    before = task.title

    client.post(f"/accounts/{org.ref}/tasks/{task.id}/cell/title", data={"title": "Renamed"})
    batch = _latest_batch(conn)
    assert batch is not None
    batches_svc.revert(conn, batch.ref, now=db.utc_now())
    assert tasks_repo.get(conn, task.id).title == before


def test_a_task_cell_blanking_the_required_title_is_refused_intact(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]
    before = task.title
    response = client.post(f"/accounts/{org.ref}/tasks/{task.id}/cell/title", data={"title": ""})
    assert response.status_code == 200
    assert "required" in response.text
    assert tasks_repo.get(conn, task.id).title == before


def test_a_task_cell_bare_number_due_date_is_refused_intact(app_and_org):
    """A bare number is not a date, on every surface."""
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]
    before = task.due_on
    response = client.post(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/due_on", data={"due_on": "5"}
    )
    assert response.status_code == 200
    assert "a bare number is ambiguous" in response.text
    assert tasks_repo.get(conn, task.id).due_on == before


def test_a_non_editable_task_key_is_404_not_a_write(app_and_org):
    """priority is on the whole-record form but not in TASK_FIELDS — which
    fields are inline-editable is not a per-surface choice."""
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]
    response = client.post(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/priority", data={"priority": "1"}
    )
    assert response.status_code == 404
    assert tasks_repo.get(conn, task.id).priority == task.priority


def test_adding_a_task_writes_one_web_batch(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    before = len(tasks_repo.open_tasks_for_client(conn, org.id))
    response = client.post(
        f"/accounts/{org.ref}/tasks/new",
        data={"title": "Prepare submission", "description": "", "category": "",
              "due_on": "2026-09-01", "priority": "2", "org_id": org.id, "detail": ""},
    )
    assert response.status_code == 200
    assert len(tasks_repo.open_tasks_for_client(conn, org.id)) == before + 1

    batch = _latest_batch(conn)
    assert batch is not None and batch.source == "web"


def test_a_task_with_a_bare_number_due_date_is_refused(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo

    before = len(tasks_repo.open_tasks_for_client(conn, org.id))
    response = client.post(
        f"/accounts/{org.ref}/tasks/new",
        data={"title": "Chase the loss runs", "description": "", "category": "",
              "due_on": "5", "priority": "2", "org_id": org.id, "detail": ""},
    )
    assert response.status_code == 200
    assert "a bare number is ambiguous" in response.text
    assert "Chase the loss runs" in response.text
    assert len(tasks_repo.open_tasks_for_client(conn, org.id)) == before


def test_completing_a_task_is_one_web_batch_and_reverts(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import tasks as tasks_repo
    from bookkit.services import batches as batches_svc

    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]
    response = client.post(f"/accounts/{org.ref}/tasks/{task.id}/done")
    assert response.status_code == 200
    assert tasks_repo.get(conn, task.id).status != task.status

    batch = _latest_batch(conn)
    assert batch is not None, "completion wrote outside any batch"
    assert batch.source == "web"

    batches_svc.revert(conn, batch.ref, now=db.utc_now())
    assert tasks_repo.get(conn, task.id).status == task.status


# --- requests: whole-form add and edit ---------------------------------------


def test_creating_a_request_writes_one_web_batch(app_and_org):
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import rfi as rfi_repo

    before = len(rfi_repo.requests_for_org(conn, org.id))
    response = client.post(
        f"/accounts/{org.ref}/requests/new",
        data={"title": "Renewal questionnaire", "requested_on": "2026-08-14",
              "due_on": "", "market_org_id": "", "placement_id": "",
              "project_id": "", "cancelled_at": "", "notes": ""},
    )
    assert response.status_code == 200
    assert len(rfi_repo.requests_for_org(conn, org.id)) == before + 1

    batch = _latest_batch(conn)
    assert batch is not None and batch.source == "web"


def test_a_request_with_a_bad_date_is_refused_intact(app_and_org):
    """A bare number is not a date, on every surface."""
    client, org, _ = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import rfi as rfi_repo

    before = len(rfi_repo.requests_for_org(conn, org.id))
    response = client.post(
        f"/accounts/{org.ref}/requests/new",
        data={"title": "Renewal questionnaire", "requested_on": "5",
              "due_on": "", "market_org_id": "", "placement_id": "",
              "project_id": "", "cancelled_at": "", "notes": "chase Friday"},
    )
    assert response.status_code == 200
    assert "a bare number is ambiguous" in response.text
    assert "Renewal questionnaire" in response.text
    assert "chase Friday" in response.text
    assert len(rfi_repo.requests_for_org(conn, org.id)) == before


def test_editing_a_request_writes_one_web_batch_and_reverts(app_and_org):
    client, org, request = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import rfi as rfi_repo
    from bookkit.services import batches as batches_svc

    before_title = request.title
    response = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/edit",
        data={"title": "Loss run refresh — round 2", "requested_on": request.requested_on,
              "due_on": request.due_on or "", "market_org_id": request.market_org_id or "",
              "placement_id": request.placement_id or "", "project_id": request.project_id or "",
              "cancelled_at": request.cancelled_at or "", "notes": request.notes or ""},
    )
    assert response.status_code == 200
    assert rfi_repo.get_request(conn, request.id).title == "Loss run refresh — round 2"

    batch = _latest_batch(conn)
    assert batch is not None, "the edit wrote outside any batch — `R` cannot reach it"
    assert batch.source == "web"

    batches_svc.revert(conn, batch.ref, now=db.utc_now())
    assert rfi_repo.get_request(conn, request.id).title == before_title


def test_a_refused_request_edit_keeps_every_value_and_writes_nothing(app_and_org):
    """The refusal contract on the form that shipped after Task 8. It routes
    through the shared `_save`, so this asserts the seam is actually taken —
    a green suite proves nothing broke, not that the new path is used.

    The batch count is not redundant with the row check: `_save` opens its
    batch BEFORE calling `write`, so a swallowed refusal strands an empty
    EventBatch in RECENT CHANGES with nothing for `Revert` to put back, even
    when the record itself is untouched."""
    client, org, request = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import batches as batches_repo
    from bookkit.repo import rfi as rfi_repo

    before_batches = len(batches_repo.recent(conn, since="", limit=50))

    response = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/edit",
        data={"title": "", "requested_on": request.requested_on,
              "due_on": "2026-08-31", "market_org_id": "", "placement_id": "",
              "project_id": "", "cancelled_at": "", "notes": "chase Friday"},
    )

    assert response.status_code == 200, "htmx drops 4xx — a refusal must be a 200"
    assert '<p class="form-error" role="alert">request is required</p>' in response.text
    # every other value survives the refusal
    assert "2026-08-31" in response.text
    assert "chase Friday" in response.text
    # and nothing was written — not the record, not a stranded batch
    assert rfi_repo.get_request(conn, request.id).title == request.title
    assert rfi_repo.get_request(conn, request.id).due_on == request.due_on
    assert len(batches_repo.recent(conn, since="", limit=50)) == before_batches


# --- items: detail page, inline editing, mark received -----------------------


def test_request_detail_lists_its_items(app_and_org):
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo

    items = rfi_repo.items_for_request(client.app.state.conn, request.id)
    response = client.get(f"/accounts/{org.ref}/requests/{request.id}")
    assert response.status_code == 200
    assert items[0].prompt in response.text


def test_editing_an_item_cell_writes_one_web_batch(app_and_org):
    client, org, request = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import batches as batches_repo
    from bookkit.repo import rfi as rfi_repo

    item = rfi_repo.items_for_request(conn, request.id)[0]
    response = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/cell/category",
        data={"category": "Loss history"},
    )
    assert response.status_code == 200
    assert rfi_repo.get_item(conn, item.id).category == "Loss history"

    batch = _latest_batch(conn)
    assert batch is not None, "the edit wrote outside any batch — `R` cannot reach it"
    assert batch.source == "web"
    assert batches_repo.events_for(conn, batch.id), "the batch carries no events to revert"


def test_a_non_editable_item_key_is_404_not_a_write(app_and_org):
    """status is deliberately not in RFI_ITEM_FIELDS — apply_rfi_item owns
    the status/received_on pair, and a bare cell write would let them
    disagree."""
    client, org, request = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import rfi as rfi_repo

    item = rfi_repo.items_for_request(conn, request.id)[0]
    before = item.status
    response = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/cell/status",
        data={"status": "received"},
    )
    assert response.status_code == 404
    assert rfi_repo.get_item(conn, item.id).status == before


def test_an_item_cell_bare_number_due_date_is_refused_intact(app_and_org):
    client, org, request = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import rfi as rfi_repo

    item = rfi_repo.items_for_request(conn, request.id)[0]
    before = item.due_on
    response = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/cell/due_on",
        data={"due_on": "5"},
    )
    assert response.status_code == 200
    assert "a bare number is ambiguous" in response.text
    assert rfi_repo.get_item(conn, item.id).due_on == before


def test_adding_an_item_writes_one_web_batch(app_and_org):
    client, org, request = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import rfi as rfi_repo

    before = len(rfi_repo.items_for_request(conn, request.id))
    response = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/new",
        data={"prompt": "SOV update", "kind": "document", "category": "Property",
              "due_on": "", "detail": "", "status": "outstanding",
              "received_on": "", "response": ""},
    )
    assert response.status_code == 200
    assert len(rfi_repo.items_for_request(conn, request.id)) == before + 1

    batch = _latest_batch(conn)
    assert batch is not None and batch.source == "web"


def test_marking_an_item_received_stamps_the_date_in_one_batch_and_reverts(app_and_org):
    """status OWNS received_on — the pair can never disagree. Both writes land
    in one batch, so a revert restores both."""
    client, org, request = app_and_org
    conn = client.app.state.conn
    from bookkit.repo import rfi as rfi_repo
    from bookkit.services import batches as batches_svc

    item = rfi_repo.items_for_request(conn, request.id)[0]
    assert item.status == "outstanding"

    response = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/received"
    )
    assert response.status_code == 200

    after = rfi_repo.get_item(conn, item.id)
    assert after.status == "received"
    assert after.received_on, "received without a date — the pair disagreed"

    batch = _latest_batch(conn)
    assert batch is not None and batch.source == "web"

    batches_svc.revert(conn, batch.ref, now=db.utc_now())
    restored = rfi_repo.get_item(conn, item.id)
    assert restored.status == "outstanding"
    assert not restored.received_on, "revert left a stale received date"


# --- the assignee cell ------------------------------------------------------


def test_the_work_tab_shows_and_saves_an_assignee(app_and_org) -> None:
    """The Work tab is where tasks are worked on the web, so the assignee has
    to be editable there — and it is the one cell whose save is not the
    generic single-column update every other cell uses (three columns hold
    the answer). A route that forgot that would 500 or, worse, write a name
    into a column that does not exist."""
    from bookkit.repo import assignees, contacts
    from bookkit.repo import tasks as tasks_repo

    client, org, _ = app_and_org
    conn = client.app.state.conn
    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]
    person = contacts.for_org(conn, org.id)[0]

    page = client.get(f"/accounts/{org.ref}/work")
    assert page.status_code == 200
    assert "<th>Assignee</th>" in page.text

    # the editor offers the account's own people, qualified
    editor = client.get(f"/accounts/{org.ref}/tasks/{task.id}/cell/assignee/edit")
    assert editor.status_code == 200
    label = f"{person.name} — {org.name}"
    assert label in editor.text, editor.text[:800]

    saved = client.post(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/assignee", data={"assignee": label}
    )
    assert saved.status_code == 200
    stored = tasks_repo.get(conn, task.id)
    assert stored.assignee_kind == "contact"
    assert stored.assignee_id == person.id
    assert stored.assignee_name is None
    # and the cell that came back shows them
    assert person.name in saved.text

    # one writer action, one undo unit — and all three columns inside it
    batch = _latest_batch(conn)
    assert batch is not None and batch.source == "web"
    fields = {
        row["field"]
        for row in conn.execute(
            "SELECT field FROM event_log WHERE batch_id = ?", (batch.id,)
        )
    }
    assert fields == {"assignee_kind", "assignee_id"}, fields
    assert assignees.label_of(conn, stored) == label


def test_an_unresolvable_assignee_is_kept_as_typed(app_and_org) -> None:
    """Freeform is the requirement, not the fallback: the AE has to be able
    to name somebody the book has never heard of. It is stored where nothing
    can read it as an identity, so the client's Owner column cannot move."""
    from bookkit.repo import tasks as tasks_repo

    client, org, _ = app_and_org
    conn = client.app.state.conn
    task = tasks_repo.open_tasks_for_client(conn, org.id)[0]

    client.post(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/assignee",
        data={"assignee": "Marisa at Lockton"},
    )
    stored = tasks_repo.get(conn, task.id)
    assert stored.assignee_kind is None and stored.assignee_id is None
    assert stored.assignee_name == "Marisa at Lockton"


# --- a posted select must name one of the form's own options ------------------


def test_an_item_status_outside_the_vocabulary_is_refused(app_and_org):
    """The reviewer's second reproduction. `status="NOT_A_STATUS"` was
    storable, and an item that is neither outstanding nor received reads as
    closed: the request drops off every attention queue, silently, because
    nothing downstream expects a value outside the vocabulary."""
    from bookkit.repo import rfi as rfi_repo

    client, org, request = app_and_org
    conn = client.app.state.conn
    before = len(rfi_repo.items_for_request(conn, request.id))

    response = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/new",
        data={"prompt": "loss runs 2026", "kind": "question",
              "status": "NOT_A_STATUS"},
    )

    assert response.status_code == 200
    assert "form-error" in response.text
    # commit-in-place: what was typed comes back, and nothing was written
    assert "loss runs 2026" in response.text
    assert len(rfi_repo.items_for_request(conn, request.id)) == before


def test_an_item_status_inside_the_vocabulary_still_saves(app_and_org):
    from bookkit.models import RFI_ITEM_STATUSES
    from bookkit.repo import rfi as rfi_repo

    client, org, request = app_and_org
    conn = client.app.state.conn
    assert "waived" in RFI_ITEM_STATUSES

    response = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/new",
        data={"prompt": "certificate of insurance", "kind": "question",
              "status": "waived"},
    )

    assert response.status_code == 200
    items = rfi_repo.items_for_request(conn, request.id)
    saved = next(i for i in items if i.prompt == "certificate of insurance")
    assert saved.status == "waived"


def test_a_task_priority_outside_the_vocabulary_is_refused(app_and_org):
    """Not only the RFI form: the guard is in forms.spec.parse_value, so every
    select on every form on both surfaces inherits it."""
    from bookkit.repo import tasks as tasks_repo

    client, org, _request = app_and_org
    conn = client.app.state.conn
    before = len(tasks_repo.open_tasks_for_client(conn, org.id))

    response = client.post(
        f"/accounts/{org.ref}/tasks/new",
        data={"title": "escalate", "priority": "99"},
    )

    assert response.status_code == 200
    assert "form-error" in response.text
    assert len(tasks_repo.open_tasks_for_client(conn, org.id)) == before


# --- the column class survives an edit ---------------------------------------
#
# A cell is built in three places per table: the panel's first render, the
# display route htmx swaps back after a save, and the editor. They each used to
# carry their own literal class map, so a class present on the first render
# could vanish the moment the cell was edited — and the column would change
# shape mid-session with nothing to explain it. These tests compare what the
# panel rendered against what the routes return, rather than asserting a
# hard-coded class list in three places of their own.


def _cell_classes(html: str, field: str) -> set[str]:
    """The classes on the cell for `field`, read out of rendered HTML."""
    import re

    match = re.search(rf'<td class="([^"]*)"[^>]*data-field="{field}"', html)
    assert match, f"no cell rendered for {field} in:\n{html[:2000]}"
    return set(match.group(1).split())


@pytest.mark.parametrize("field", ["prompt", "response", "due_on", "category"])
def test_an_item_cell_comes_back_from_a_save_dressed_as_it_started(app_and_org, field):
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo

    item = rfi_repo.items_for_request(client.app.state.conn, request.id)[0]
    panel = client.get(f"/accounts/{org.ref}/requests/{request.id}").text
    swapped = client.get(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/cell/{field}"
    ).text

    assert _cell_classes(swapped, field) == _cell_classes(panel, field)


@pytest.mark.parametrize("field", ["title", "description", "due_on", "assignee"])
def test_a_task_cell_comes_back_from_a_save_dressed_as_it_started(app_and_org, field):
    client, org, _request = app_and_org
    from bookkit.repo import tasks as tasks_repo

    task = tasks_repo.open_tasks_for_client(client.app.state.conn, org.id)[0]
    panel = client.get(f"/accounts/{org.ref}/work").text
    swapped = client.get(f"/accounts/{org.ref}/tasks/{task.id}/cell/{field}").text

    assert _cell_classes(swapped, field) == _cell_classes(panel, field)


def test_the_question_and_the_answer_are_marked_as_prose(app_and_org):
    """Not decoration: `prose` is what lets those two columns wrap. Without it
    the table sizes itself to the longest unwrapped sentence and the columns to
    its right — Status, Response, and the Mark received button — render past
    the panel's edge, live and invisible."""
    client, org, request = app_and_org

    panel = client.get(f"/accounts/{org.ref}/requests/{request.id}").text

    assert "prose" in _cell_classes(panel, "prompt")
    assert "prose" in _cell_classes(panel, "response")


def test_the_tables_that_carry_prose_declare_that_they_fit(app_and_org):
    """`rows-fit` is the other half of the same fix — prose can only wrap if
    the table is not sizing itself to max-content."""
    client, org, request = app_and_org

    for url in (f"/accounts/{org.ref}/work", f"/accounts/{org.ref}/requests/{request.id}"):
        assert 'class="rows rows-fit"' in client.get(url).text, url


# --- the roomy door onto the answer -------------------------------------------


def _first_item(client, request):
    from bookkit.repo import rfi as rfi_repo

    return rfi_repo.items_for_request(client.app.state.conn, request.id)[0]


def test_the_answer_form_opens_with_what_is_already_there(app_and_org):
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo

    item = _first_item(client, request)
    conn = client.app.state.conn
    with db.transaction(conn):
        rfi_repo.update_item(conn, item.id, response="partial: 2023 onward sent")

    form = client.get(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/answer"
    )

    assert form.status_code == 200
    assert "partial: 2023 onward sent" in form.text
    assert "<textarea" in form.text, "a long answer needs a box, not a one-line input"


def test_answering_writes_the_response_and_leaves_the_status_alone(app_and_org):
    """Grant, 2026-08-19: notes may just be notes. An answer that quietly
    marked the item received would empty a chase queue nobody had cleared."""
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo

    item = _first_item(client, request)
    assert item.status == "outstanding"

    saved = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/answer",
        data={"response": "Controller is pulling the class-code split; expect Friday."},
    )

    assert saved.status_code == 200
    after = rfi_repo.get_item(client.app.state.conn, item.id)
    assert after.response == "Controller is pulling the class-code split; expect Friday."
    assert after.status == "outstanding"
    assert after.received_on is None


def test_answering_is_one_revertible_web_batch(app_and_org):
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo
    from bookkit.services import batches as batches_svc

    item = _first_item(client, request)
    conn = client.app.state.conn

    client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/answer",
        data={"response": "sent 12 Aug"},
    )

    batch = _latest_batch(conn)
    assert batch is not None and batch.source == "web"
    assert item.prompt[:20] in batch.summary

    batches_svc.revert(conn, batch.ref, now=db.utc_now())
    assert rfi_repo.get_item(conn, item.id).response != "sent 12 Aug"


def test_every_item_offers_the_bigger_box(app_and_org):
    """Including a received one — an answer is edited after the fact more
    often than it is right first time."""
    client, org, request = app_and_org

    panel = client.get(f"/accounts/{org.ref}/requests/{request.id}").text

    for item in _items(client, request):
        assert f"/items/{item.id}/answer" in panel, item.prompt


def _items(client, request):
    from bookkit.repo import rfi as rfi_repo

    return rfi_repo.items_for_request(client.app.state.conn, request.id)


# --- taking back an ask filed in error (2026-08-19) ---------------------------
#
# The web half of what MCP got the same day. rfi_repo.delete_request and
# delete_item had no caller anywhere, which is why an RFI the MCP filed by
# mistake could not be taken off Grant's book from any surface at all.


def test_the_confirm_step_writes_nothing(app_and_org):
    """A step, not a browser confirm(): removing a request takes its items with
    it, and the plan for that has to be visible before it happens."""
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo

    confirm = client.get(f"/accounts/{org.ref}/requests/{request.id}/remove")

    assert confirm.status_code == 200
    assert request.title in confirm.text
    assert rfi_repo.find_request(client.app.state.conn, request.id) is not None


def test_the_confirm_step_says_how_many_items_go_with_it(app_and_org):
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo

    count = len(rfi_repo.items_for_request(client.app.state.conn, request.id))
    assert count, "the fixture request has no items — the plan would say nothing"

    confirm = client.get(f"/accounts/{org.ref}/requests/{request.id}/remove").text

    assert str(count) in confirm


def test_removing_a_request_takes_its_items_and_is_revertible(app_and_org):
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo
    from bookkit.services import batches as batches_svc

    conn = client.app.state.conn
    removed = client.post(f"/accounts/{org.ref}/requests/{request.id}/remove")

    assert removed.status_code == 200
    assert rfi_repo.find_request(conn, request.id) is None
    batch = _latest_batch(conn)
    assert batch.source == "web" and batch.tool == "request_remove"

    batches_svc.revert(conn, batch.ref, now=db.utc_now())
    assert rfi_repo.find_request(conn, request.id) is not None
    assert rfi_repo.items_for_request(conn, request.id)


def test_removing_an_answered_request_is_refused_in_the_page(app_and_org):
    """A refusal SAYS SOMETHING. htmx drops 4xx, so a destructive control that
    refuses with a status code produces no swap, no message and no change."""
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo

    conn = client.app.state.conn
    item = rfi_repo.items_for_request(conn, request.id)[0]
    with db.transaction(conn):
        rfi_repo.update_item(conn, item.id, response="sent 12 Aug")

    refused = client.post(f"/accounts/{org.ref}/requests/{request.id}/remove")

    assert refused.status_code == 200
    assert "answered" in refused.text
    assert rfi_repo.find_request(conn, request.id) is not None


def test_one_item_can_be_taken_off_a_request(app_and_org):
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo

    conn = client.app.state.conn
    item = rfi_repo.items_for_request(conn, request.id)[0]

    removed = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/remove"
    )

    assert removed.status_code == 200
    left = [i.id for i in rfi_repo.items_for_request(conn, request.id)]
    assert item.id not in left
    assert rfi_repo.find_request(conn, request.id) is not None, "the request went too"
    assert _latest_batch(conn).tool == "request_item_remove"


def test_removing_an_answered_item_is_refused_in_the_page(app_and_org):
    client, org, request = app_and_org
    from bookkit.repo import rfi as rfi_repo

    conn = client.app.state.conn
    item = rfi_repo.items_for_request(conn, request.id)[0]
    with db.transaction(conn):
        rfi_repo.update_item(conn, item.id, response="sent 12 Aug")

    refused = client.post(
        f"/accounts/{org.ref}/requests/{request.id}/items/{item.id}/remove"
    )

    assert refused.status_code == 200
    assert "answered" in refused.text
    assert item.id in [i.id for i in rfi_repo.items_for_request(conn, request.id)]


def test_the_page_offers_both_removals(app_and_org):
    """A route nothing renders is a route nobody can use — the lesson from the
    market cell editor earlier the same day."""
    client, org, request = app_and_org

    work = client.get(f"/accounts/{org.ref}/work").text
    items = client.get(f"/accounts/{org.ref}/requests/{request.id}").text

    assert f"/requests/{request.id}/remove" in work
    assert "/remove" in items and "items/" in items


# --- the Work tab's own export ----------------------------------------------


def test_the_work_tab_offers_an_export(app_and_org):
    """Grant, 2026-08-21: "an export to .xlsx on the Work page for open tasks
    and information requests"."""
    client, org, _ = app_and_org
    page = client.get(f"/accounts/{org.ref}/work").text
    assert f"/accounts/{org.ref}/export/work.xlsx" in page


def test_the_work_export_carries_the_two_tables_and_nothing_else(app_and_org):
    """"just these tables" — no Schedule of Insurance, no Projects. Those are
    the client deliverable's business and this file is the working one."""
    from io import BytesIO

    from openpyxl import load_workbook

    client, org, request = app_and_org
    body = client.get(f"/accounts/{org.ref}/export/work.xlsx")

    assert body.status_code == 200
    assert "attachment" in body.headers["content-disposition"]
    wb = load_workbook(BytesIO(body.content))
    assert any(name.startswith("Open Items") for name in wb.sheetnames), wb.sheetnames
    assert "Information Requests" in wb.sheetnames
    assert "Schedule of Insurance" not in wb.sheetnames
    assert "Projects" not in wb.sheetnames


def test_the_work_export_and_the_deliverable_cannot_disagree(app_and_org, tmp_path):
    """THE REASON THE SHEET BUILDERS WERE EXTRACTED. Two writers each composing
    "open items" would each look right read on its own and drift apart on the
    day one of them learned a rule the other did not.

    Compared cell by cell rather than by sheet name: a shared name proves the
    tabs match, not the rows.
    """
    from datetime import date as _date
    from io import BytesIO

    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    client, org, _ = app_and_org
    conn = client.app.state.conn
    today = _date.today()

    work = load_workbook(
        BytesIO(client.get(f"/accounts/{org.ref}/export/work.xlsx").content)
    )
    full = load_workbook(write(conn, org.id, tmp_path / "full.xlsx", today))

    for name in ("Information Requests",):
        assert name in work.sheetnames and name in full.sheetnames
        assert [[c.value for c in row] for row in work[name].iter_rows()] == [
            [c.value for c in row] for row in full[name].iter_rows()
        ], f"{name} differs between the two workbooks"

    open_items_work = next(n for n in work.sheetnames if n.startswith("Open Items"))
    open_items_full = next(n for n in full.sheetnames if n.startswith("Open Items"))
    assert [[c.value for c in row] for row in work[open_items_work].iter_rows()] == [
        [c.value for c in row] for row in full[open_items_full].iter_rows()
    ], "the Open Items sheets differ between the two workbooks"


def test_the_work_export_withholds_internal_items(app_and_org):
    """It is narrower than the client deliverable, not looser: it inherits
    `compose()`'s Internal rule, so our own file notes stay ours."""
    from io import BytesIO

    from openpyxl import load_workbook

    from bookkit.repo import tasks as tasks_repo

    client, org, _ = app_and_org
    conn = client.app.state.conn
    tasks_repo.create(
        conn, "our own file note", org_id=org.id, category="Internal"
    )

    wb = load_workbook(
        BytesIO(client.get(f"/accounts/{org.ref}/export/work.xlsx").content)
    )
    values = [
        str(c.value)
        for name in wb.sheetnames
        for row in wb[name].iter_rows()
        for c in row
        if c.value is not None
    ]
    assert not any("our own file note" in v for v in values)


# --- the assignee cell reads as a person, and edits as a resolvable one -------


def _input_value(html: str, name: str) -> str:
    """The `value` of the `<input name="...">` in a rendered editor cell.

    Deliberately NOT a substring search over the whole fragment: the editor
    also renders a datalist whose options carry the same strings, so a naive
    `in` check cannot tell the pre-filled value from a suggestion."""
    match = re.search(
        rf'<input[^>]*\bname="{re.escape(name)}"[^>]*>', html
    )
    assert match, f"no <input name={name!r}> in the editor fragment"
    value = re.search(r'\bvalue="([^"]*)"', match.group(0))
    return value.group(1) if value else ""


def _a_team_member(conn):
    """Somebody on OUR team, so the qualified label carries "— our team" —
    the exact suffix Grant asked to stop seeing on the row."""
    from bookkit.repo import team

    members = team.list_members(conn)
    assert members, "fixture drifted — no team members to assign"
    return members[0]


def test_the_assignee_cell_shows_the_person_without_the_qualifier(app_and_org):
    """Grant, 2026-08-21: "i do not like the ' — our team' suffix that is added.
    It should just be the person"."""
    from bookkit.repo import assignees as assignees_repo
    from bookkit.repo import tasks as tasks_repo

    client, org, _ = app_and_org
    conn = client.app.state.conn
    member = _a_team_member(conn)
    task = tasks_repo.create(conn, "Bind the layer", org_id=org.id)
    assignees_repo.set_on_task(
        conn, task.id, f"{member.name} — our team", org_id=org.id
    )

    cell = client.get(f"/accounts/{org.ref}/tasks/{task.id}/cell/assignee").text

    assert member.name in cell
    assert "our team" not in cell, "the qualifier is back on the row"


def test_the_assignee_EDITOR_still_prefills_the_qualified_label(app_and_org):
    """THE HALF THAT MUST NOT MOVE, and the reason the suffix could not simply
    be deleted.

    What a form pre-fills has to be a value its own resolver accepts back
    unchanged. Pre-fill the plain name and opening a task and pressing save
    silently downgrades a resolved assignee to freeform — the same failure
    CLAUDE.md's ENTRY ACCEPTS CENTS rule describes on a different field.

    The two forms can differ at all only because the editor is fetched from the
    server in its own request, so what the row displays never reaches the
    input.
    """
    from bookkit.repo import assignees as assignees_repo
    from bookkit.repo import tasks as tasks_repo

    client, org, _ = app_and_org
    conn = client.app.state.conn
    member = _a_team_member(conn)
    task = tasks_repo.create(conn, "Draft the strategy", org_id=org.id)
    assignees_repo.set_on_task(
        conn, task.id, f"{member.name} — our team", org_id=org.id
    )

    editor = client.get(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/assignee/edit"
    ).text

    # THE INPUT's value, not merely the string somewhere on the fragment. The
    # first version of this test asserted `value="… — our team"` appeared in
    # the response and passed with the editor pre-filling the PLAIN name —
    # because the qualified label is also one of the datalist's <option
    # value="…"> suggestions. A mutation caught it (2026-08-21).
    assert _input_value(editor, "assignee") == f"{member.name} — our team", editor[:500]


def test_saving_an_untouched_assignee_keeps_it_resolved(app_and_org):
    """The round trip the two rules above exist to protect, end to end: open
    the cell, save what it pre-filled, and the assignee is still the resolved
    team member.

    THE COLLISION IS THE POINT. An unambiguous name resolves perfectly well on
    its own, so this test proved nothing until a contact at the client was
    given the SAME name as our colleague — a mutation pre-filling the plain
    name passed without it (2026-08-21). With two candidates, only the
    qualified label says which person, and a plain name is a refusal to guess.
    """
    from bookkit.repo import assignees as assignees_repo
    from bookkit.repo import contacts as contacts_repo
    from bookkit.repo import tasks as tasks_repo

    client, org, _ = app_and_org
    conn = client.app.state.conn
    member = _a_team_member(conn)
    first, _, last = member.name.partition(" ")
    contacts_repo.create(conn, org.id, first_name=first, last_name=last)
    task = tasks_repo.create(conn, "Confirm the SIR", org_id=org.id)
    assignees_repo.set_on_task(
        conn, task.id, f"{member.name} — our team", org_id=org.id
    )

    editor = client.get(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/assignee/edit"
    ).text
    prefilled = _input_value(editor, "assignee")
    client.post(
        f"/accounts/{org.ref}/tasks/{task.id}/cell/assignee",
        data={"assignee": prefilled},
    )

    saved = tasks_repo.get(conn, task.id)
    assert saved.assignee_id == member.id, "the assignee stopped resolving"
