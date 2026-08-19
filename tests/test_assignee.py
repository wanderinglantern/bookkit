"""Who is chasing a task — and which side of the client's table it lands on.

The load-bearing test in here is the last group: a workbook is GENERATED and
read back, and an assignee who is a contact on the client's own account
renders `You` while a team member and a market contact render `Us`. Asserting
on the row objects would prove the composition and not the deliverable, and
the deliverable is the thing a client reads."""

from __future__ import annotations

import sqlite3
from datetime import date
from pathlib import Path

import pytest

from bookkit import db
from bookkit.models import AssigneeKind
from bookkit.repo import assignees, contacts, orgs, tasks, team

TODAY = date(2026, 8, 18)


@pytest.fixture
def book(conn: sqlite3.Connection) -> dict[str, object]:
    """One client with a contact, one market with an underwriter, one
    colleague — the three suggestion sources, one row each."""
    client = orgs.create(conn, name="Atomic Industries", kind="client")
    market = orgs.create(conn, name="Zurich", kind="market")
    theirs = contacts.create(conn, client.id, first_name="Rae", last_name="Okafor")
    underwriter = contacts.create(conn, market.id, first_name="Jo", last_name="Chen")
    ours = team.create_member(conn, "Dana Reyes")
    return {
        "client": client, "market": market, "theirs": theirs,
        "underwriter": underwriter, "ours": ours,
    }


# --- the candidate list -----------------------------------------------------


def test_the_picker_offers_team_the_account_and_the_markets(conn, book) -> None:
    """THE correction that reshaped this design. An earlier version of the
    spec said freeform mattered because "underwriters and wholesalers are not
    records in the book" — they are: contacts.for_org never filtered by org
    kind and the markets screen has bound `w` → add_underwriter since it was
    written. A picker without them is missing exactly the people the AE
    spends the day chasing."""
    labels = [c.label for c in assignees.candidates(conn, book["client"].id)]
    assert labels == [
        "Dana Reyes — our team",
        "Jo Chen — Zurich",
        "Rae Okafor — Atomic Industries",
    ], labels


def test_a_market_contact_is_offered_even_with_no_account_in_hand(conn, book) -> None:
    """ctrl+t opens the task form before an account has been chosen. The two
    unscoped sources are still offered; the account's own people cannot be,
    because there is no account yet."""
    labels = [c.label for c in assignees.candidates(conn, None)]
    assert labels == ["Dana Reyes — our team", "Jo Chen — Zurich"]


def test_another_clients_contacts_are_not_offered(conn, book) -> None:
    """Scoped to THIS account. Every client contact in the book would be a
    list nobody can read and a picker that can name the wrong company's
    risk manager."""
    other = orgs.create(conn, name="Borealis Foods", kind="client")
    contacts.create(conn, other.id, first_name="Sam", last_name="Ruiz")
    labels = [c.label for c in assignees.candidates(conn, book["client"].id)]
    assert "Sam Ruiz — Borealis Foods" not in labels


# --- resolution, and what a collision does ----------------------------------


def test_the_qualified_label_resolves_to_that_person(conn, book) -> None:
    picked = assignees.resolve(
        assignees.candidates(conn, book["client"].id),
        "Rae Okafor — Atomic Industries",
    )
    assert picked is not None
    assert (picked.kind, picked.id) == (AssigneeKind.CONTACT, book["theirs"].id)


def test_a_bare_name_resolves_when_it_names_exactly_one_person(conn, book) -> None:
    """Typing past the picker still works — the field is freeform by
    requirement and most names are not ambiguous."""
    picked = assignees.resolve(
        assignees.candidates(conn, book["client"].id), "dana reyes"
    )
    assert picked is not None
    assert (picked.kind, picked.id) == (AssigneeKind.TEAM, book["ours"].id)


def test_a_bare_name_two_people_share_resolves_to_neither(conn, book) -> None:
    """THE collision question, decided.

    A colleague and a client contact with the same name are the exact pair
    that would flip the client-facing Owner column, and taking the first match
    is the bug repo/team.py's uniqueness guard was written to stop. So the
    bare name refuses; the qualified labels still resolve, both of them."""
    client = book["client"]
    twin = contacts.create(conn, client.id, first_name="Dana", last_name="Reyes")
    pool = assignees.candidates(conn, client.id)

    assert assignees.resolve(pool, "Dana Reyes") is None

    theirs = assignees.resolve(pool, "Dana Reyes — Atomic Industries")
    ours = assignees.resolve(pool, "Dana Reyes — our team")
    assert theirs is not None and theirs.id == twin.id
    assert ours is not None and ours.id == book["ours"].id


def test_an_unresolved_name_is_stored_freeform_with_no_kind(conn, book) -> None:
    """A third party who is genuinely not in the book. The name is kept — the
    AE has to be able to write "Marisa at Lockton" — and it is kept somewhere
    that can never be read as an identity."""
    t = tasks.create(conn, "chase the binder", org_id=book["client"].id)
    assignees.set_on_task(conn, t.id, "Marisa at Lockton", org_id=book["client"].id)
    saved = tasks.get(conn, t.id)
    assert saved.assignee_kind is None
    assert saved.assignee_id is None
    assert saved.assignee_name == "Marisa at Lockton"


def test_clearing_an_assignee_nulls_the_id_the_last_one_left(conn, book) -> None:
    """The three columns only mean something together. A clear that wrote the
    name and left the id would leave a task pointing at somebody nobody can
    see, and the export would still read that stale kind."""
    t = tasks.create(conn, "chase the binder", org_id=book["client"].id)
    assignees.set_on_task(conn, t.id, "Dana Reyes", org_id=book["client"].id)
    assert tasks.get(conn, t.id).assignee_id == book["ours"].id

    assignees.set_on_task(conn, t.id, "", org_id=book["client"].id)
    saved = tasks.get(conn, t.id)
    assert (saved.assignee_kind, saved.assignee_id, saved.assignee_name) == (
        None, None, None
    )


def test_a_renamed_colleague_is_renamed_on_every_task_they_hold(conn, book) -> None:
    """The name is never denormalized onto the task. It is read live, so a
    rename in the team screen needs no sweep and leaves nothing anywhere
    still saying the old name."""
    t = tasks.create(conn, "chase the binder", org_id=book["client"].id)
    assignees.set_on_task(conn, t.id, "Dana Reyes", org_id=book["client"].id)
    team.update_member(conn, book["ours"].id, name="Dana Okonkwo")
    assert assignees.name_of(conn, tasks.get(conn, t.id)) == "Dana Okonkwo"


def test_a_removed_contact_reads_as_unassigned_not_as_a_crash(conn, book) -> None:
    """Both target tables soft-delete, so an id can outlive the row it names.
    A list of somebody's open work must not blow up on it."""
    t = tasks.create(conn, "chase the binder", org_id=book["client"].id)
    assignees.set_on_task(
        conn, t.id, "Rae Okafor — Atomic Industries", org_id=book["client"].id
    )
    contacts.delete(conn, book["theirs"].id)
    assert assignees.name_of(conn, tasks.get(conn, t.id)) == ""


def test_the_editor_prefills_a_value_its_own_resolver_accepts_back(conn, book) -> None:
    """Opening a task and pressing save must not quietly downgrade a resolved
    assignee to freeform. Same rule as ENTRY ACCEPTS CENTS: a form that
    pre-fills a value its own parser refuses corrupts the record on save.

    Driven through a COLLIDING name on purpose: a unique one round-trips
    whether the prefill is qualified or not, so it would pass over a prefill
    that is only accidentally acceptable."""
    contacts.create(conn, book["market"].id, first_name="Rae", last_name="Okafor")
    t = tasks.create(conn, "chase the binder", org_id=book["client"].id)
    assignees.set_on_task(
        conn, t.id, "Rae Okafor — Atomic Industries", org_id=book["client"].id
    )
    prefilled = assignees.label_of(conn, tasks.get(conn, t.id))
    assignees.set_on_task(conn, t.id, prefilled, org_id=book["client"].id)
    saved = tasks.get(conn, t.id)
    assert saved.assignee_kind is AssigneeKind.CONTACT
    assert saved.assignee_id == book["theirs"].id
    assert saved.assignee_name is None


# --- one writer action, one undo unit ---------------------------------------


def test_assigning_is_one_batch_and_reverts_all_of_it(conn, book) -> None:
    """Three columns change; `R` puts all three back or none of them. Three
    separate updates would be three events `u` returns one at a time, and the
    halfway state (a kind with no id) means nothing."""
    from bookkit.services import batches as batches_svc

    t = tasks.create(conn, "chase the binder", org_id=book["client"].id)
    assignees.set_on_task(conn, t.id, "Marisa at Lockton", org_id=book["client"].id)

    with batches_svc.open_batch(
        conn, source="tui", tool="edit_task", summary="set assignee",
        org_id=book["client"].id,
    ) as batch:
        assignees.set_on_task(
            conn, t.id, "Rae Okafor — Atomic Industries", org_id=book["client"].id
        )
    moved = tasks.get(conn, t.id)
    assert moved.assignee_id == book["theirs"].id and moved.assignee_name is None

    result = batches_svc.revert(conn, batch.ref, db.utc_now())
    assert result.applied and not result.refused, result.refused
    # all three columns, in ONE revert — a batch that only put back the
    # column that happened to be logged last is the halfway state
    assert {c.field for c in result.reverted} == {
        "assignee_kind", "assignee_id", "assignee_name"
    }
    back = tasks.get(conn, t.id)
    assert (back.assignee_kind, back.assignee_id, back.assignee_name) == (
        None, None, "Marisa at Lockton"
    )


# --- the migration ----------------------------------------------------------


def test_migration_013_is_additive_only() -> None:
    """CLAUDE.md's rule and Grant's standing ruling: additive migrations only,
    nothing rewrites existing rows. A DROP, an UPDATE or a table rebuild in
    this file would take data with it and there is no undo. Same assertion
    012 carries, on the same terms."""
    migrations = Path(__file__).resolve().parent.parent / "migrations"
    sql = (migrations / "013_task_assignee.sql").read_text()
    body = " ".join(
        line.strip().upper()
        for line in sql.splitlines()
        if line.strip() and not line.strip().startswith("--")
    )
    for forbidden in ("DROP ", "DELETE ", "UPDATE ", "RENAME ", "INSERT "):
        assert forbidden not in body, f"013 is not additive: it contains {forbidden!r}"
    for column in ("ASSIGNEE_KIND", "ASSIGNEE_ID", "ASSIGNEE_NAME"):
        assert f"ALTER TABLE TASK ADD COLUMN {column}" in body


def test_the_migration_is_additive_and_snapshots_the_book_first(
    tmp_path: Path,
) -> None:
    """CLAUDE.md's rule, on the one choke point every surface reaches
    migrations through. A book that already holds rows is copied into
    backups/ before 013 touches its shape, and every row it held survives
    with the three new columns reading NULL — which is "nobody has said whose
    this is", the state the export renders as Us."""
    path = tmp_path / "book.db"
    conn = db.connect(path, migrate=False)
    conn.execute(
        "CREATE TABLE IF NOT EXISTS schema_version"
        " (version INTEGER PRIMARY KEY, applied_at TEXT NOT NULL)"
    )
    pending = sorted(db.pending_migrations(conn))
    assert pending[-1][0] == 13, "013 is not the newest migration any more"
    for version, sql_path in pending[:-1]:
        conn.executescript(sql_path.read_text())
        conn.execute(
            "INSERT INTO schema_version (version, applied_at) VALUES (?, ?)",
            (version, db.utc_now()),
        )
    org = orgs.create(conn, name="Atomic Industries", kind="client")
    before = tasks.create(conn, "a task that predates the assignee", org_id=org.id)
    conn.close()

    assert not (tmp_path / "backups").exists()
    migrated = db.connect(path)
    try:
        assert db.schema_version(migrated) == 13
        backups = list((tmp_path / "backups").glob("*.bak"))
        assert len(backups) == 1, backups

        survived = tasks.get(migrated, before.id)
        assert survived.title == "a task that predates the assignee"
        assert (
            survived.assignee_kind,
            survived.assignee_id,
            survived.assignee_name,
        ) == (None, None, None)

        # and the snapshot is a real book, openable, holding the pre-013 rows
        copy = sqlite3.connect(backups[0])
        try:
            assert copy.execute(
                "SELECT title FROM task WHERE id = ?", (before.id,)
            ).fetchone()[0] == "a task that predates the assignee"
            columns = {r[1] for r in copy.execute("PRAGMA table_info(task)")}
            assert "assignee_kind" not in columns, "the snapshot is post-migration"
        finally:
            copy.close()
    finally:
        migrated.close()


# --- the client's Owner column, through a generated workbook ----------------


def _owner_column(path: Path) -> dict[str, str]:
    """{item: owner} read off sheet 1 of a REAL .xlsx.

    Through the file, not through the row objects: the composition can be
    right while the writer's column tuple puts the value under the wrong
    header, which is precisely how a bool nearly reached the Status column."""
    from openpyxl import load_workbook  # test-only import; src never imports it

    sheet = load_workbook(path).active
    header = [c.value for c in sheet[1]]
    assert header[-1] == "Owner", header
    return {
        str(row[0].value): str(row[-1].value)
        for row in sheet.iter_rows(min_row=2)
        if row[0].value and row[-1].value in ("You", "Us")
    }


def test_the_owner_column_is_derived_from_the_kind_not_from_a_name(
    conn, book, tmp_path
) -> None:
    """THE test this feature exists for.

    Four tasks, four assignees, one workbook. A contact on the CLIENT'S OWN
    account is theirs; our colleague, the underwriter at Zurich and a
    freeform third party are all ours — and so is a task nobody has claimed,
    because unassigned work is ours until someone says otherwise."""
    from bookkit.services.export_open_items import write

    client = book["client"]
    theirs = tasks.create(conn, "Return signed TRIA form", org_id=client.id)
    ours = tasks.create(conn, "Draft the renewal strategy", org_id=client.id)
    market = tasks.create(conn, "Chase Zurich on the loss run", org_id=client.id)
    third = tasks.create(conn, "Get the appraisal back", org_id=client.id)
    tasks.create(conn, "Nobody has claimed this", org_id=client.id)

    assignees.set_on_task(conn, theirs.id, "Rae Okafor — Atomic Industries",
                          org_id=client.id)
    assignees.set_on_task(conn, ours.id, "Dana Reyes — our team", org_id=client.id)
    assignees.set_on_task(conn, market.id, "Jo Chen — Zurich", org_id=client.id)
    assignees.set_on_task(conn, third.id, "Marisa at Lockton", org_id=client.id)

    owners = _owner_column(write(conn, client.id, tmp_path / "c.xlsx", TODAY))
    assert owners == {
        "Return signed TRIA form": "You",
        "Draft the renewal strategy": "Us",
        "Chase Zurich on the loss run": "Us",
        "Get the appraisal back": "Us",
        "Nobody has claimed this": "Us",
    }, owners


def test_the_same_name_on_both_sides_does_not_flip_the_column(
    conn, book, tmp_path
) -> None:
    """The collision, end to end and on the client's own copy.

    Two people called Dana Reyes — one ours, one theirs. The two tasks are
    told apart by the QUALIFIED label, so they land on opposite sides even
    though the string a name-matcher would compare is identical. And a task
    assigned by the bare ambiguous name reads Us: a refusal to guess lands on
    the recoverable side of a document the client reads."""
    from bookkit.services.export_open_items import write

    client = book["client"]
    contacts.create(conn, client.id, first_name="Dana", last_name="Reyes")

    theirs = tasks.create(conn, "Countersign the schedule", org_id=client.id)
    ours = tasks.create(conn, "Bind the layer", org_id=client.id)
    guessed = tasks.create(conn, "Confirm the SIR", org_id=client.id)
    assignees.set_on_task(conn, theirs.id, "Dana Reyes — Atomic Industries",
                          org_id=client.id)
    assignees.set_on_task(conn, ours.id, "Dana Reyes — our team", org_id=client.id)
    assignees.set_on_task(conn, guessed.id, "Dana Reyes", org_id=client.id)

    owners = _owner_column(write(conn, client.id, tmp_path / "d.xlsx", TODAY))
    assert owners == {
        "Countersign the schedule": "You",
        "Bind the layer": "Us",
        "Confirm the SIR": "Us",
    }, owners


def test_a_contact_at_ANOTHER_client_is_not_this_clients_problem(
    conn, book, tmp_path
) -> None:
    """The comparison is against the account being EXPORTED, not against
    "is a client contact at all". Rae works at Atomic; on Borealis's workbook
    she is not the reader's own person and the row is not theirs."""
    from bookkit.services.export_open_items import write

    other = orgs.create(conn, name="Borealis Foods", kind="client")
    t = tasks.create(conn, "Confirm the values", org_id=other.id)
    assignees.set_on_task(
        conn, t.id, "Rae Okafor — Atomic Industries", org_id=book["client"].id
    )
    assert tasks.get(conn, t.id).assignee_id == book["theirs"].id  # it DID resolve

    owners = _owner_column(write(conn, other.id, tmp_path / "e.xlsx", TODAY))
    assert owners == {"Confirm the values": "Us"}, owners


def test_a_submission_row_carries_an_owner_too(conn, book, tmp_path) -> None:
    """Every row on the sheet answers the question, or the column has holes
    the client reads as "we don't know". A submission out at market is ours
    by nature — nothing on it can carry an assignee."""
    from bookkit.repo import opportunities, submissions
    from bookkit.services.export_open_items import write

    client = book["client"]
    opp = opportunities.create(conn, client.id, "Casualty renewal")
    submissions.create(conn, book["market"].id, "2026-07-01", opportunity_id=opp.id)
    owners = _owner_column(write(conn, client.id, tmp_path / "g.xlsx", TODAY))
    assert owners == {"Submission to Zurich": "Us"}, owners


def test_the_assignees_NAME_never_reaches_the_client(conn, book, tmp_path) -> None:
    """The internal fact stays internal. The CFO asked whose an item is, not
    who at our firm is handling it — and a client-facing sheet naming our
    staff, or naming an underwriter we are negotiating with, is a disclosure
    nobody signed off."""
    from openpyxl import load_workbook

    from bookkit.services.export_open_items import write

    client = book["client"]
    t = tasks.create(conn, "Draft the renewal strategy", org_id=client.id)
    assignees.set_on_task(conn, t.id, "Dana Reyes — our team", org_id=client.id)
    m = tasks.create(conn, "Chase the loss run", org_id=client.id)
    assignees.set_on_task(conn, m.id, "Jo Chen — Zurich", org_id=client.id)

    wb = load_workbook(write(conn, client.id, tmp_path / "h.xlsx", TODAY))
    values = [
        str(c.value)
        for name in wb.sheetnames
        for row in wb[name].iter_rows()
        for c in row
        if c.value is not None
    ]
    assert not any("Dana Reyes" in v for v in values), "our colleague is named"
    assert not any("Jo Chen" in v for v in values), "the underwriter is named"


# --- the add / edit form ----------------------------------------------------


def _field(spec, key: str):
    return {f.key: f for f in spec.fields}[key]


def test_the_task_form_offers_the_assignee_on_both_halves(conn, book) -> None:
    """CLAUDE.md: a vocabulary field completes from existing records, wired on
    BOTH halves — the autocomplete dropdown and the ghost text. Both read
    `Field.suggestions`, so this is the one thing that has to be there; a
    field with an empty tuple renders as a bare box that teaches nobody the
    picker exists."""
    from bookkit.forms.entities import task_form

    spec = task_form(conn=conn, default_org_id=book["client"].id)
    assignee = _field(spec, "assignee")
    assert assignee.suggestions == (
        "Dana Reyes — our team",
        "Jo Chen — Zurich",
        "Rae Okafor — Atomic Industries",
    ), assignee.suggestions


def test_the_form_saves_a_picked_assignee_as_an_identity(conn, book) -> None:
    """The create path. `assignee` is not a Task column, so a form that
    handed it straight to the repo would raise — and one that dropped it
    would lose the answer with nothing saying so."""
    from bookkit.forms.entities import apply_task

    task = apply_task(
        conn,
        {"title": "Return signed TRIA form",
         "assignee": "Rae Okafor — Atomic Industries"},
        org_id=book["client"].id,
    )
    assert task.assignee_kind is AssigneeKind.CONTACT
    assert task.assignee_id == book["theirs"].id
    assert task.assignee_name is None


def test_the_form_reopens_on_the_person_it_saved(conn, book) -> None:
    """Edit is the same form. It must pre-fill the assignee — an edit form
    that opens blank on a filled field reads as "nobody is on this", and
    saving it would then be true."""
    from bookkit.forms.entities import apply_task, task_form

    task = apply_task(
        conn,
        {"title": "Return signed TRIA form",
         "assignee": "Rae Okafor — Atomic Industries"},
        org_id=book["client"].id,
    )
    spec = task_form(task, conn=conn)
    assert spec.initial["assignee"] == "Rae Okafor — Atomic Industries"


def test_editing_a_task_offers_ITS_accounts_people_not_the_cursors(
    conn, book
) -> None:
    """A task edited from a list that spans accounts (the navigator's
    attention pane) must offer the contacts of the account the TASK belongs
    to. Offering the last-looked-at account's people would put another
    client's risk manager in the picker."""
    from bookkit.forms.entities import apply_task, task_form

    other = orgs.create(conn, name="Borealis Foods", kind="client")
    contacts.create(conn, other.id, first_name="Sam", last_name="Ruiz")
    task = apply_task(conn, {"title": "Confirm the values"}, org_id=other.id)

    spec = task_form(task, conn=conn, default_org_id=book["client"].id)
    labels = _field(spec, "assignee").suggestions
    assert "Sam Ruiz — Borealis Foods" in labels
    assert "Rae Okafor — Atomic Industries" not in labels


def test_the_form_can_clear_an_assignee(conn, book) -> None:
    """Blanking the field has to mean "nobody", not "leave it alone" —
    forms.spec.dropped() strips a None so an optional blank does not
    overwrite, which is right for every other field and wrong for this one."""
    from bookkit.forms.entities import apply_task

    task = apply_task(
        conn, {"title": "Chase the binder", "assignee": "Dana Reyes — our team"},
        org_id=book["client"].id,
    )
    assert task.assignee_id == book["ours"].id
    cleared = apply_task(
        conn, {"title": "Chase the binder", "assignee": None},
        org_id=book["client"].id, existing=task,
    )
    assert (
        cleared.assignee_kind, cleared.assignee_id, cleared.assignee_name
    ) == (None, None, None)


def test_an_unassigned_cell_prints_the_dash_and_not_a_blank() -> None:
    """The same fix the renewals pane's `_cover` needed. Unassigned is a real
    and common state — it is the DEFAULT — so the cell that says so has to
    look like an empty field on purpose rather than like a rendering fault.
    The dash is free."""
    from bookkit.tui import theme

    assert str(theme.assignee_text("")) == "—"
    assert str(theme.assignee_text("Dana Reyes")) == "Dana Reyes"


# --- assigning through MCP (2026-08-19) ---------------------------------------
#
# The gap that caused a wrong record to be written in Grant's real book: a task
# could be assigned in the TUI and on the web and NOT through MCP, so an
# assistant asked to file an assigned open item could not do it — and filed an
# information request instead. A capability gap the model routes around is
# worse than one it reports, because what it does instead lands in the book.


def test_a_task_can_be_created_with_an_assignee(conn) -> None:
    from bookkit import mcpserver
    from bookkit.repo import assignees, orgs, tasks

    org = orgs.create(conn, kind="client", name="Atomic Industries")
    mcpserver._member_create(conn, "Dana Okafor")

    out = mcpserver._task_create(
        conn, "chase the loss runs", client="Atomic Industries", assignee="Dana Okafor"
    )

    task = tasks.get(conn, out["task_ref"])
    assert assignees.name_of(conn, task) == "Dana Okafor"
    assert task.assignee_kind is not None, "resolved to a person, not free text"
    assert out["assignee"] == "Dana Okafor"
    del org


def test_a_task_is_created_unassigned_when_nobody_is_named(conn) -> None:
    """THE BACKSTOP. Assignment is a bonus, never a precondition — an open item
    nobody has picked up yet is the normal state of a task, and refusing to
    file one because no name was given is what sent an assistant looking for
    another record type to write instead."""
    from bookkit import mcpserver
    from bookkit.repo import orgs, tasks

    orgs.create(conn, kind="client", name="Atomic Industries")

    out = mcpserver._task_create(conn, "chase the loss runs", client="Atomic Industries")

    task = tasks.get(conn, out["task_ref"])
    assert task.assignee_kind is None
    assert task.assignee_id is None
    assert task.assignee_name is None


def test_an_unknown_name_is_kept_as_typed_rather_than_refused(conn) -> None:
    """A name the book does not know yet is a note, not an error. Refusing it
    would put the task's existence at the mercy of whether somebody has been
    added to the team table yet — which is precisely the failure this fixes."""
    from bookkit import mcpserver
    from bookkit.repo import orgs, tasks

    orgs.create(conn, kind="client", name="Atomic Industries")

    out = mcpserver._task_create(
        conn, "chase the loss runs", client="Atomic Industries",
        assignee="Somebody Not In The Book",
    )

    task = tasks.get(conn, out["task_ref"])
    assert task.assignee_name == "Somebody Not In The Book"
    assert task.assignee_kind is None


def test_an_existing_task_can_be_assigned_and_unassigned(conn) -> None:
    from bookkit import mcpserver
    from bookkit.repo import assignees, orgs, tasks

    orgs.create(conn, kind="client", name="Atomic Industries")
    mcpserver._member_create(conn, "Dana Okafor")
    ref = mcpserver._task_create(
        conn, "chase the loss runs", client="Atomic Industries"
    )["task_ref"]

    mcpserver._task_assign(conn, ref, "Dana Okafor")
    assert assignees.name_of(conn, tasks.get(conn, ref)) == "Dana Okafor"

    mcpserver._task_assign(conn, ref, None)
    cleared = tasks.get(conn, ref)
    assert cleared.assignee_kind is None
    assert cleared.assignee_id is None
    assert cleared.assignee_name is None


def test_assigning_is_one_revertible_batch(conn) -> None:
    from bookkit import db, mcpserver
    from bookkit.repo import assignees, orgs, tasks
    from bookkit.services import batches as batches_svc

    orgs.create(conn, kind="client", name="Atomic Industries")
    mcpserver._member_create(conn, "Dana Okafor")
    ref = mcpserver._task_create(
        conn, "chase the loss runs", client="Atomic Industries"
    )["task_ref"]

    out = mcpserver._task_assign(conn, ref, "Dana Okafor")
    batches_svc.revert(conn, out["batch"], now=db.utc_now())

    assert assignees.name_of(conn, tasks.get(conn, ref)) != "Dana Okafor"
