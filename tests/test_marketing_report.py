"""The marketing report, built against the worked example the design settled
on — so the numbers in the artifact and the numbers this ships are the same
numbers.

General Liability, expiring $412,000 on $41.0M of sales; this year $48.5M.
Six markets, one of them reached through a wholesaler, covering every status.
"""

from __future__ import annotations

import sqlite3
from datetime import date

import pytest

from bookkit.repo import marketing, orgs, placements, submissions
from bookkit.services import marketing_report

TODAY = date(2027, 7, 28)

# Cents, and rate in micros (rate x 1,000,000). The expiring rate is the
# ROUNDED figure a broker actually writes down (412,000 / 41,000 = 10.0488),
# which is why the bridge below reconciles to within a couple of cents rather
# than exactly: it is a decomposition of a rounded input, not an identity.
EXPIRING_PREMIUM = 41_200_000
EXPIRING_EXPOSURE = 4_100_000_000
EXPIRING_RATE = 10_048_780
EXPOSURE = 4_850_000_000


def _book(conn: sqlite3.Connection):
    client = orgs.create(conn, kind="client", name="Legibility Inc", status="active")
    placement = placements.create(
        conn,
        org_id=client.id,
        program_name="2027 casualty",
        period_from="2027-09-01",
        period_to="2028-09-01",
    )
    marketing.set_placement_line(
        conn,
        placement.id,
        "general-liability",
        expiring_premium=EXPIRING_PREMIUM,
        expiring_exposure=EXPIRING_EXPOSURE,
        expiring_rate_micros=EXPIRING_RATE,
        expiring_basis="gross_sales",
        expected_exposure=EXPOSURE,
        rating_basis="gross_sales",
        rate_per=1000,
        limit_sought=100_000_000,
    )
    return client, placement


def _market(conn, name, best=None):
    org = orgs.create(conn, kind="market", name=name, status="active")
    if best:
        conn.execute(
            "INSERT INTO market_profile (org_id, am_best_rating, market_type)"
            " VALUES (?, ?, 'carrier')",
            (org.id, best),
        )
    return org


def _approach(conn, placement_id, market, **fields):
    sub = submissions.create(
        conn, market_org_id=market.id, sent_on="2027-07-07", placement_id=placement_id
    )
    return marketing.create_response(
        conn, sub.id, "general-liability", market_org_id=market.id, **fields
    )


def _full_book(conn):
    client, placement = _book(conn)
    travelers = _market(conn, "Travelers", "A++ XV")
    cna = _market(conn, "CNA", "A XV")
    chubb = _market(conn, "Chubb", "A++ XV")
    zurich = _market(conn, "Zurich", "A+ XV")
    liberty = _market(conn, "Liberty Mutual", "A XV")
    aig = _market(conn, "AIG", "A XV")
    rt = _market(conn, "RT Specialty")

    _approach(
        conn, placement.id, travelers, status="quoted", responded_on="2027-07-21",
        rate_micros=8_100_000, premium=39_285_000, tria_premium=785_000,
        policy_fees=390_000, surplus_lines_tax=0,
    )
    sub_cna = submissions.create(
        conn, market_org_id=cna.id, sent_on="2027-07-07", placement_id=placement.id
    )
    marketing.create_response(
        conn, sub_cna.id, "general-liability", market_org_id=cna.id, via_org_id=rt.id,
        status="quoted", responded_on="2027-07-24", rate_micros=7_950_000,
        premium=38_557_500, tria_premium=771_000, policy_fees=40_000,
        surplus_lines_tax=1_928_000,
    )
    _approach(
        conn, placement.id, chubb, status="indicated", responded_on="2027-07-18",
        rate_micros=7_900_000, premium=38_315_000,
    )
    _approach(conn, placement.id, zurich, status="pending")
    _approach(
        conn, placement.id, liberty, status="declined", responded_on="2027-07-15",
        decline_reason="underwriter hates the loss runs, off the record",
        decline_reason_public="loss_history",
    )
    _approach(conn, placement.id, aig, status="non_response")
    return placement


def test_the_block_reads_live_options_first(conn) -> None:
    placement = _full_book(conn)
    report = marketing_report.compose(conn, placement.id, TODAY)
    block = report.blocks[0]
    assert block.line_name == "General Liability"
    assert [row.status for row in block.rows] == [
        "Quoted", "Quoted", "Indicated", "Pending", "Declined", "Non-response",
    ]
    # Within Quoted, cheapest first.
    assert [row.market for row in block.rows[:2]] == ["CNA", "Travelers"]


def test_a_wholesaler_is_named_beside_the_paper(conn) -> None:
    placement = _full_book(conn)
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    cna = next(row for row in block.rows if row.market == "CNA")
    assert cna.market_cell == "CNA (via RT Specialty)"


def test_the_rate_change_is_what_the_broker_achieved(conn) -> None:
    """Premium fell 4.6%; the RATE fell 19.4% while the client's sales grew
    18.3%. A premium-only report tells the client the wrong story."""
    placement = _full_book(conn)
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    travelers = next(row for row in block.rows if row.market == "Travelers")
    assert travelers.rate_move.pct is not None
    assert round(travelers.rate_move.pct, 1) == -19.4
    assert travelers.rate_move.cell == "-19.4%"
    assert block.exposure_move.pct is not None
    assert round(block.exposure_move.pct, 1) == 18.3


def test_the_bridge_splits_rate_from_growth_and_reconciles(conn) -> None:
    placement = _full_book(conn)
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    bridge = block.bridge
    assert bridge is not None
    assert bridge.expiring_premium == EXPIRING_PREMIUM
    # The LEADING quote is CNA, cheapest premium among the live options.
    assert bridge.market.startswith("CNA")
    assert round(bridge.rate_effect / 100) == -86_050       # dollars
    assert round(bridge.exposure_effect / 100) == 59_625
    walked = bridge.expiring_premium + bridge.rate_effect + bridge.exposure_effect
    assert abs(walked - bridge.quoted_premium) <= 100  # within a dollar


def test_a_total_cost_appears_only_where_every_component_is_known(conn) -> None:
    """THE CHEAPER PREMIUM IS THE DEARER PLACEMENT. CNA quotes $385,575
    against Travelers' $392,850 — and lands $8,365 higher once its surplus
    lines tax is on, because it is E&S paper through a wholesaler. A premium
    column alone recommends the wrong market. Chubb has given an indication
    and nothing else, so its total is BLANK, not zero."""
    placement = _full_book(conn)
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    rows = {row.market: row for row in block.rows}
    cna, travelers, chubb = rows["CNA"], rows["Travelers"], rows["Chubb"]

    assert cna.premium == 38_557_500 and travelers.premium == 39_285_000
    assert cna.premium < travelers.premium          # cheaper on the quote
    assert cna.total_cost == 41_296_500             # $412,965
    assert travelers.total_cost == 40_460_000       # $404,600
    assert cna.total_cost > travelers.total_cost    # dearer on the money paid
    assert chubb.total_cost is None


def test_a_rate_change_across_two_bases_is_refused_in_words(conn) -> None:
    """Last year on payroll, this year on sales. A percentage here would be
    the same lie as ranking two carriers rated on different bases."""
    _, placement = _book(conn)
    marketing.set_placement_line(
        conn, placement.id, "general-liability", expiring_basis="payroll"
    )
    market = _market(conn, "Travelers")
    _approach(conn, placement.id, market, status="quoted", rate_micros=8_100_000,
              premium=39_285_000)
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    assert block.rows[0].rate_move.pct is None
    assert block.rows[0].rate_move.cell == "basis changed"


def test_a_missing_expiring_rate_leaves_the_comparison_blank(conn) -> None:
    """Most of the book starts here. Assuming exposure was flat would put a
    premium change in the rate column wearing a rate's clothes."""
    client = orgs.create(conn, kind="client", name="No History Co", status="active")
    placement = placements.create(
        conn, org_id=client.id, program_name="2027", period_from="2027-01-01",
        period_to="2028-01-01",
    )
    marketing.set_placement_line(
        conn, placement.id, "general-liability",
        # THE DENOMINATOR, because a rate cannot be stored without one any
        # more (repo.marketing._stamp_rate_per). It is not what this test is
        # about; it is what makes the row it is about writable at all.
        rate_per=100, expiring_premium=41_200_000,
    )
    market = _market(conn, "Travelers")
    _approach(conn, placement.id, market, status="quoted", rate_micros=8_100_000,
              premium=39_285_000)
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    assert block.rows[0].rate_move.cell == "no expiring rate recorded"
    assert block.bridge is None


def test_the_client_sheet_withholds_what_the_internal_one_shows(conn) -> None:
    placement = _full_book(conn)
    client_report = marketing_report.compose(conn, placement.id, TODAY)
    internal = marketing_report.compose(conn, placement.id, TODAY, audience="internal")

    liberty_client = next(r for r in client_report.blocks[0].rows if r.market == "Liberty Mutual")
    liberty_internal = next(r for r in internal.blocks[0].rows if r.market == "Liberty Mutual")

    assert liberty_client.public_reason == "Loss history"
    assert liberty_client.internal_reason == ""
    assert "off the record" in liberty_internal.internal_reason

    client_text = "\n".join(
        "\t".join(cells)
        for section in marketing_report.to_sections(client_report)
        for cells in section.rows
    )
    assert "off the record" not in client_text


def test_a_line_with_no_approaches_says_so_in_words(conn) -> None:
    """An empty table is ambiguous: a client cannot tell a line nobody has
    gone to market on from a rendering bug."""
    _, placement = _book(conn)
    marketing.set_placement_line(conn, placement.id, "auto")
    report = marketing_report.compose(conn, placement.id, TODAY)
    sections = marketing_report.to_sections(report)
    auto = next(s for s in sections if s.label.startswith("Commercial Auto"))
    assert "No markets approached" in auto.rows[0][0]


def test_the_section_label_carries_the_header_facts(conn) -> None:
    placement = _full_book(conn)
    report = marketing_report.compose(conn, placement.id, TODAY)
    label = marketing_report.to_sections(report)[0].label
    assert "General Liability" in label
    # THE SEND DATE IS NOT IN THE HEADING ANY MORE. It collapsed into the
    # label only when every package on the line went out the same day and
    # printed NOWHERE otherwise, so the client's workbook lost it on any line
    # marketed over more than one day (D6, 2026-08-26). It is a `Sent` column
    # on the row now — the same un-collapsing the grid did — and the row
    # assertion below is what holds it.
    assert "submitted" not in label
    assert "Gross sales" in label
    assert "+18.3%" in label
    assert "expiring $412,000 at 10.05" in label

    headers = [h for h, _, _ in marketing_report.columns(marketing_report.CLIENT)]
    section = marketing_report.to_sections(report)[0]
    assert section.rows[0][headers.index("Sent")] == "7 Jul"


# --- through a REAL workbook ----------------------------------------------


def _sheet_text(path) -> str:
    """Everything on sheet 1 of a real .xlsx, as one string.

    Through the FILE, not the row objects: the composition can be right while
    the writer's column spec puts a value under the wrong header, which is
    the class of bug the Owner column nearly shipped."""
    from openpyxl import load_workbook  # test-only import; src never imports it

    sheet = load_workbook(path).active
    return "\n".join(
        "\t".join("" if cell.value is None else str(cell.value) for cell in row)
        for row in sheet.iter_rows()
    )


def test_the_workbook_carries_the_block_and_its_numbers(conn, tmp_path) -> None:
    placement = _full_book(conn)
    out = marketing_report.write(
        conn, placement.id, tmp_path / "marketing.xlsx", TODAY
    )
    assert out.exists()
    text = _sheet_text(out)

    # the block header, and the rate story it exists to tell
    assert "General Liability" in text
    assert "Gross sales" in text
    assert "+18.3%" in text
    # the markets, live options first, the wholesaler named beside the paper
    assert "CNA (via RT Specialty)" in text
    assert "Travelers" in text
    # the money, and the E&S inversion
    # format_cents keeps cents only when present, so whole dollars have no .00
    assert "$385,575" in text and "$392,850" in text
    assert "$412,965" in text and "$404,600" in text
    # the bridge
    assert "Rate effect" in text and "Exposure effect" in text
    # column headers come from the renderer, once, not once per block
    assert text.count("Est. premium") == 1


def test_the_client_workbook_never_carries_the_underwriter_s_words(
    conn, tmp_path
) -> None:
    """The one failure this whole two-field design exists to prevent."""
    placement = _full_book(conn)
    client_file = marketing_report.write(
        conn, placement.id, tmp_path / "client.xlsx", TODAY
    )
    internal_file = marketing_report.write(
        conn, placement.id, tmp_path / "internal.xlsx", TODAY, audience="internal"
    )
    assert "off the record" not in _sheet_text(client_file)
    assert "Loss history" in _sheet_text(client_file)
    assert "off the record" in _sheet_text(internal_file)


# --- a basis that is not money --------------------------------------------


def _fleet(conn):
    """A trucking fleet: 350 power units, rated per unit. The whole worked
    example above is `gross_sales`, where cents happen to be right — which is
    exactly why the count case shipped broken."""
    from bookkit.repo import placements as placements_repo

    client = orgs.create(conn, kind="client", name="Haulage Co", status="active")
    placement = placements_repo.create(
        conn, org_id=client.id, program_name="2027 fleet",
        period_from="2027-01-01", period_to="2028-01-01",
    )
    marketing.set_placement_line(
        conn, placement.id, "auto",
        rating_basis="power_units", rate_per=1,
        expected_exposure=350, expiring_exposure=300,
        # THE EXPIRING SIDE'S BASIS TOO. It was absent, and every comparison
        # this fixture reaches was therefore refused for an axis nobody had
        # stated (silence is not agreement) — which is right, and made the
        # fixture unable to exercise the one thing the count case is here for.
        expiring_basis="power_units",
        # THE PREMIUMS RECONCILE TO THE RATE AND THE COUNT: 300 units at
        # 175.00 per unit is $52,500. They were stated a factor of TEN apart
        # from the rate beside them, and nothing noticed, because the premium
        # bridge could not be computed on a count basis at all — the one check
        # that would have objected (D7).
        expiring_premium=5_250_000, expiring_rate_micros=175_000_000,
    )
    market = _market(conn, "Progressive", "A+ XV")
    sub = submissions.create(
        conn, market_org_id=market.id, sent_on="2027-07-07", placement_id=placement.id
    )
    marketing.create_response(
        conn, sub.id, "auto", market_org_id=market.id, status="quoted",
        # 350 units at 180.00 per unit is $63,000.
        responded_on="2027-07-20", rate_micros=180_000_000, premium=6_300_000,
        # The ORDINARY override: this carrier counted the fleet itself and got
        # the same number, and did NOT restate the basis.
        exposure_amount=350,
    )
    return placement


def test_the_premium_bridge_walks_on_a_count_basis_too(conn) -> None:
    """THE BRIDGE WAS DEAD ON EVERY NON-MONETARY BASIS, silently.

    `_premium_from` multiplied a rate by an exposure and returned the product
    as CENTS. That is right where the exposure is money — cents times a
    dollar denominator cancels the hundred — and a hundredfold LOW where the
    exposure is a COUNT: 320 power units at 1287.50 per unit came out $4,120
    for $412,000. `_reconciles` then dropped the walk, every time, so nothing
    wrong was ever printed and the feature simply never appeared for a
    trucking fleet (D7, found 2026-08-26). Failing safe is not working.

    The figures reconcile to the cent on purpose: 300 units at 175.00 is
    $52,500 and 350 at 180.00 is $63,000, so a bridge that does not appear
    here cannot be excused as rounding.
    """
    placement = _fleet(conn)
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]

    bridge = block.bridge
    assert bridge is not None, "the bridge is dropped on a count basis"
    assert bridge.expiring_premium == 5_250_000
    # (180.00 - 175.00) x 300 units = $1,500
    assert bridge.rate_effect == 150_000
    # 180.00 x (350 - 300) units = $9,000
    assert bridge.exposure_effect == 900_000
    assert bridge.quoted_premium == 6_300_000
    assert (
        bridge.expiring_premium + bridge.rate_effect + bridge.exposure_effect
        == bridge.quoted_premium
    )


def test_a_count_is_never_rendered_as_money(conn) -> None:
    """350 power units printed to a client as "$3.50" — a hundredfold
    mis-render on the client sheet AND the client workbook, found 2026-08-25.

    `ReportBlock` carried the basis LABEL and not the basis KEY, so
    `_block_label` had nothing to pass and `_fmt_exposure` fell back to
    format_cents; the row cell hit the same thing whenever a market overrode
    the exposure without restating the basis. models.RatingBasis.monetary is
    the ONE place that decides cents-or-count, and both call sites had thrown
    the key away before asking it."""
    placement = _fleet(conn)
    report = marketing_report.compose(conn, placement.id, TODAY)
    block = report.blocks[0]

    assert block.basis_key == "power_units"
    label = marketing_report.to_sections(report)[0].label
    assert "350 power units" in label
    assert "$3.50" not in label and "$350" not in label

    # and the same figure in the row cell a market overrode
    assert block.rows[0].exposure_override == "350 power units"


def test_the_basis_column_prints_only_what_a_market_overrode(conn) -> None:
    """The row falls back to the line's basis so it can RENDER the exposure
    correctly — but the Basis COLUMN must stay blank unless this market stated
    its own, or every row repeats the header and "blank means as above" dies."""
    placement = _fleet(conn)
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    assert block.rows[0].basis_override == ""


def test_a_count_never_reaches_the_workbook_as_money(conn, tmp_path) -> None:
    """The composer and the writer share `to_sections`, so this is the same
    defect reaching the file a client is actually sent."""
    placement = _fleet(conn)
    out = marketing_report.write(conn, placement.id, tmp_path / "fleet.xlsx", TODAY)
    text = _sheet_text(out)
    assert "350 power units" in text
    assert "$3.50" not in text


# --- what may NOT be printed ----------------------------------------------


def test_an_exposure_change_across_two_bases_is_refused_in_words(conn) -> None:
    """THE THIRD COMPARISON, and the one that shipped without the guard.

    Gross sales last term, power units this. `_rate_move` refuses this exact
    pair in words and `_bridge` returns None for it — while the exposure
    comparison divided 350 power units by $41,000,000 and printed "exposure
    down 100%" on the sheet a CLIENT is sent.
    """
    _, placement = _book(conn)
    # Basis and figure in ONE call, which is the ordinary correction and the
    # only way past `repo.marketing._basis_guard`.
    marketing.set_placement_line(
        conn, placement.id, "general-liability",
        rating_basis="power_units", expected_exposure=350,
    )
    report = marketing_report.compose(conn, placement.id, TODAY)
    block = report.blocks[0]

    assert block.exposure_move.pct is None
    assert block.exposure_move.cell == "basis changed"
    label = marketing_report.to_sections(report)[0].label
    assert "basis changed" in label
    assert "-100.0%" not in label, "a comparison across two bases reached the sheet"


def test_an_exposure_change_within_one_basis_is_still_printed(conn) -> None:
    """The other half: the refusal must not swallow the ordinary case. Same
    basis both terms, $41.0M to $48.5M."""
    _, placement = _book(conn)
    report = marketing_report.compose(conn, placement.id, TODAY)
    assert report.blocks[0].exposure_move.cell == "+18.3%"


def test_a_bridge_that_does_not_add_up_is_not_printed(conn) -> None:
    """`Bridge` PROMISES it reconciles — expiring + rate + exposure = quoted —
    and nothing checked that it still did. Four lines that do not add up are
    worse than no bridge: the client reads them as the explanation of the
    number above them.

    A premium typed with a digit adrift is the ordinary way to reach it, and a
    denominator or basis relabelled under a stored rate is the dangerous one.
    """
    _, placement = _book(conn)
    market = _market(conn, "Travelers")
    response = _approach(
        conn, placement.id, market, status="quoted", rate_micros=8_100_000,
        premium=39_285_000,
    )
    assert marketing_report.compose(conn, placement.id, TODAY).blocks[0].bridge

    marketing.edit_response(conn, response.id, {"premium": 50_000_000})

    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    assert block.bridge is None, "a bridge that misses by $107,000 was printed"


def test_a_date_outside_the_placement_s_own_window_prints_its_year(
    conn, tmp_path
) -> None:
    """"12 Aug" was every date this report printed, so 2001, 2027 and 2099 all
    rendered identically — on the grid AND in the client's workbook. The year
    prints where it is news: outside the window this placement's marketing can
    honestly fall in.
    """
    _, placement = _book(conn)  # 2027-09-01 to 2028-09-01
    ordinary = _approach(
        conn, placement.id, _market(conn, "Travelers"),
        status="quoted", responded_on="2027-07-21", premium=39_285_000,
    )
    _approach(
        conn, placement.id, _market(conn, "Zurich"),
        status="quoted", responded_on="2099-08-12", premium=41_000_000,
    )
    report = marketing_report.compose(conn, placement.id, TODAY)

    assert report.window.holds(ordinary.responded_on)
    text = _sheet_text(
        marketing_report.write(conn, placement.id, tmp_path / "dates.xlsx", TODAY)
    )
    assert "21 Jul" in text and "21 Jul 2027" not in text
    assert "12 Aug 2099" in text, "a mistyped year is invisible on the client sheet"


def test_a_reply_cannot_predate_the_submission_it_answers(conn) -> None:
    """CONSISTENCY IS THE THIN CATEGORY. A market cannot answer a package it
    has not been sent, and a mistyped year is the ordinary way to record one
    that did — the guard is in repo/ so every surface inherits it."""
    _, placement = _book(conn)
    sub = submissions.create(
        conn, market_org_id=_market(conn, "Travelers").id,
        sent_on="2027-07-07", placement_id=placement.id,
    )
    with pytest.raises(ValueError, match="cannot answer"):
        marketing.create_response(
            conn, sub.id, "general-liability",
            market_org_id=_market(conn, "Travelers").id,
            status="declined", responded_on="2026-07-08",
        )
    response = marketing.create_response(
        conn, sub.id, "general-liability",
        market_org_id=_market(conn, "Travelers").id, status="pending",
    )
    with pytest.raises(ValueError, match="cannot answer"):
        marketing.edit_response(conn, response.id, {"responded_on": "2027-07-06"})
    # The day the submission went out is an answer, not a contradiction.
    marketing.edit_response(conn, response.id, {"responded_on": "2027-07-07"})
    assert marketing.get_response(conn, response.id).responded_on == "2027-07-07"


def test_a_retired_line_of_coverage_keeps_its_block(conn, tmp_path) -> None:
    """A SOFT-DELETED VOCABULARY ROW IS NOT A LICENCE TO DELETE HISTORY.
    Retiring a line of coverage touches neither the responses nor the
    expectations recorded against it, but the composer read the LIVING
    vocabulary, found no name for the id and dropped the whole block — a bound
    quote at $392,850 gone from the Program tab AND from the client's own
    workbook, which then showed one line where three had been.
    """
    from bookkit.repo import base

    _, placement = _book(conn)
    _approach(
        conn, placement.id, _market(conn, "Chubb", "A++ XV"),
        status="bound", responded_on="2027-07-21", premium=39_285_000,
    )
    base.soft_delete(conn, "line_of_coverage", "general-liability")

    blocks = marketing_report.compose(conn, placement.id, TODAY).blocks
    assert [b.line_name for b in blocks] == ["General Liability"]
    assert blocks[0].line_retired
    assert [r.market for r in blocks[0].rows] == ["Chubb"]
    assert blocks[0].rows[0].premium == 39_285_000

    text = _sheet_text(
        marketing_report.write(conn, placement.id, tmp_path / "retired.xlsx", TODAY)
    )
    assert "Chubb" in text and "General Liability" in text
    # THE CLIENT SHEET DOES NOT SAY IT. "Property (retired)" on a sheet a
    # client reads is a sentence about their COVER, and that would be a lie —
    # the retirement is a fact about this book's own vocabulary. What the
    # client sheet needed was the rows, which is what it had lost.
    assert "retired" not in text.lower()


def test_marketing_counts_as_a_reference_to_a_line_of_coverage(conn) -> None:
    """SURFACE, DON'T GUESS — `repo/lines.usage` exists so a confirm can say
    what points at a line BEFORE anything happens to it, and it was blind to
    both marketing tables: all zeros while a bound quote hung off the line
    about to be merged away. A merge still does not MOVE them (a quote is
    filed under the cover it was asked about), so what it returns is what it
    moved, and the block it leaves behind is visible rather than deleted.
    """
    from bookkit.repo import lines

    client, placement = _book(conn)
    duplicate = lines.create(conn, "Gen Liability")
    marketing.set_placement_line(conn, placement.id, duplicate, expiring_premium=41_200_000)
    submission = submissions.create(
        conn, market_org_id=_market(conn, "Chubb").id,
        sent_on="2027-07-07", placement_id=placement.id,
    )
    marketing.create_response(
        conn, submission.id, duplicate,
        market_org_id=orgs.find_by_name(conn, "Chubb").id,
        status="bound", premium=39_285_000,
    )

    counts = lines.usage(conn, duplicate)
    assert counts["market_response"] == 1
    assert counts["placement_line"] == 1

    moved = lines.merge(conn, duplicate, "general-liability")
    assert "market_response" not in moved, "merge claimed to move a quote it left"

    blocks = marketing_report.compose(conn, placement.id, TODAY).blocks
    retired = next(b for b in blocks if b.line_id == duplicate)
    assert retired.line_retired and retired.rows[0].premium == 39_285_000


def test_a_row_can_always_name_the_market_that_answered(conn) -> None:
    """THE COLUMN THAT SAYS WHOSE ANSWER IT IS. Deleting the carrier org left
    `market_cell` empty — a row carrying a premium, a status and an A.M. Best
    rating with nothing identifying it, which reads as a rendering fault.
    Deleting the org does not unmake the quote, so the name it had is what the
    report prints; the floor under that is a sentence, never a blank."""
    from dataclasses import replace

    _, placement = _book(conn)
    chubb = _market(conn, "Chubb", "A++ XV")
    _approach(
        conn, placement.id, chubb, status="quoted", responded_on="2027-07-21",
        premium=39_285_000,
    )
    orgs.delete(conn, chubb.id)

    row = marketing_report.compose(conn, placement.id, TODAY).blocks[0].rows[0]
    assert row.market_cell == "Chubb"
    assert replace(row, market="", via=None).market_cell == "market not on file"


def test_the_agent_can_name_a_retired_line_and_a_deleted_market(conn) -> None:
    """THE THIRD SURFACE IS AN AGENT. `marketing_report`'s `responses` index is
    what names a row to the assistant — it is the only place a `response_id`
    comes from, and `market_responded` writes against it. Read through the
    LIVING lookups it printed a raw slug for a retired line of coverage and
    `null` for a market whose org had been deleted, so the one thing the index
    exists to do — say WHICH row — stopped working on exactly the rows the
    panel had already stopped showing."""
    from bookkit import mcpserver
    from bookkit.repo import base

    _, placement = _book(conn)
    chubb = _market(conn, "Chubb", "A++ XV")
    _approach(
        conn, placement.id, chubb, status="bound", responded_on="2027-07-21",
        premium=39_285_000,
    )
    base.soft_delete(conn, "line_of_coverage", "general-liability")
    orgs.delete(conn, chubb.id)

    index = mcpserver._marketing_report(
        conn, placement.id, audience="internal", as_of=TODAY.isoformat()
    )["responses"]

    assert [r["line"] for r in index] == ["General Liability"]
    assert [r["market"] for r in index] == ["Chubb"]


# Two responses tied on layer and premium, with the SQL order (attach, id)
# deliberately opposite the report's own order (status, premium, reply date).
# The ids are stated rather than minted because a ULID's low bits are random:
# two rows written in the same millisecond come back in either order, and a
# test that only fails half the time is not a test.
_TIED_FIRST = "01" + "A" * 24
_TIED_SECOND = "01" + "B" * 24
# Chubb's 9.00 against the expiring 10.05 walks to this, to two cents.
TIED_PREMIUM = 43_650_000
CHUBB_RATE_EFFECT = -4_299_998
TRAVELERS_RATE_EFFECT = 8_000_002


def test_two_quotes_tied_on_layer_and_premium_bridge_the_right_one(conn) -> None:
    """A ROW IS IDENTIFIED BY ITS ID, never re-found by matching two of its
    values.

    The bridge took its leading row out of the SORTED rows and then went
    looking for it again in the SQL-ordered responses by `layer ==` and
    `premium ==`. Two markets quoting the same band at the same figure — the
    ordinary end of a competitive marketing exercise — satisfy that pair
    twice, and the two orderings then disagree: the bridge printed under
    "Chubb" on the CLIENT's workbook was decomposed from Travelers' rate.
    `_reconciles` cannot catch it, because the walk it checks adds up
    perfectly and simply belongs to somebody else.
    """
    _, placement = _book(conn)
    travelers = _market(conn, "Travelers", "A++ XV")
    chubb = _market(conn, "Chubb", "A++ XV")
    # First in the SQL order, last in the report's: it replied a week later.
    _approach(
        conn, placement.id, travelers, id=_TIED_FIRST, status="quoted",
        responded_on="2027-07-21", rate_micros=12_000_000, premium=TIED_PREMIUM,
    )
    _approach(
        conn, placement.id, chubb, id=_TIED_SECOND, status="quoted",
        responded_on="2027-07-12", rate_micros=9_000_000, premium=TIED_PREMIUM,
    )

    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    assert [row.market for row in block.rows] == ["Chubb", "Travelers"]
    assert block.rows[0].layer == block.rows[1].layer
    assert block.rows[0].premium == block.rows[1].premium

    bridge = block.bridge
    assert bridge is not None
    assert bridge.market.startswith("Chubb")
    assert bridge.rate_effect == CHUBB_RATE_EFFECT, (
        "the bridge named one market and decomposed the other's rate"
    )
    assert bridge.rate_effect != TRAVELERS_RATE_EFFECT

    # The other half, and the reported shape: the leading row's OWN walk does
    # not add up, so there must be no bridge at all — not the tied row's walk
    # printed under the leading row's name.
    marketing.edit_response(conn, block.rows[0].response_id, {"rate_micros": 12_000_000})
    marketing.edit_response(conn, block.rows[1].response_id, {"rate_micros": 9_000_000})
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    assert block.rows[0].market == "Chubb"
    assert block.bridge is None, "a walk belonging to the row below was printed"


def test_a_rate_across_two_denominators_is_refused_in_words(conn) -> None:
    """RATES COMPARE ONLY WITHIN ONE BASIS **AND** ONE DENOMINATOR.

    `_rate_move` checked the basis and never the `rate_per`, and the route
    into it is the very sequence `repo.marketing._rate_per_guard`'s refusal
    names: clear the expiring rate, move the denominator, enter the rate
    again. The response rate stored against per-$1,000 was then divided by an
    expiring rate stated per-$100 and printed as an 88% reduction nobody
    achieved — on the client's workbook (found 2026-08-26).
    """
    _, placement = _book(conn)          # per $1,000, expiring 10.05
    market = _market(conn, "Travelers")
    _approach(
        conn, placement.id, market, status="quoted", responded_on="2027-07-21",
        rate_micros=11_880_000, premium=39_285_000,
    )
    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    assert block.rows[0].rate_move.cell == "+18.2%"

    # The three clicks the guard tells the broker to make.
    marketing.set_placement_line(
        conn, placement.id, "general-liability", expiring_rate_micros=None
    )
    marketing.set_placement_line(conn, placement.id, "general-liability", rate_per=100)
    marketing.set_placement_line(
        conn, placement.id, "general-liability", expiring_rate_micros=100_500_000
    )

    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    assert block.rows[0].rate_move.pct is None
    assert block.rows[0].rate_move.cell == "denominator changed"
    assert "-88.2%" not in marketing_report.to_sections(
        marketing_report.compose(conn, placement.id, TODAY)
    )[0].rows[0], "a comparison across two denominators reached the sheet"
    assert block.bridge is None


def test_a_stored_rate_carries_the_denominator_it_was_typed_against(conn) -> None:
    """The write half of the rule above. Nothing was writing
    `market_response.rate_per`, so every quoted rate inherited the LINE's
    denominator live at read time and moved with it — which is what made the
    comparison above unable to see that the two sides had parted company.
    A rate CLEARED takes its denominator with it: a denominator with no rate
    under it marks nothing."""
    _, placement = _book(conn)          # per $1,000
    market = _market(conn, "Travelers")
    response = _approach(
        conn, placement.id, market, status="quoted", rate_micros=8_100_000,
    )
    assert response.rate_per == 1000

    # Both halves in one act — the ordinary correction `_rate_per_guard` allows.
    marketing.set_placement_line(
        conn, placement.id, "general-liability",
        rate_per=100, expiring_rate_micros=100_500_000,
    )
    fresh = marketing.edit_response(conn, response.id, {"rate_micros": 7_000_000})
    assert fresh.rate_per == 100, "the rate did not carry the denominator it was typed against"

    cleared = marketing.edit_response(conn, response.id, {"rate_micros": None})
    assert cleared.rate_per is None


def test_an_exposure_with_no_basis_is_never_rendered_as_money(conn) -> None:
    """THE HUNDREDFOLD MISTAKE, ONE BRANCH AWAY. `fmt_exposure` fell back to
    `format_cents` whenever the basis was None — the same default that printed
    350 power units to a client as "$3.50", still sitting under the one input
    nobody can currently reach. Both refusals hold today (the exposure cell
    refuses while the basis is unknown, and `_basis_guard` refuses clearing a
    basis out from under a figure), so this branch is unreachable; it is the
    NEXT surface that composes an exposure that this is for. Nothing in the
    integer says whether it is money or a count, so the honest rendering says
    the digits and says the unit is unknown."""
    rendered = marketing_report.fmt_exposure(4_850_000_000, None)
    assert "$" not in rendered, "an unmarked integer was printed as money"
    assert "4,850,000,000" in rendered
    assert "basis not set" in rendered


def test_a_bridge_across_two_denominators_is_not_printed(conn) -> None:
    """`rate_effect` SUBTRACTS the expiring rate from the quoted one, which is
    arithmetic only while both are stated per the same unit — and `_reconciles`
    cannot be the thing that catches it, because the premium is free: a walk
    built from two denominators adds up to whatever it adds up to, and the
    client reads those four lines as the explanation of the figure above them.

    The rate here was typed per $1,000 (and stamped so) and the line now
    states its expiring rate per $100. The walk lands EXACTLY on the quote, so
    the reconcile guard waves it through; the only thing that can refuse it is
    knowing the two rates are not on the same denominator.
    """
    _, placement = _book(conn)              # per $1,000
    market = _market(conn, "Travelers")
    _approach(
        conn, placement.id, market, status="quoted", responded_on="2027-07-21",
        rate_micros=11_880_000, premium=94_697_500,
    )
    # Denominator and expiring rate restated together — the one correction
    # `_rate_per_guard` allows. The RESPONSE rate is left where it was, which
    # is the door the guard deliberately leaves open.
    marketing.set_placement_line(
        conn, placement.id, "general-liability",
        rate_per=100, expiring_rate_micros=1_005_000,
    )

    block = marketing_report.compose(conn, placement.id, TODAY).blocks[0]
    assert block.bridge is None, "a walk built from two denominators reconciled and printed"


def test_the_rate_column_is_wide_enough_for_the_words_it_prints() -> None:
    """A REFUSAL SAYS SOMETHING — and a sentence clipped by the populated cell
    beside it says half of it. The Rate Δ column prints a percentage most of
    the time and one of `REFUSAL_NOTES` the rest of it, and those are the rows
    where the words ARE the answer. Its width is derived from them so a new
    refusal cannot be added without the column growing to hold it."""
    width = next(
        w for header, w, _ in marketing_report.columns(marketing_report.CLIENT)
        if header == "Rate Δ"
    )
    longest = max(len(note) for note in marketing_report.REFUSAL_NOTES)
    assert width >= longest, "the column clips the sentence it exists to print"
