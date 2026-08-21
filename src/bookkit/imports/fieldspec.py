"""One registry per import flow: keys, aliases, examples — the single source
of truth that drives BOTH header mapping (tablemap) and template export, so a
template filled in and re-imported always maps with zero fuzz."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class FieldSpec:
    key: str
    required: bool
    example: str
    aliases: tuple[str, ...] = ()


BOOK_FIELDS: tuple[FieldSpec, ...] = (
    FieldSpec(
        "account", True, "Atomic Industries",
        ("insured", "client", "named insured", "account name", "company"),
    ),
    FieldSpec("industry", False, "Manufacturing", ("sector",)),
    FieldSpec("naics", False, "3345"),
    FieldSpec("domain", False, "atomic.example.com"),
    FieldSpec("website", False, "https://atomic.example.com", ("url",)),
    FieldSpec("owner", False, "grant", ("producer", "account exec")),
    FieldSpec("status", False, "prospect"),
    FieldSpec("contact_name", False, "Rosa Silva", ("contact", "primary contact")),
    FieldSpec("contact_email", False, "rosa@atomic.example.com", ("email",)),
    FieldSpec("contact_phone", False, "(312) 555-0142", ("phone",)),
    FieldSpec("contact_title", False, "Risk Manager", ("title",)),
    FieldSpec("program", False, "Property", ("program name", "line of business", "lob")),
    FieldSpec(
        "inception", False, "2026-10-01",
        ("eff date", "effective", "effective date", "inception date"),
    ),
    FieldSpec("expiry", False, "2027-10-01", ("exp date", "expiration", "expiry date")),
    FieldSpec("premium", False, "250,000", ("annual premium", "total premium")),
    FieldSpec("limit", False, "10M", ("total limit",)),
    FieldSpec("commission", False, "15%", ("comm", "commission %")),
    FieldSpec("placement_status", False, "bound"),
)


# The data sheet the pipeline reads, and the sheet the worked example lives
# on. Keeping them apart is the whole point: an example that sits in the data
# sheet IS data.
DATA_SHEET = "Import"
EXAMPLE_SHEET = "Example (never imported)"
EXAMPLE_BANNER = (
    "EXAMPLE ONLY — this sheet is never imported. "
    f"Fill in the {DATA_SHEET!r} sheet."
)


def write_template(specs: tuple[FieldSpec, ...], path: Path) -> Path:
    """Blank populate-and-reimport workbook: canonical headers (bold, required
    filled), frozen header, and the worked example on its OWN sheet.

    The example used to be appended to the data sheet as row 2. Nothing marked
    it and staging had no filter, so filling rows underneath it and
    re-importing created a real "Atomic Industries" account carrying a real
    BOUND $250,000 placement — a figure nobody typed, off no document. The
    example still has to be visible (it is what shows the expected shape), so
    it moved rather than went away."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = DATA_SHEET
    required_fill = PatternFill("solid", fgColor="F9E6A0")
    for col, spec in enumerate(specs, start=1):
        cell = ws.cell(row=1, column=col, value=spec.key)
        cell.font = Font(bold=True)
        if spec.required:
            cell.fill = required_fill
        ws.column_dimensions[cell.column_letter].width = max(12, len(spec.key) + 2)
    ws.freeze_panes = "A2"

    example = wb.create_sheet(EXAMPLE_SHEET)
    banner = example.cell(row=1, column=1, value=EXAMPLE_BANNER)
    banner.font = Font(bold=True, color="B03A2E")
    for col, spec in enumerate(specs, start=1):
        header = example.cell(row=2, column=col, value=spec.key)
        header.font = Font(bold=True)
        example.cell(row=3, column=col, value=spec.example)
        example.column_dimensions[header.column_letter].width = max(
            12, len(spec.key) + 2
        )
    wb.save(path)
    return path
