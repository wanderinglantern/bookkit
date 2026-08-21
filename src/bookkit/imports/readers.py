"""Files → RawTable. Zero interpretation: original headers, cells as the file
gives them, empty cells absent. Everything downstream works on RawTable, so
adding a source format never touches the pipeline."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from ..sync import file_sha256


@dataclass(frozen=True)
class RawTable:
    source: str  # basename, for provenance and reports
    sha256: str
    headers: list[str]
    rows: list[dict[str, object]]  # keyed by ORIGINAL header; empty cells absent


def read_table(path: Path) -> RawTable:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError(
        f"cannot read {path.name!r}: {suffix or 'no'} suffix — "
        "bookkit imports .xlsx and .csv; export the sheet as one of those"
    )


def _read_csv(path: Path) -> RawTable:
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        headers = [h.strip() for h in reader.fieldnames or []]
        rows: list[dict[str, object]] = []
        for raw in reader:
            row: dict[str, object] = {
                k.strip(): v
                for k, v in raw.items()
                if k is not None and v is not None and str(v).strip() != ""
            }
            if row:
                rows.append(row)
    return RawTable(path.name, file_sha256(path), headers, rows)


def _read_xlsx(path: Path) -> RawTable:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _data_sheet(wb)
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        headers = [str(h or "").strip() for h in header_row or ()]
        rows: list[dict[str, object]] = []
        for raw in rows_iter:
            # read-only sheets yield ragged rows; short rows are fine
            row = {
                k: v
                for k, v in zip(headers, raw, strict=False)
                if k and v is not None and str(v).strip() != ""
            }
            if row:
                rows.append(row)
        return RawTable(path.name, file_sha256(path), [h for h in headers if h], rows)
    finally:
        wb.close()


def _data_sheet(wb: Any) -> Any:
    """The sheet named `Import` when the workbook has one, else the first.

    Tab order decides `worksheets[0]`, and tabs get dragged. The template
    ships a data sheet and a separate worked-example sheet; naming the data
    sheet is what stops a reordered workbook from feeding the EXAMPLE into
    the pipeline as if a human had typed it.
    """
    from .fieldspec import DATA_SHEET

    sheets = wb.worksheets
    for sheet in sheets:
        if str(sheet.title).strip().lower() == DATA_SHEET.lower():
            return sheet
    return sheets[0]
